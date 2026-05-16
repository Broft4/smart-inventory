from __future__ import annotations

import asyncio
import calendar
import hashlib
import httpx
import hmac
import os
import logging
import re
from collections import defaultdict
from dataclasses import dataclass
from time import monotonic
from datetime import date, datetime, timedelta, timezone
from typing import Any
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from fastapi import HTTPException
from sqlalchemy import and_, delete, func, inspect, or_, select, text, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.config import settings
from app.moysklad import DEFAULT_CATEGORY_NAME, DEFAULT_SUBCATEGORY_NAME, ms_client
from app.models import AdminLocationAccess, CategoryAssignment, CheckResult, LocationPoint, PasswordResetRequest, ProductCostOverride, ProductFinancialCache, Report, ReportEmployeeCompletion, ReportEmployeeStart, ReportTargetSnapshot, SelectionCycle, SelectionTarget, SelectionTargetDay, User, VerifyAttemptProgress
logger = logging.getLogger(__name__)


from app.schemas import (
    AdminCycleTargetCategory,
    AdminCycleTargetItem,
    AdminCycleTargetsResponse,
    AdminReport,
    AssignSelectionResponse,
    CategoryModel,
    CategoryResult,
    CreateLocationRequest,
    CreateLocationResponse,
    DeleteResponse,
    DiscrepancyItem,
    CompletedSubcategoryInfo,
    InProgressSubcategoryInfo,
    EmployeeReportSummary,
    InventoryStructureResponse,
    ItemModel,
    LocationListResponse,
    LocationPointModel,
    MeResponse,
    ReopenEmployeeAccessResponse,
    ReportHistoryItem,
    ReportHistoryResponse,
    ResetSelectionCycleResponse,
    SaveCycleTargetsRequest,
    SaveCycleTargetsResponse,
    RoleEnum,
    StartReportResponse,
    StatusEnum,
    StoreListResponse,
    UpdateLocationRequest,
    UpdateLocationResponse,
    StoreOption,
    SubcategoryModel,
    UpdateDiscrepancyCostOverrideRequest,
    UpdateDiscrepancyCostOverrideResponse,
    UpdateDiscrepancyRequest,
    UpdateDiscrepancyResponse,
    UserActionResponse,
    UserCreateRequest,
    UserInfo,
    UserListResponse,
    UserResponse,
    UserUpdateRequest,
    VerifyRequest,
    VerifyResponse,
)



@dataclass(slots=True)
class RuntimeCacheEntry:
    value: Any
    expires_at: float
    inventory_identity: int | None = None


STRIPPED_INVENTORY_CACHE_TTL = max(15, int(settings.ms_inventory_cache_ttl_seconds or 120))
TARGET_LOOKUP_CACHE_TTL = max(60, STRIPPED_INVENTORY_CACHE_TTL)

_stripped_inventory_cache: dict[str, RuntimeCacheEntry] = {}
_target_lookup_cache: dict[str, RuntimeCacheEntry] = {}

DAILY_REPORT_TYPE = 'daily'
FINAL_REPORT_TYPE = 'final'
PERIOD_REPORT_TYPE = 'period'


def _ms_client_enabled(token: str | None = None, *, location: str | None = None) -> bool:
    enabled_attr = getattr(ms_client, "enabled", None)
    if callable(enabled_attr):
        for args, kwargs in (
            ((token,), {'location': location}),
            ((token,), {}),
            ((), {'location': location}),
            ((), {}),
        ):
            try:
                return bool(enabled_attr(*args, **kwargs))
            except TypeError:
                continue
    return bool(enabled_attr)


def _normalize_optional_ms_value(value: Any) -> str | None:
    raw = str(value or '').strip()
    if not raw or raw.lower() in {'none', 'null', 'undefined'}:
        return None
    return raw


async def _get_location_ms_credentials(location: str, db: AsyncSession | None = None) -> tuple[str | None, str | None]:
    normalized = _normalize_location(location)
    if db is not None:
        point = await db.scalar(select(LocationPoint).where(LocationPoint.name == normalized).limit(1))
        if point:
            token = _normalize_optional_ms_value(point.ms_token)
            store_id = _normalize_optional_ms_value(point.ms_store_id)
            return token, store_id
    token = getattr(settings, 'moysklad_token', None)
    if normalized == 'Дмитров':
        return token, _normalize_optional_ms_value(settings.store_dmitrov_id)
    if normalized == 'Дубна':
        return token, _normalize_optional_ms_value(settings.store_dubna_id)
    return token, None


def _extract_meta_id(meta_or_obj: dict[str, Any] | None) -> str | None:
    if not meta_or_obj:
        return None
    if isinstance(meta_or_obj.get('meta'), dict):
        meta_or_obj = meta_or_obj['meta']
    obj_id = meta_or_obj.get('id')
    if obj_id:
        return str(obj_id)
    href = meta_or_obj.get('href')
    if not href:
        return None
    return str(href).rstrip('/').split('/')[-1]


def _iter_inventory_items(inventory: dict[str, Any]) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    for category in inventory.get('categories', []):
        for subcategory in category.get('subcategories', []):
            for item in subcategory.get('items', []):
                item_id = str(item.get('id') or '').strip()
                if not item_id:
                    continue
                items.append({
                    'id': item_id,
                    'name': str(item.get('name') or '').strip() or item_id,
                })
    return items


@dataclass(slots=True)
class CycleContext:
    cycle_version: int
    started_at: date
    target_date: date


async def _resolve_cycle_context(location: str, db: AsyncSession, requested_date: date | None = None) -> CycleContext:
    normalized = _normalize_location(location)
    cycle = await _get_or_create_selection_cycle(normalized, db)
    if requested_date is None:
        return CycleContext(
            cycle_version=cycle.cycle_version,
            started_at=cycle.started_at,
            target_date=_resolve_cycle_target_date(cycle.started_at, None),
        )

    current_min_date, current_max_date = _cycle_date_bounds(cycle.started_at)
    if current_min_date <= requested_date <= current_max_date:
        return CycleContext(
            cycle_version=cycle.cycle_version,
            started_at=cycle.started_at,
            target_date=_resolve_cycle_target_date(cycle.started_at, requested_date),
        )

    historical_start, historical_end = _cycle_bounds_for_date(requested_date)
    historical_cycle_version = await db.scalar(
        select(Report.cycle_version)
        .where(Report.location == normalized)
        .where(Report.report_date >= historical_start)
        .where(Report.report_date <= historical_end)
        .order_by(Report.report_date.desc(), Report.id.desc())
        .limit(1)
    )
    if historical_cycle_version is None:
        historical_cycle_version = await db.scalar(
            select(SelectionTargetDay.cycle_version)
            .where(SelectionTargetDay.location == normalized)
            .where(SelectionTargetDay.target_date >= historical_start)
            .where(SelectionTargetDay.target_date <= historical_end)
            .order_by(SelectionTargetDay.target_date.desc(), SelectionTargetDay.id.desc())
            .limit(1)
        )
    if historical_cycle_version is None:
        cycle_delta = max(0, _cycle_order_value(cycle.started_at) - _cycle_order_value(requested_date))
        historical_cycle_version = max(1, int(cycle.cycle_version or 1) - cycle_delta)

    return CycleContext(
        cycle_version=int(historical_cycle_version or 1),
        started_at=historical_start,
        target_date=requested_date,
    )


def _get_cached_stripped_inventory(location: str, raw_inventory: dict[str, Any]) -> dict[str, Any]:
    cache_key = _normalize_location(location)
    inventory_identity = id(raw_inventory)
    cached = _stripped_inventory_cache.get(cache_key)
    now = monotonic()
    if cached and cached.inventory_identity == inventory_identity and cached.expires_at > now:
        return cached.value

    stripped = _strip_ignored_inventory_branches(raw_inventory)
    _stripped_inventory_cache[cache_key] = RuntimeCacheEntry(
        value=stripped,
        expires_at=now + STRIPPED_INVENTORY_CACHE_TTL,
        inventory_identity=inventory_identity,
    )
    return stripped


def _build_target_lookup(inventory: dict[str, Any]) -> dict[str, tuple[str, str, str | None, str | None, str, str, float]]:
    lookup: dict[str, tuple[str, str, str | None, str | None, str, str, float]] = {}
    for category in inventory.get('categories', []):
        for subcategory in category.get('subcategories', []):
            expected_total = float(sum(float(item.get('expected_qty') or 0) for item in subcategory.get('items', [])))
            lookup[subcategory['id']] = (
                category['id'],
                category['name'],
                subcategory['id'],
                subcategory['name'],
                'subcategory',
                subcategory['name'],
                expected_total,
            )
            for item in subcategory.get('items', []):
                lookup[item['id']] = (
                    category['id'],
                    category['name'],
                    subcategory['id'],
                    subcategory['name'],
                    'item',
                    item['name'],
                    float(item.get('expected_qty') or 0),
                )
    return lookup


def _get_target_lookup(location: str, inventory: dict[str, Any]) -> dict[str, tuple[str, str, str | None, str | None, str, str, float]]:
    cache_key = _normalize_location(location)
    inventory_identity = id(inventory)
    cached = _target_lookup_cache.get(cache_key)
    now = monotonic()
    if cached and cached.inventory_identity == inventory_identity and cached.expires_at > now:
        return cached.value

    lookup = _build_target_lookup(inventory)
    _target_lookup_cache[cache_key] = RuntimeCacheEntry(
        value=lookup,
        expires_at=now + TARGET_LOOKUP_CACHE_TTL,
        inventory_identity=inventory_identity,
    )
    return lookup


def _invalidate_runtime_inventory_cache(location: str | None = None) -> None:
    if location is None:
        _stripped_inventory_cache.clear()
        _target_lookup_cache.clear()
        return

    normalized = _normalize_location(location)
    _stripped_inventory_cache.pop(normalized, None)
    _target_lookup_cache.pop(normalized, None)


MOCK_INVENTORY: dict[str, dict[str, Any]] = {
    'Дубна': {
        'categories': [
            {
                'id': 'cat-drinks',
                'name': 'Напитки',
                'subcategories': [
                    {
                        'id': 'sub-soda',
                        'name': 'Газировка',
                        'items': [
                            {'id': 'dub-cola', 'name': 'Кола 0.5', 'expected_qty': 6},
                            {'id': 'dub-fanta', 'name': 'Фанта 0.5', 'expected_qty': 4},
                        ],
                    },
                    {
                        'id': 'sub-juice',
                        'name': 'Соки',
                        'items': [
                            {'id': 'dub-apple', 'name': 'Сок яблочный 1л', 'expected_qty': 5},
                            {'id': 'dub-orange', 'name': 'Сок апельсиновый 1л', 'expected_qty': 3},
                        ],
                    },
                ],
            },
            {
                'id': 'cat-snacks',
                'name': 'Снеки',
                'subcategories': [
                    {
                        'id': 'sub-chips',
                        'name': 'Чипсы',
                        'items': [
                            {'id': 'dub-chips-crab', 'name': 'Lays Краб', 'expected_qty': 7},
                            {'id': 'dub-chips-cheese', 'name': 'Lays Сыр', 'expected_qty': 5},
                        ],
                    },
                    {
                        'id': 'sub-croutons',
                        'name': 'Сухарики',
                        'items': [
                            {'id': 'dub-croutons-jelly', 'name': 'Сухарики Холодец', 'expected_qty': 4},
                            {'id': 'dub-croutons-bacon', 'name': 'Сухарики Бекон', 'expected_qty': 6},
                        ],
                    },
                ],
            },
        ]
    },
    'Дмитров': {
        'categories': [
            {
                'id': 'cat-coffee',
                'name': 'Кофе',
                'subcategories': [
                    {
                        'id': 'sub-ice-coffee',
                        'name': 'Холодный кофе',
                        'items': [
                            {'id': 'dm-ice-latte', 'name': 'Айс латте', 'expected_qty': 8},
                            {'id': 'dm-ice-capp', 'name': 'Айс капучино', 'expected_qty': 5},
                        ],
                    }
                ],
            },
            {
                'id': 'cat-food',
                'name': 'Еда',
                'subcategories': [
                    {
                        'id': 'sub-shawarma',
                        'name': 'Шаурма',
                        'items': [
                            {'id': 'dm-shawarma-chicken', 'name': 'Шаурма с курицей', 'expected_qty': 9},
                            {'id': 'dm-shawarma-cheese', 'name': 'Шаурма сырная', 'expected_qty': 3},
                        ],
                    },
                    {
                        'id': 'sub-sandwich',
                        'name': 'Сэндвичи',
                        'items': [
                            {'id': 'dm-sandwich-ham', 'name': 'Сэндвич с ветчиной', 'expected_qty': 4},
                            {'id': 'dm-sandwich-tuna', 'name': 'Сэндвич с тунцом', 'expected_qty': 2},
                        ],
                    },
                ],
            },
        ]
    },
}

MSK_SHIFT = timedelta(hours=3)
SELECTION_CYCLE_DAYS = 15

MSK_TZ = timezone(MSK_SHIFT)
RUS_MONTH_NAMES = [
    'Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь',
    'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь', 'Декабрь',
]

def get_moscow_today() -> date:
    return datetime.now(MSK_TZ).date()


def _cycle_bounds_for_date(target_date: date) -> tuple[date, date]:
    if target_date.day <= 15:
        return target_date.replace(day=1), target_date.replace(day=15)
    month_last_day = calendar.monthrange(target_date.year, target_date.month)[1]
    return target_date.replace(day=16), target_date.replace(day=month_last_day)


def _cycle_index_for_date(target_date: date) -> int:
    return 1 if target_date.day <= 15 else 2


def _cycle_order_value(target_date: date) -> int:
    return target_date.year * 24 + (target_date.month - 1) * 2 + (_cycle_index_for_date(target_date) - 1)


def _cycle_days_left_for_date(target_date: date) -> int:
    _, cycle_end = _cycle_bounds_for_date(target_date)
    return max(0, (cycle_end - target_date).days + 1)


def _cycle_label_for_date(target_date: date) -> str:
    cycle_start, cycle_end = _cycle_bounds_for_date(target_date)
    month_label = f"{RUS_MONTH_NAMES[target_date.month - 1]} {target_date.year}"
    return f"{month_label} · Цикл {_cycle_index_for_date(target_date)} ({cycle_start.day}–{cycle_end.day})"


def hash_password(password: str) -> str:
    salt = os.urandom(16)
    digest = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), salt, 200_000)
    return f"{salt.hex()}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    try:
        salt_hex, digest_hex = stored_hash.split('$', 1)
    except ValueError:
        return False
    digest = hashlib.pbkdf2_hmac('sha256', password.encode('utf-8'), bytes.fromhex(salt_hex), 200_000)
    return hmac.compare_digest(digest.hex(), digest_hex)


_EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')


def _normalize_email(value: str | None) -> str | None:
    raw = str(value or '').strip().lower()
    return raw or None


def _validate_email(value: str | None) -> str | None:
    email = _normalize_email(value)
    if email and not _EMAIL_RE.match(email):
        raise HTTPException(status_code=400, detail='Укажите корректный email.')
    return email


async def _ensure_email_is_unique(email: str | None, db: AsyncSession, *, exclude_user_id: int | None = None) -> None:
    if not email:
        return
    query = select(User.id).where(func.lower(User.email) == email.lower())
    if exclude_user_id is not None:
        query = query.where(User.id != exclude_user_id)
    duplicate_id = await db.scalar(query.limit(1))
    if duplicate_id:
        raise HTTPException(status_code=400, detail='Пользователь с таким email уже существует.')


def _normalize_location(location: str) -> str:
    return location.strip().title()


def _strip_ignored_inventory_branches(inventory: dict[str, Any]) -> dict[str, Any]:
    categories: list[dict[str, Any]] = []
    for category in inventory.get('categories', []):
        if category.get('name') == DEFAULT_CATEGORY_NAME:
            continue
        filtered_subcategories: list[dict[str, Any]] = []
        for subcategory in category.get('subcategories', []):
            if subcategory.get('name') == DEFAULT_SUBCATEGORY_NAME:
                continue
            filtered_subcategories.append({
                'id': subcategory['id'],
                'name': subcategory['name'],
                'items': [dict(item) for item in subcategory.get('items', [])],
            })
        if filtered_subcategories:
            categories.append({
                'id': category['id'],
                'name': category['name'],
                'subcategories': filtered_subcategories,
            })
    return {'location': inventory.get('location'), 'categories': categories}


def _env_integration_rows() -> list[dict[str, str | None]]:
    rows: list[dict[str, str | None]] = []
    seen: set[str] = set()
    shared_token = (settings.moysklad_token or '').strip() or None
    for raw_name, raw_store_id in [
        (settings.store_dubna, settings.store_dubna_id),
        (settings.store_dmitrov, settings.store_dmitrov_id),
    ]:
        normalized = _normalize_location(raw_name or '') if raw_name else ''
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        rows.append({
            'name': normalized,
            'ms_token': shared_token,
            'ms_store_id': (str(raw_store_id or '').strip() or None),
            'ms_store_name': normalized,
        })
    return rows


async def _sync_location_points_from_env(db: AsyncSession) -> None:
    env_rows = _env_integration_rows()
    if not env_rows:
        return

    existing_rows = (await db.scalars(select(LocationPoint).order_by(LocationPoint.id.asc()))).all()
    existing_by_name = {_normalize_location(row.name): row for row in existing_rows}
    dirty = False

    for payload in env_rows:
        point = existing_by_name.get(payload['name'])
        if point is None:
            db.add(LocationPoint(
                name=payload['name'],
                ms_token=payload['ms_token'],
                ms_store_id=payload['ms_store_id'],
                ms_store_name=payload['ms_store_name'],
            ))
            dirty = True
            continue

        next_token = payload['ms_token']
        next_store_id = payload['ms_store_id']
        next_store_name = payload['ms_store_name']
        if (point.ms_token or None) != next_token:
            point.ms_token = next_token
            dirty = True
        if (point.ms_store_id or None) != next_store_id:
            point.ms_store_id = next_store_id
            dirty = True
        if (point.ms_store_name or None) != next_store_name:
            point.ms_store_name = next_store_name
            dirty = True

    if dirty:
        await db.commit()


async def _ensure_default_location_points(db: AsyncSession) -> None:
    await _sync_location_points_from_env(db)

    existing_count = await db.scalar(select(func.count()).select_from(LocationPoint))
    if (existing_count or 0) > 0:
        return

    candidates: list[tuple[str, str | None]] = []
    seen: set[str] = set()
    for row in _env_integration_rows():
        normalized = row['name'] or ''
        if normalized and normalized not in seen:
            candidates.append((normalized, row['ms_store_id']))
            seen.add(normalized)

    if not candidates:
        for name in sorted(MOCK_INVENTORY.keys()):
            normalized = _normalize_location(name)
            if normalized not in seen:
                candidates.append((normalized, None))
                seen.add(normalized)

    for name, store_id in candidates:
        db.add(LocationPoint(
            name=name,
            ms_token=(settings.moysklad_token or '').strip() or None,
            ms_store_id=store_id,
            ms_store_name=name,
        ))
    await db.commit()




def _moscow_datetime(dt: datetime | None) -> datetime | None:
    if not dt:
        return None
    return dt if dt.tzinfo is not None else dt.replace(tzinfo=timezone.utc)


def _moscow_date_for_storage(dt: datetime | None) -> date | None:
    value = _moscow_datetime(dt)
    if not value:
        return None
    return value.astimezone(MSK_TZ).date()


def _sanitize_historical_daily_report_snapshots(
    report: Report,
    report_snapshots: list[ReportTargetSnapshot],
    results: list[CheckResult],
) -> list[ReportTargetSnapshot]:
    if not report_snapshots:
        return report_snapshots

    report_day = report.report_date
    result_keys = {
        (row.target_type, row.target_id, row.checked_by_user_id)
        for row in results
        if row.checked_by_user_id is not None
    }

    sanitized: list[ReportTargetSnapshot] = []
    for row in report_snapshots:
        created_on = _moscow_date_for_storage(row.created_at)
        if created_on == report_day:
            sanitized.append(row)
            continue
        if (row.target_type, row.target_id, row.assigned_user_id_snapshot) in result_keys:
            sanitized.append(row)
            continue
    return sanitized

def _format_moscow_datetime(dt: datetime | None) -> str:
    if not dt:
        return '-'
    return (dt + MSK_SHIFT).strftime('%d.%m.%Y %H:%M')


def _report_status_label(status: str) -> str:
    if status == 'created':
        return 'Назначена'
    if status == 'completed':
        return 'Завершена'
    return 'В процессе'


def bootstrap_schema_and_admin(sync_conn) -> None:
    inspector = inspect(sync_conn)
    tables = set(inspector.get_table_names())

    if 'users' in tables:
        columns = {c['name'] for c in inspector.get_columns('users')}
        required = {'id', 'full_name', 'birth_date', 'username', 'password_hash', 'role', 'location', 'is_active', 'created_at'}
        if not required.issubset(columns):
            sync_conn.execute(text('DROP TABLE IF EXISTS users'))
        elif 'email' not in columns:
            sync_conn.execute(text('ALTER TABLE users ADD COLUMN email VARCHAR(255)'))
            sync_conn.execute(text('CREATE INDEX IF NOT EXISTS ix_users_email ON users (email)'))

    if 'password_reset_requests' in tables:
        cols = {c['name'] for c in inspector.get_columns('password_reset_requests')}
        required = {'id', 'user_id', 'request_id', 'code_hash', 'reset_token_hash', 'attempts', 'expires_at', 'created_at', 'last_sent_at', 'verified_at', 'used_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS password_reset_requests'))

    reset_reports = False
    if 'reports' in tables:
        cols = {c['name'] for c in inspector.get_columns('reports')}
        if not {'id', 'location', 'report_date', 'cycle_version', 'status', 'date_created'}.issubset(cols):
            reset_reports = True
        elif 'report_type' not in cols:
            sync_conn.execute(text("ALTER TABLE reports ADD COLUMN report_type VARCHAR(20) NOT NULL DEFAULT 'daily'"))

    if 'check_results' in tables:
        cols = {c['name'] for c in inspector.get_columns('check_results')}
        required = {'id', 'report_id', 'category_id', 'category_name', 'subcategory_id', 'subcategory_name', 'target_type', 'target_id', 'target_name', 'expected_qty', 'actual_qty', 'diff', 'status', 'attempts_used', 'checked_by_user_id', 'checked_by_name_snapshot', 'created_at'}
        if not required.issubset(cols):
            reset_reports = True

    reset_assignments = False
    if 'category_assignments' in tables:
        cols = {c['name'] for c in inspector.get_columns('category_assignments')}
        required = {'id', 'location', 'cycle_version', 'category_id', 'category_name', 'subcategory_id', 'subcategory_name', 'target_type', 'target_id', 'target_name', 'user_id', 'user_full_name_snapshot', 'assigned_at'}
        if not required.issubset(cols):
            reset_assignments = True

    if 'selection_cycles' in tables:
        cols = {c['name'] for c in inspector.get_columns('selection_cycles')}
        required = {'id', 'location', 'cycle_version', 'started_at', 'updated_at'}
        if not required.issubset(cols):
            reset_assignments = True
    elif 'selection_cycles' not in tables:
        reset_assignments = True if 'category_assignments' in tables else reset_assignments

    if 'selection_targets' in tables:
        cols = {c['name'] for c in inspector.get_columns('selection_targets')}
        required = {'id', 'location', 'cycle_version', 'category_id', 'category_name', 'subcategory_id', 'subcategory_name', 'target_type', 'target_id', 'target_name', 'created_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS selection_targets'))


    if 'selection_target_days' in tables:
        cols = {c['name'] for c in inspector.get_columns('selection_target_days')}
        required = {'id', 'location', 'cycle_version', 'target_date', 'category_id', 'category_name', 'subcategory_id', 'subcategory_name', 'target_type', 'target_id', 'target_name', 'created_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS selection_target_days'))

    if 'location_points' in tables:
        cols = {c['name'] for c in inspector.get_columns('location_points')}
        required = {'id', 'name', 'ms_token', 'ms_store_id', 'ms_store_name', 'created_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS location_points'))

    if 'admin_location_access' in tables:
        cols = {c['name'] for c in inspector.get_columns('admin_location_access')}
        required = {'id', 'admin_user_id', 'location_point_id', 'granted_by_user_id', 'created_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS admin_location_access'))

    if 'verify_attempt_progress' in tables:
        cols = {c['name'] for c in inspector.get_columns('verify_attempt_progress')}
        required = {'id', 'report_id', 'target_type', 'target_id', 'checked_by_user_id', 'attempts_used', 'created_at', 'updated_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS verify_attempt_progress'))

    if 'report_target_snapshots' in tables:
        cols = {c['name'] for c in inspector.get_columns('report_target_snapshots')}
        required = {'id', 'report_id', 'category_id', 'category_name', 'subcategory_id', 'subcategory_name', 'target_type', 'target_id', 'target_name', 'assigned_user_id_snapshot', 'assigned_user_name_snapshot', 'created_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS report_target_snapshots'))

    if 'report_employee_completions' in tables:
        cols = {c['name'] for c in inspector.get_columns('report_employee_completions')}
        required = {'id', 'report_id', 'user_id', 'user_full_name_snapshot', 'finished_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS report_employee_completions'))

    if 'product_financial_cache' in tables:
        cols = {c['name'] for c in inspector.get_columns('product_financial_cache')}
        required = {'id', 'location_point_id', 'item_id', 'item_name', 'item_code', 'cost_price', 'retail_price', 'source_refreshed_at', 'created_at', 'updated_at'}
        if not required.issubset(cols):
            sync_conn.execute(text('DROP TABLE IF EXISTS product_financial_cache'))


    if 'expense_templates' in tables:
        cols = {c['name'] for c in inspector.get_columns('expense_templates')}
        if 'created_by_user_id' not in cols:
            sync_conn.execute(text('ALTER TABLE expense_templates ADD COLUMN created_by_user_id INTEGER'))
        if 'day_of_month' not in cols:
            sync_conn.execute(text('ALTER TABLE expense_templates ADD COLUMN day_of_month INTEGER NOT NULL DEFAULT 1'))

    if 'monthly_expense_entries' in tables:
        cols = {c['name'] for c in inspector.get_columns('monthly_expense_entries')}
        if 'custom_name' not in cols:
            sync_conn.execute(text('ALTER TABLE monthly_expense_entries ADD COLUMN custom_name VARCHAR(255)'))
        if 'comment' not in cols:
            sync_conn.execute(text('ALTER TABLE monthly_expense_entries ADD COLUMN comment TEXT'))
        if 'created_by_user_id' not in cols:
            sync_conn.execute(text('ALTER TABLE monthly_expense_entries ADD COLUMN created_by_user_id INTEGER'))
        if 'expense_date' not in cols:
            sync_conn.execute(text('ALTER TABLE monthly_expense_entries ADD COLUMN expense_date DATE'))
            sync_conn.execute(text('UPDATE monthly_expense_entries SET expense_date = month_start WHERE expense_date IS NULL'))
        if 'distribution_mode' not in cols:
            sync_conn.execute(text("ALTER TABLE monthly_expense_entries ADD COLUMN distribution_mode VARCHAR(20) NOT NULL DEFAULT 'spread'"))
            sync_conn.execute(text("UPDATE monthly_expense_entries SET distribution_mode = 'single_day' WHERE template_id IS NULL"))

    if reset_reports:
        sync_conn.execute(text('DROP TABLE IF EXISTS check_results'))
        sync_conn.execute(text('DROP TABLE IF EXISTS reports'))

    if reset_assignments:
        sync_conn.execute(text('DROP TABLE IF EXISTS category_assignments'))
        sync_conn.execute(text('DROP TABLE IF EXISTS selection_cycles'))

    from app.database import Base
    Base.metadata.create_all(sync_conn)


async def _assign_admin_location_access_by_ids(admin_user_id: int, location_ids: list[int], db: AsyncSession, granted_by_user_id: int | None = None) -> None:
    await db.execute(delete(AdminLocationAccess).where(AdminLocationAccess.admin_user_id == admin_user_id))
    unique_ids: list[int] = []
    seen: set[int] = set()
    for location_id in location_ids:
        normalized_id = int(location_id)
        if normalized_id not in seen:
            unique_ids.append(normalized_id)
            seen.add(normalized_id)
    for location_id in unique_ids:
        db.add(AdminLocationAccess(
            admin_user_id=admin_user_id,
            location_point_id=location_id,
            granted_by_user_id=granted_by_user_id,
        ))


async def _get_admin_location_rows(admin_user_id: int, db: AsyncSession) -> list[LocationPoint]:
    return (
        await db.scalars(
            select(LocationPoint)
            .join(AdminLocationAccess, AdminLocationAccess.location_point_id == LocationPoint.id)
            .where(AdminLocationAccess.admin_user_id == admin_user_id)
            .order_by(LocationPoint.name.asc())
        )
    ).all()


async def get_user_accessible_locations(user: User, db: AsyncSession) -> list[str]:
    if user.role == RoleEnum.SUPERADMIN.value:
        rows = (await db.scalars(select(LocationPoint).order_by(LocationPoint.name.asc()))).all()
        return [row.name for row in rows]
    if user.role == RoleEnum.ADMIN.value:
        return [row.name for row in await _get_admin_location_rows(user.id, db)]
    if user.location:
        return [_normalize_location(user.location)]
    return []


async def get_user_accessible_location_ids(user: User, db: AsyncSession) -> list[int]:
    if user.role == RoleEnum.SUPERADMIN.value:
        rows = (await db.scalars(select(LocationPoint.id).order_by(LocationPoint.name.asc()))).all()
        return [int(row) for row in rows]
    if user.role == RoleEnum.ADMIN.value:
        rows = (await db.scalars(
            select(AdminLocationAccess.location_point_id)
            .where(AdminLocationAccess.admin_user_id == user.id)
            .order_by(AdminLocationAccess.location_point_id.asc())
        )).all()
        return [int(row) for row in rows]
    if not user.location:
        return []
    location_id = await db.scalar(select(LocationPoint.id).where(LocationPoint.name == _normalize_location(user.location)).limit(1))
    return [int(location_id)] if location_id else []


async def user_has_location_access(user: User, location: str, db: AsyncSession) -> bool:
    normalized = _normalize_location(location)
    if user.role == RoleEnum.SUPERADMIN.value:
        return True
    if user.role == RoleEnum.ADMIN.value:
        return normalized in set(await get_user_accessible_locations(user, db))
    return normalized == _normalize_location(user.location or '')


async def ensure_user_can_access_location(user: User, location: str, db: AsyncSession) -> None:
    if not await user_has_location_access(user, location, db):
        raise HTTPException(status_code=403, detail='Нет доступа к выбранной точке.')


async def _validate_location_ids(location_ids: list[int], db: AsyncSession) -> list[LocationPoint]:
    unique_ids = sorted({int(location_id) for location_id in location_ids})
    if not unique_ids:
        return []
    rows = (await db.scalars(select(LocationPoint).where(LocationPoint.id.in_(unique_ids)).order_by(LocationPoint.name.asc()))).all()
    if len(rows) != len(unique_ids):
        raise HTTPException(status_code=400, detail='Выбраны несуществующие точки.')
    return rows


async def _build_user_response(user: User, db: AsyncSession) -> UserResponse:
    admin_rows = await _get_admin_location_rows(user.id, db) if user.role == RoleEnum.ADMIN.value else []
    return UserResponse(
        id=user.id,
        full_name=user.full_name,
        birth_date=user.birth_date,
        username=user.username,
        email=user.email,
        role=RoleEnum(user.role),
        location=user.location,
        is_active=user.is_active,
        admin_location_ids=[row.id for row in admin_rows],
        admin_locations=[row.name for row in admin_rows],
    )


async def _migrate_admin_location_access(db: AsyncSession) -> None:
    admins = (await db.scalars(select(User).where(User.role == RoleEnum.ADMIN.value))).all()
    changed = False
    for admin in admins:
        access_count = await db.scalar(select(func.count()).select_from(AdminLocationAccess).where(AdminLocationAccess.admin_user_id == admin.id))
        if (access_count or 0) > 0:
            continue
        if not admin.location:
            continue
        point = await db.scalar(select(LocationPoint).where(LocationPoint.name == _normalize_location(admin.location)).limit(1))
        if not point:
            continue
        db.add(AdminLocationAccess(admin_user_id=admin.id, location_point_id=point.id, granted_by_user_id=None))
        changed = True
    if changed:
        await db.commit()


async def ensure_default_admin(db: AsyncSession) -> None:
    await _ensure_default_location_points(db)

    superadmin = await db.scalar(select(User).where(User.role == RoleEnum.SUPERADMIN.value).limit(1))
    if not superadmin:
        default_user = await db.scalar(select(User).where(User.username == settings.default_admin_username).limit(1))
        if default_user:
            default_user.role = RoleEnum.SUPERADMIN.value
            default_user.location = None
            default_user.is_active = True
            await db.commit()
        else:
            existing_admin = await db.scalar(select(User).where(User.role == RoleEnum.ADMIN.value).order_by(User.id.asc()).limit(1))
            if existing_admin:
                existing_admin.role = RoleEnum.SUPERADMIN.value
                existing_admin.location = None
                existing_admin.is_active = True
                await db.commit()
            else:
                admin_user = User(
                    full_name=settings.default_admin_full_name,
                    birth_date=date.fromisoformat(settings.default_admin_birth_date),
                    username=settings.default_admin_username,
                    password_hash=hash_password(settings.default_admin_password),
                    role=RoleEnum.SUPERADMIN.value,
                    location=None,
                    is_active=True,
                )
                db.add(admin_user)
                await db.commit()

    await _migrate_admin_location_access(db)


def user_to_schema(user: User) -> UserInfo:
    return UserInfo(
        id=user.id,
        full_name=user.full_name,
        birth_date=user.birth_date,
        username=user.username,
        email=user.email,
        role=RoleEnum(user.role),
        location=user.location,
        is_active=user.is_active,
        admin_location_ids=[],
        admin_locations=[],
    )


async def authenticate_user(username: str, password: str, db: AsyncSession) -> User | None:
    user = await db.scalar(select(User).where(User.username == username).limit(1))
    if not user or not user.is_active:
        return None
    if not verify_password(password, user.password_hash):
        return None
    return user


async def get_me_response(user: User | None) -> MeResponse:
    return MeResponse(authenticated=bool(user), user=user_to_schema(user) if user else None)


async def prewarm_inventory_cache(location: str | None) -> None:
    if not location:
        return
    normalized = _normalize_location(location)
    if _ms_client_enabled():
        await ms_client.prewarm_inventory(normalized)


async def list_users(db: AsyncSession, current_user: User) -> UserListResponse:
    accessible_locations = set(await get_user_accessible_locations(current_user, db))

    if current_user.role == RoleEnum.SUPERADMIN.value:
        users = (await db.scalars(select(User))).all()
    else:
        conditions = [User.id == current_user.id]
        if accessible_locations:
            conditions.append(and_(User.role == RoleEnum.EMPLOYEE.value, User.location.in_(accessible_locations)))
        users = (await db.scalars(select(User).where(or_(*conditions)))).all()

    role_order = {
        RoleEnum.SUPERADMIN.value: 0,
        RoleEnum.ADMIN.value: 1,
        RoleEnum.EMPLOYEE.value: 2,
    }
    ordered_users = sorted(users, key=lambda item: (role_order.get(item.role, 99), item.full_name.lower()))
    return UserListResponse(users=[await _build_user_response(user, db) for user in ordered_users])


async def create_user(payload: UserCreateRequest, db: AsyncSession, current_user: User) -> UserActionResponse:
    normalized_username = payload.username.strip()
    normalized_email = _validate_email(payload.email)
    existing = await db.scalar(select(User).where(User.username == normalized_username).limit(1))
    if existing:
        raise HTTPException(status_code=400, detail='Пользователь с таким логином уже существует.')
    await _ensure_email_is_unique(normalized_email, db)

    requested_role = payload.role.value
    normalized_location = _normalize_location(payload.location) if payload.location else None
    requested_access_ids = sorted({int(location_id) for location_id in payload.admin_location_ids})

    if current_user.role == RoleEnum.ADMIN.value and requested_role != RoleEnum.EMPLOYEE.value:
        raise HTTPException(status_code=403, detail='Обычный управляющий может создавать только сотрудников.')
    if requested_role == RoleEnum.SUPERADMIN.value and current_user.role != RoleEnum.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail='Создавать главного управляющего может только главный управляющий.')
    if requested_role == RoleEnum.ADMIN.value and current_user.role != RoleEnum.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail='Создавать управляющих может только главный управляющий.')

    if requested_role == RoleEnum.EMPLOYEE.value:
        if not normalized_location:
            raise HTTPException(status_code=400, detail='Сотруднику нужно назначить точку.')
        location_point = await db.scalar(select(LocationPoint).where(LocationPoint.name == normalized_location).limit(1))
        if not location_point:
            raise HTTPException(status_code=400, detail='Выбрана несуществующая точка.')
        if current_user.role == RoleEnum.ADMIN.value:
            await ensure_user_can_access_location(current_user, normalized_location, db)
    else:
        normalized_location = None

    admin_location_rows: list[LocationPoint] = []
    if requested_role == RoleEnum.ADMIN.value:
        admin_location_rows = await _validate_location_ids(requested_access_ids, db)
        if not admin_location_rows:
            raise HTTPException(status_code=400, detail='Управляющему нужно назначить хотя бы одну точку.')

    user = User(
        full_name=payload.full_name.strip(),
        birth_date=payload.birth_date,
        username=normalized_username,
        email=normalized_email,
        password_hash=hash_password(payload.password),
        role=requested_role,
        location=normalized_location,
        is_active=payload.is_active,
    )
    db.add(user)
    await db.flush()

    if requested_role == RoleEnum.ADMIN.value:
        await _assign_admin_location_access_by_ids(user.id, [row.id for row in admin_location_rows], db, granted_by_user_id=current_user.id)

    await db.commit()
    await db.refresh(user)
    return UserActionResponse(success=True, message='Пользователь создан.', user=await _build_user_response(user, db))


async def update_user(user_id: int, payload: UserUpdateRequest, db: AsyncSession, current_user: User) -> UserActionResponse:
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail='Пользователь не найден.')

    normalized_username = payload.username.strip()
    normalized_email = _validate_email(payload.email)
    duplicate = await db.scalar(select(User).where(User.username == normalized_username, User.id != user_id).limit(1))
    if duplicate:
        raise HTTPException(status_code=400, detail='Пользователь с таким логином уже существует.')
    await _ensure_email_is_unique(normalized_email, db, exclude_user_id=user_id)

    requested_role = payload.role.value
    normalized_location = _normalize_location(payload.location) if payload.location else None
    requested_access_ids = sorted({int(location_id) for location_id in payload.admin_location_ids})

    if current_user.role == RoleEnum.ADMIN.value:
        accessible_locations = set(await get_user_accessible_locations(current_user, db))
        if user.id == current_user.id:
            if user.role != RoleEnum.ADMIN.value or requested_role != RoleEnum.ADMIN.value:
                raise HTTPException(status_code=400, detail='Нельзя снять роль admin у своего аккаунта.')
        else:
            if user.role != RoleEnum.EMPLOYEE.value:
                raise HTTPException(status_code=403, detail='Обычный управляющий может редактировать только сотрудников.')
            if user.location and _normalize_location(user.location) not in accessible_locations:
                raise HTTPException(status_code=403, detail='Нет доступа к пользователю из другой точки.')
            if requested_role != RoleEnum.EMPLOYEE.value:
                raise HTTPException(status_code=403, detail='Обычный управляющий не может менять роль сотрудника.')
            if not normalized_location:
                raise HTTPException(status_code=400, detail='Сотруднику нужно назначить точку.')
            if normalized_location not in accessible_locations:
                raise HTTPException(status_code=403, detail='Нельзя назначить сотруднику чужую точку.')

    if requested_role == RoleEnum.SUPERADMIN.value and current_user.role != RoleEnum.SUPERADMIN.value:
        raise HTTPException(status_code=403, detail='Назначать роль главного управляющего может только главный управляющий.')
    if requested_role == RoleEnum.ADMIN.value and current_user.role != RoleEnum.SUPERADMIN.value and user.id != current_user.id:
        raise HTTPException(status_code=403, detail='Назначать роль управляющего может только главный управляющий.')

    if user.id == current_user.id and user.role == RoleEnum.SUPERADMIN.value and requested_role != RoleEnum.SUPERADMIN.value:
        raise HTTPException(status_code=400, detail='Нельзя снять роль главного управляющего у своего аккаунта.')

    old_superadmin = user.role == RoleEnum.SUPERADMIN.value

    admin_location_rows: list[LocationPoint] = []
    if requested_role == RoleEnum.ADMIN.value:
        if current_user.role == RoleEnum.SUPERADMIN.value:
            admin_location_rows = await _validate_location_ids(requested_access_ids, db)
            if user.id != current_user.id and not admin_location_rows:
                raise HTTPException(status_code=400, detail='Управляющему нужно назначить хотя бы одну точку.')
        normalized_location = None
    elif requested_role == RoleEnum.EMPLOYEE.value:
        if not normalized_location:
            raise HTTPException(status_code=400, detail='Сотруднику нужно назначить точку.')
        location_point = await db.scalar(select(LocationPoint).where(LocationPoint.name == normalized_location).limit(1))
        if not location_point:
            raise HTTPException(status_code=400, detail='Выбрана несуществующая точка.')
    else:
        normalized_location = None

    old_name = user.full_name
    user.full_name = payload.full_name.strip()
    user.birth_date = payload.birth_date
    user.username = normalized_username
    user.email = normalized_email
    user.role = requested_role
    user.location = normalized_location
    user.is_active = payload.is_active
    if payload.password:
        user.password_hash = hash_password(payload.password)

    if old_name != user.full_name:
        await db.execute(update(CategoryAssignment).where(CategoryAssignment.user_id == user.id).values(user_full_name_snapshot=user.full_name))
        await db.execute(update(CheckResult).where(CheckResult.checked_by_user_id == user.id).values(checked_by_name_snapshot=user.full_name))

    if requested_role == RoleEnum.ADMIN.value:
        if current_user.role == RoleEnum.SUPERADMIN.value:
            if admin_location_rows:
                await _assign_admin_location_access_by_ids(user.id, [row.id for row in admin_location_rows], db, granted_by_user_id=current_user.id)
            elif user.id != current_user.id:
                await _assign_admin_location_access_by_ids(user.id, [], db, granted_by_user_id=current_user.id)
        user.location = None
    else:
        await db.execute(delete(AdminLocationAccess).where(AdminLocationAccess.admin_user_id == user.id))

    if old_superadmin and requested_role != RoleEnum.SUPERADMIN.value:
        superadmin_count = await db.scalar(select(func.count()).select_from(User).where(User.role == RoleEnum.SUPERADMIN.value))
        if (superadmin_count or 0) <= 1:
            raise HTTPException(status_code=400, detail='Нельзя снять роль у последнего главного управляющего.')

    await db.commit()
    await db.refresh(user)
    return UserActionResponse(success=True, message='Пользователь обновлён.', user=await _build_user_response(user, db))


async def delete_user(user_id: int, db: AsyncSession, current_user: User) -> DeleteResponse:
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail='Пользователь не найден.')

    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail='Нельзя удалить собственный аккаунт.')

    if current_user.role == RoleEnum.ADMIN.value:
        accessible_locations = set(await get_user_accessible_locations(current_user, db))
        if user.role != RoleEnum.EMPLOYEE.value:
            raise HTTPException(status_code=403, detail='Обычный управляющий может удалять только сотрудников.')
        if not user.location or _normalize_location(user.location) not in accessible_locations:
            raise HTTPException(status_code=403, detail='Нет доступа к пользователю из другой точки.')

    if user.role == RoleEnum.SUPERADMIN.value:
        superadmin_count = await db.scalar(select(func.count()).select_from(User).where(User.role == RoleEnum.SUPERADMIN.value))
        if (superadmin_count or 0) <= 1:
            raise HTTPException(status_code=400, detail='Нельзя удалить последнего главного управляющего.')

    await db.execute(delete(CategoryAssignment).where(CategoryAssignment.user_id == user.id))
    await db.execute(update(CheckResult).where(CheckResult.checked_by_user_id == user.id).values(checked_by_user_id=None))
    await db.execute(delete(AdminLocationAccess).where(AdminLocationAccess.admin_user_id == user.id))
    await db.execute(delete(PasswordResetRequest).where(PasswordResetRequest.user_id == user.id))
    await db.delete(user)
    await db.commit()
    return DeleteResponse(success=True, message='Пользователь удалён.')


async def _get_inventory_for(location: str, db: AsyncSession | None = None) -> dict[str, Any]:
    normalized = _normalize_location(location)
    started = monotonic()
    token, store_id = await _get_location_ms_credentials(normalized, db)
    if _ms_client_enabled(token, location=normalized):
        logger.info(
            'Загрузка inventory началась. location=%s source=moysklad token_source=%s store_source=%s',
            normalized,
            'db' if token else 'fallback',
            'db' if store_id else 'fallback',
        )
        try:
            inventory = await ms_client.get_inventory(normalized, token=token, store_id=store_id)
            stripped = _get_cached_stripped_inventory(normalized, inventory)
            duration_ms = round((monotonic() - started) * 1000, 1)
            logger.info(
                'Загрузка inventory завершена. location=%s source=moysklad categories=%s duration_ms=%s',
                normalized,
                len(stripped.get('categories', [])),
                duration_ms,
            )
            return stripped
        except ValueError as exc:
            duration_ms = round((monotonic() - started) * 1000, 1)
            logger.warning('Inventory не найден. location=%s duration_ms=%s detail=%s', normalized, duration_ms, exc)
            raise HTTPException(status_code=404, detail=str(exc)) from exc
        except httpx.HTTPError as exc:
            duration_ms = round((monotonic() - started) * 1000, 1)
            logger.exception('Ошибка загрузки inventory из МойСклад. location=%s duration_ms=%s', normalized, duration_ms)
            raise HTTPException(status_code=502, detail='Не удалось получить данные из МойСклад. Попробуйте ещё раз.') from exc
        except Exception:
            duration_ms = round((monotonic() - started) * 1000, 1)
            logger.exception('Непредвиденная ошибка загрузки inventory. location=%s duration_ms=%s', normalized, duration_ms)
            raise

    if normalized not in MOCK_INVENTORY:
        logger.warning('Inventory не найден в mock-данных. location=%s', normalized)
        raise HTTPException(status_code=404, detail='Неизвестная точка.')

    stripped = _get_cached_stripped_inventory(normalized, MOCK_INVENTORY[normalized])
    duration_ms = round((monotonic() - started) * 1000, 1)
    logger.info(
        'Загрузка inventory завершена. location=%s source=mock categories=%s duration_ms=%s',
        normalized,
        len(stripped.get('categories', [])),
        duration_ms,
    )
    return stripped


async def _find_category(location: str, category_id: str, db: AsyncSession | None = None) -> dict[str, Any]:
    inventory = await _get_inventory_for(location, db=db)
    for category in inventory['categories']:
        if category['id'] == category_id:
            return category
    raise HTTPException(status_code=404, detail='Категория не найдена.')


async def _find_subcategory(location: str, category_id: str, subcategory_id: str, db: AsyncSession | None = None) -> dict[str, Any]:
    category = await _find_category(location, category_id, db=db)
    for sub in category['subcategories']:
        if sub['id'] == subcategory_id:
            return sub
    raise HTTPException(status_code=404, detail='Подкатегория не найдена.')


async def _find_target(location: str, target_id: str, db: AsyncSession | None = None) -> tuple[str, str, str | None, str | None, str, str, float]:
    inventory = await _get_inventory_for(location, db=db)
    target = _get_target_lookup(location, inventory).get(target_id)
    if target is not None:
        return target
    raise HTTPException(status_code=404, detail='Цель проверки не найдена.')


async def get_inventory_diagnostics_details(location: str, db: AsyncSession | None = None) -> list[dict[str, Any]]:
    inventory = await _get_inventory_for(location, db=db)
    rows: list[dict[str, Any]] = []
    normalized_location = inventory.get('location') or _normalize_location(location)

    for category in inventory['categories']:
        for subcategory in category['subcategories']:
            category_is_default = category['name'] == DEFAULT_CATEGORY_NAME
            subcategory_is_default = subcategory['name'] == DEFAULT_SUBCATEGORY_NAME
            if not (category_is_default or subcategory_is_default):
                continue

            issue_parts: list[str] = []
            if category_is_default:
                issue_parts.append('без категории')
            if subcategory_is_default:
                issue_parts.append('без подкатегории')
            issue_label = ' и '.join(issue_parts)

            for item in subcategory['items']:
                diagnostics = item.get('diagnostics') or {}
                folder_chain = diagnostics.get('folder_chain') or []
                folder_path = ' → '.join(part.get('name', '') for part in folder_chain if part.get('name'))
                rows.append({
                    'location': normalized_location,
                    'issue_type': issue_label,
                    'category_name': category['name'],
                    'subcategory_name': subcategory['name'],
                    'item_id': item['id'],
                    'item_name': item['name'],
                    'expected_qty': item.get('expected_qty', 0),
                    'reason': diagnostics.get('reason') or (
                        'У товара не определилась категория или подкатегория. Проверьте папку товара в МойСклад.'
                        if category_is_default or subcategory_is_default
                        else 'Товар размечен корректно.'
                    ),
                    'folder_path': folder_path or '-',
                    'folder_source': diagnostics.get('folder_source') or '-',
                    'assortment_lookup': diagnostics.get('assortment_lookup') or '-',
                })

    rows.sort(key=lambda row: (str(row['issue_type']).lower(), str(row['category_name']).lower(), str(row['subcategory_name']).lower(), str(row['item_name']).lower()))
    return rows


async def get_inventory_diagnostics_rows(location: str, db: AsyncSession | None = None) -> list[dict[str, Any]]:
    rows = await get_inventory_diagnostics_details(location, db=db)
    return [
        {
            'location': row['location'],
            'issue_type': row['issue_type'],
            'category_name': row['category_name'],
            'subcategory_name': row['subcategory_name'],
            'item_id': row['item_id'],
            'item_name': row['item_name'],
            'expected_qty': row['expected_qty'],
            'reason': row['reason'],
            'folder_path': row['folder_path'],
            'folder_source': row['folder_source'],
            'assortment_lookup': row['assortment_lookup'],
        }
        for row in rows
    ]



async def _load_report_target_snapshots(report_id: int, db: AsyncSession) -> list[ReportTargetSnapshot]:
    return (
        await db.scalars(
            select(ReportTargetSnapshot)
            .where(ReportTargetSnapshot.report_id == report_id)
            .order_by(ReportTargetSnapshot.id.asc())
        )
    ).all()


async def _upsert_report_target_snapshot(
    *,
    report_id: int,
    category_id: str,
    category_name: str,
    subcategory_id: str | None,
    subcategory_name: str | None,
    target_type: str,
    target_id: str,
    target_name: str,
    assigned_user_id_snapshot: int | None,
    assigned_user_name_snapshot: str | None,
    db: AsyncSession,
) -> None:
    existing = await db.scalar(
        select(ReportTargetSnapshot)
        .where(ReportTargetSnapshot.report_id == report_id)
        .where(ReportTargetSnapshot.target_type == target_type)
        .where(ReportTargetSnapshot.target_id == target_id)
        .where(ReportTargetSnapshot.assigned_user_id_snapshot == assigned_user_id_snapshot)
        .limit(1)
    )
    if existing:
        existing.category_id = category_id
        existing.category_name = category_name
        existing.subcategory_id = subcategory_id
        existing.subcategory_name = subcategory_name
        existing.target_name = target_name
        existing.assigned_user_name_snapshot = assigned_user_name_snapshot
        return

    db.add(ReportTargetSnapshot(
        report_id=report_id,
        category_id=category_id,
        category_name=category_name,
        subcategory_id=subcategory_id,
        subcategory_name=subcategory_name,
        target_type=target_type,
        target_id=target_id,
        target_name=target_name,
        assigned_user_id_snapshot=assigned_user_id_snapshot,
        assigned_user_name_snapshot=assigned_user_name_snapshot,
    ))


async def _bootstrap_report_target_snapshots(
    report: Report,
    assignments: list[CategoryAssignment],
    results: list[CheckResult],
    db: AsyncSession,
) -> list[ReportTargetSnapshot]:
    existing = await _load_report_target_snapshots(report.id, db)
    existing_keys = {
        (row.target_type, row.target_id, row.assigned_user_id_snapshot)
        for row in existing
    }
    created = False

    for assignment in assignments:
        key = (assignment.target_type, assignment.target_id, assignment.user_id)
        if key in existing_keys:
            continue
        db.add(ReportTargetSnapshot(
            report_id=report.id,
            category_id=assignment.category_id,
            category_name=assignment.category_name,
            subcategory_id=assignment.subcategory_id,
            subcategory_name=assignment.subcategory_name,
            target_type=assignment.target_type,
            target_id=assignment.target_id,
            target_name=assignment.target_name,
            assigned_user_id_snapshot=assignment.user_id,
            assigned_user_name_snapshot=assignment.user_full_name_snapshot,
        ))
        existing_keys.add(key)
        created = True

    for row in results:
        if row.checked_by_user_id is None and not row.checked_by_name_snapshot:
            continue
        key = (row.target_type, row.target_id, row.checked_by_user_id)
        if key in existing_keys:
            continue
        db.add(ReportTargetSnapshot(
            report_id=report.id,
            category_id=row.category_id,
            category_name=row.category_name,
            subcategory_id=row.subcategory_id,
            subcategory_name=row.subcategory_name,
            target_type=row.target_type,
            target_id=row.target_id,
            target_name=row.target_name,
            assigned_user_id_snapshot=row.checked_by_user_id,
            assigned_user_name_snapshot=row.checked_by_name_snapshot,
        ))
        existing_keys.add(key)
        created = True

    if created:
        await db.commit()
        return await _load_report_target_snapshots(report.id, db)
    return existing


def _build_retained_scope_for_user(
    user_id: int,
    snapshots: list[ReportTargetSnapshot],
    results: list[CheckResult],
) -> tuple[set[str], dict[str, set[str]], dict[str, dict[str, set[str]]]]:
    retained_category_ids: set[str] = set()
    retained_subcategory_ids: dict[str, set[str]] = defaultdict(set)
    retained_item_ids: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))

    for row in snapshots:
        if row.assigned_user_id_snapshot != user_id:
            continue
        if row.target_type == 'category':
            retained_category_ids.add(row.category_id)
        elif row.target_type == 'subcategory' and row.subcategory_id:
            retained_subcategory_ids[row.category_id].add(row.subcategory_id)
        elif row.target_type == 'item' and row.subcategory_id:
            retained_item_ids[row.category_id][row.subcategory_id].add(row.target_id)

    for row in results:
        if row.checked_by_user_id != user_id:
            continue
        if row.target_type == 'category':
            retained_category_ids.add(row.category_id)
        elif row.target_type == 'subcategory' and row.subcategory_id:
            retained_subcategory_ids[row.category_id].add(row.subcategory_id)
        elif row.target_type == 'item' and row.subcategory_id:
            retained_item_ids[row.category_id][row.subcategory_id].add(row.target_id)

    return retained_category_ids, retained_subcategory_ids, retained_item_ids


def _report_snapshot_maps(
    snapshots: list[ReportTargetSnapshot],
) -> tuple[
    dict[str, set[int]],
    dict[str, dict[str, set[int]]],
    dict[str, dict[str, dict[str, set[int]]]],
    dict[str, str],
    dict[str, dict[str, str]],
    dict[str, dict[str, dict[str, str]]],
]:
    category_user_ids: dict[str, set[int]] = defaultdict(set)
    subcategory_user_ids: dict[str, dict[str, set[int]]] = defaultdict(lambda: defaultdict(set))
    item_user_ids: dict[str, dict[str, dict[str, set[int]]]] = defaultdict(lambda: defaultdict(lambda: defaultdict(set)))
    category_owner_names: dict[str, str] = {}
    subcategory_owner_names: dict[str, dict[str, str]] = defaultdict(dict)
    item_owner_names: dict[str, dict[str, dict[str, str]]] = defaultdict(lambda: defaultdict(dict))

    for row in snapshots:
        if row.target_type == 'category':
            if row.assigned_user_id_snapshot is not None:
                category_user_ids[row.category_id].add(int(row.assigned_user_id_snapshot))
            if row.assigned_user_name_snapshot and row.category_id not in category_owner_names:
                category_owner_names[row.category_id] = row.assigned_user_name_snapshot
        elif row.target_type == 'subcategory' and row.subcategory_id:
            if row.assigned_user_id_snapshot is not None:
                subcategory_user_ids[row.category_id][row.subcategory_id].add(int(row.assigned_user_id_snapshot))
            if row.assigned_user_name_snapshot and row.subcategory_id not in subcategory_owner_names[row.category_id]:
                subcategory_owner_names[row.category_id][row.subcategory_id] = row.assigned_user_name_snapshot
        elif row.target_type == 'item' and row.subcategory_id:
            if row.assigned_user_id_snapshot is not None:
                item_user_ids[row.category_id][row.subcategory_id][row.target_id].add(int(row.assigned_user_id_snapshot))
            if row.assigned_user_name_snapshot and row.target_id not in item_owner_names[row.category_id][row.subcategory_id]:
                item_owner_names[row.category_id][row.subcategory_id][row.target_id] = row.assigned_user_name_snapshot

    return (
        category_user_ids,
        subcategory_user_ids,
        item_user_ids,
        category_owner_names,
        subcategory_owner_names,
        item_owner_names,
    )

def _subcategory_is_complete(raw_subcategory: dict[str, Any], results_by_target: dict[str, CheckResult]) -> tuple[bool, StatusEnum]:
    sub_row = results_by_target.get(raw_subcategory['id'])
    item_rows = [results_by_target.get(item['id']) for item in raw_subcategory['items']]
    final_item_rows = [row for row in item_rows if row and row.status in {'green', 'red'}]
    has_red_items = any(row and row.status == 'red' for row in item_rows)

    if sub_row and sub_row.status == 'green':
        return True, StatusEnum.GREEN

    if sub_row and sub_row.status == 'orange':
        if len(final_item_rows) == len(raw_subcategory['items']):
            return True, StatusEnum.RED if has_red_items else StatusEnum.GREEN
        return False, StatusEnum.ORANGE

    return False, StatusEnum.GREY


def _category_is_complete(raw_category: dict[str, Any], results_by_target: dict[str, CheckResult]) -> bool:
    relevant_subcategories = list(raw_category['subcategories'])
    return bool(relevant_subcategories) and all(_subcategory_is_complete(sub, results_by_target)[0] for sub in relevant_subcategories)


async def _find_available_report_date(location: str, preferred_date: date, db: AsyncSession) -> date:
    used_dates = set(
        await db.scalars(
            select(Report.report_date).where(Report.location == location)
        )
    )
    candidate = preferred_date
    while candidate in used_dates:
        candidate -= timedelta(days=1)
    return candidate


async def _ensure_cycle_final_report(location: str, cycle_version: int, cycle_started_at: date, db: AsyncSession) -> Report | None:
    normalized = _normalize_location(location)
    existing = await db.scalar(
        select(Report).where(
            Report.location == normalized,
            Report.cycle_version == cycle_version,
            Report.report_type == FINAL_REPORT_TYPE,
        ).limit(1)
    )
    if existing:
        if existing.status != 'completed':
            existing.status = 'completed'
            await db.commit()
            await db.refresh(existing)
        return existing

    daily_reports = (
        await db.scalars(
            select(Report).where(
                Report.location == normalized,
                Report.cycle_version == cycle_version,
                Report.report_type == DAILY_REPORT_TYPE,
            )
        )
    ).all()
    if not daily_reports:
        return None

    report_date = await _find_available_report_date(normalized, cycle_started_at - timedelta(days=1), db)
    final_report = Report(
        location=normalized,
        report_date=report_date,
        cycle_version=cycle_version,
        report_type=FINAL_REPORT_TYPE,
        status='completed',
    )
    db.add(final_report)
    await db.commit()
    await db.refresh(final_report)
    return final_report


async def _get_or_create_selection_cycle(location: str, db: AsyncSession) -> SelectionCycle:
    normalized = _normalize_location(location)
    cycle = await db.scalar(select(SelectionCycle).where(SelectionCycle.location == normalized).limit(1))
    today = get_moscow_today()
    expected_cycle_start, _ = _cycle_bounds_for_date(today)

    if not cycle:
        cycle = SelectionCycle(location=normalized, cycle_version=1, started_at=expected_cycle_start)
        db.add(cycle)
        await db.commit()
        await db.refresh(cycle)
        return cycle

    stored_cycle_start = cycle.started_at or expected_cycle_start
    stored_order = _cycle_order_value(stored_cycle_start)
    current_order = _cycle_order_value(today)

    if stored_order < current_order:
        old_version = cycle.cycle_version
        old_started_at = stored_cycle_start
        await _ensure_cycle_final_report(normalized, old_version, old_started_at, db)
        cycle.cycle_version += max(1, current_order - stored_order)
        cycle.started_at = expected_cycle_start
        cycle.updated_at = datetime.utcnow()
        await db.execute(delete(CategoryAssignment).where(CategoryAssignment.location == normalized, CategoryAssignment.cycle_version == old_version))
        await db.commit()
        await db.refresh(cycle)
        return cycle

    if cycle.started_at != expected_cycle_start:
        cycle.started_at = expected_cycle_start
        cycle.updated_at = datetime.utcnow()
        await db.commit()
        await db.refresh(cycle)

    return cycle


async def reset_selection_cycle(location: str, db: AsyncSession) -> ResetSelectionCycleResponse:
    normalized = _normalize_location(location)
    cycle = await _get_or_create_selection_cycle(normalized, db)
    old_version = cycle.cycle_version
    old_started_at = cycle.started_at
    await _ensure_cycle_final_report(normalized, old_version, old_started_at, db)
    expected_cycle_start, _ = _cycle_bounds_for_date(get_moscow_today())
    cycle.cycle_version += 1
    cycle.started_at = expected_cycle_start
    cycle.updated_at = datetime.utcnow()
    await db.execute(delete(CategoryAssignment).where(CategoryAssignment.location == normalized, CategoryAssignment.cycle_version == old_version))
    await db.commit()
    await db.refresh(cycle)
    return ResetSelectionCycleResponse(
        success=True,
        message='Выбор категорий и подкатегорий обновлён. Начался новый цикл месяца.',
        cycle_version=cycle.cycle_version,
        cycle_started_at=cycle.started_at.strftime('%d.%m.%Y'),
    )



async def _sync_report_status(report: Report, db: AsyncSession) -> None:
    if (report.report_type or DAILY_REPORT_TYPE) == FINAL_REPORT_TYPE:
        report.status = 'completed'
        return

    participant_user_ids = await _get_report_participant_user_ids(report.id, db)
    completion_count = await db.scalar(
        select(func.count()).select_from(ReportEmployeeCompletion).where(ReportEmployeeCompletion.report_id == report.id)
    )
    required_finish_count = len(participant_user_ids)

    if required_finish_count > 0 and (completion_count or 0) >= required_finish_count:
        report.status = 'completed'
    elif not participant_user_ids:
        report.status = 'created'
    else:
        report.status = 'in_progress'


async def get_or_create_daily_report(location: str, cycle_version: int, db: AsyncSession) -> Report:
    today = get_moscow_today()
    normalized = _normalize_location(location)

    report = await db.scalar(
        select(Report).where(
            Report.location == normalized,
            Report.report_date == today,
            Report.report_type == DAILY_REPORT_TYPE,
        ).limit(1)
    )
    if report:
        await _sync_report_status(report, db)
        await db.commit()
        return report

    report = Report(location=normalized, report_date=today, cycle_version=cycle_version, report_type=DAILY_REPORT_TYPE, status='created')
    db.add(report)
    await db.execute(
        delete(CategoryAssignment).where(
            CategoryAssignment.location == normalized,
            CategoryAssignment.cycle_version == cycle_version,
        )
    )
    await db.commit()
    await db.refresh(report)
    return report


async def _load_assignments(location: str, cycle_version: int, db: AsyncSession) -> list[CategoryAssignment]:
    return (await db.scalars(select(CategoryAssignment).where(CategoryAssignment.location == location, CategoryAssignment.cycle_version == cycle_version))).all()


async def _load_results(report_id: int, db: AsyncSession) -> list[CheckResult]:
    return (await db.scalars(select(CheckResult).where(CheckResult.report_id == report_id).order_by(CheckResult.id.asc()))).all()


async def _load_results_for_report_ids(report_ids: list[int], db: AsyncSession) -> list[CheckResult]:
    if not report_ids:
        return []
    return (
        await db.scalars(
            select(CheckResult)
            .where(CheckResult.report_id.in_(report_ids))
            .order_by(CheckResult.created_at.asc(), CheckResult.id.asc())
        )
    ).all()


async def _load_report_target_snapshots_for_report_ids(report_ids: list[int], db: AsyncSession) -> list[ReportTargetSnapshot]:
    if not report_ids:
        return []
    return (
        await db.scalars(
            select(ReportTargetSnapshot)
            .where(ReportTargetSnapshot.report_id.in_(report_ids))
            .order_by(ReportTargetSnapshot.created_at.asc(), ReportTargetSnapshot.id.asc())
        )
    ).all()


async def _load_selection_targets(location: str, cycle_version: int, db: AsyncSession) -> list[SelectionTarget]:
    return (await db.scalars(select(SelectionTarget).where(SelectionTarget.location == location, SelectionTarget.cycle_version == cycle_version))).all()


async def _load_selection_targets_for_date(location: str, cycle_version: int, target_date: date, db: AsyncSession) -> list[SelectionTargetDay]:
    return (
        await db.scalars(
            select(SelectionTargetDay)
            .where(SelectionTargetDay.location == location)
            .where(SelectionTargetDay.cycle_version == cycle_version)
            .where(SelectionTargetDay.target_date == target_date)
        )
    ).all()


async def _resolve_selection_targets_for_date(location: str, cycle_version: int, target_date: date, db: AsyncSession) -> list[SelectionTarget | SelectionTargetDay]:
    dated_targets = await _load_selection_targets_for_date(location, cycle_version, target_date, db)
    if dated_targets:
        return dated_targets
    return await _load_selection_targets(location, cycle_version, db)


async def _resolve_previous_selection_targets_for_date(
    location: str,
    cycle_version: int,
    target_date: date,
    db: AsyncSession,
) -> tuple[list[SelectionTarget | SelectionTargetDay], date | None]:
    previous_target_date = await db.scalar(
        select(SelectionTargetDay.target_date)
        .where(SelectionTargetDay.location == location)
        .where(SelectionTargetDay.cycle_version == cycle_version)
        .where(SelectionTargetDay.target_date < target_date)
        .order_by(SelectionTargetDay.target_date.desc())
        .limit(1)
    )
    if previous_target_date is not None:
        return await _load_selection_targets_for_date(location, cycle_version, previous_target_date, db), previous_target_date
    legacy_targets = await _load_selection_targets(location, cycle_version, db)
    return legacy_targets, None


def _cycle_date_bounds(cycle_started_at: date) -> tuple[date, date]:
    return _cycle_bounds_for_date(cycle_started_at)


def _resolve_cycle_target_date(cycle_started_at: date, requested_date: date | None) -> date:
    min_date, max_date = _cycle_date_bounds(cycle_started_at)
    candidate = requested_date or get_moscow_today()
    if candidate < min_date:
        return min_date
    if candidate > max_date:
        return max_date
    return candidate


async def _load_completed_subcategory_ids_for_cycle(
    location: str,
    cycle_version: int,
    inventory: dict[str, Any],
    db: AsyncSession,
    before_report_date: date | None = None,
) -> dict[str, set[str]]:
    report_query = select(Report).where(
        Report.location == location,
        Report.cycle_version == cycle_version,
    )
    if before_report_date is not None:
        report_query = report_query.where(Report.report_date < before_report_date)

    reports = (await db.scalars(report_query.order_by(Report.report_date.asc(), Report.id.asc()))).all()
    return await _load_completed_subcategory_ids_for_reports(reports, inventory, db)


async def _load_completed_subcategory_ids_before_date(
    location: str,
    inventory: dict[str, Any],
    db: AsyncSession,
    before_report_date: date | None = None,
) -> dict[str, set[str]]:
    report_query = select(Report).where(
        Report.location == location,
        Report.report_type == DAILY_REPORT_TYPE,
    )
    if before_report_date is not None:
        report_query = report_query.where(Report.report_date < before_report_date)

    reports = (await db.scalars(report_query.order_by(Report.report_date.asc(), Report.id.asc()))).all()
    return await _load_completed_subcategory_ids_for_reports(reports, inventory, db)


async def _load_completed_subcategory_ids_for_reports(
    reports: list[Report],
    inventory: dict[str, Any],
    db: AsyncSession,
) -> dict[str, set[str]]:
    if not reports:
        return {}

    report_ids = [report.id for report in reports]
    results = (
        await db.scalars(
            select(CheckResult)
            .where(CheckResult.report_id.in_(report_ids))
            .order_by(CheckResult.report_id.asc(), CheckResult.id.asc())
        )
    ).all()
    if not results:
        return {}

    category_by_id = {category['id']: category for category in inventory.get('categories', [])}
    results_by_report: dict[int, dict[str, dict[str, CheckResult]]] = defaultdict(lambda: defaultdict(dict))
    for row in results:
        results_by_report[row.report_id][row.category_id][row.target_id] = row

    completed_subcategory_ids: dict[str, set[str]] = defaultdict(set)
    for report in reports:
        rows_by_category = results_by_report.get(report.id, {})
        for category_id, result_map in rows_by_category.items():
            raw_category = category_by_id.get(category_id)
            if not raw_category or raw_category.get('name') == DEFAULT_CATEGORY_NAME:
                continue
            for raw_sub in raw_category.get('subcategories', []):
                if _is_categoryless_subcategory(raw_category, raw_sub):
                    continue
                is_completed, _ = _subcategory_is_complete(raw_sub, result_map)
                if is_completed:
                    completed_subcategory_ids[category_id].add(raw_sub['id'])

    return {category_id: set(sub_ids) for category_id, sub_ids in completed_subcategory_ids.items()}


def _selection_target_maps(targets: list[SelectionTarget]) -> tuple[set[str], dict[str, set[str]]]:
    category_ids: set[str] = set()
    subcategory_ids: dict[str, set[str]] = defaultdict(set)
    for row in targets:
        if row.target_type == 'category':
            category_ids.add(row.category_id)
        elif row.target_type == 'subcategory' and row.subcategory_id:
            subcategory_ids[row.category_id].add(row.subcategory_id)
    return category_ids, subcategory_ids


def _iter_real_subcategory_ids(raw_category: dict[str, Any]) -> list[str]:
    return [
        sub['id']
        for sub in raw_category.get('subcategories', [])
        if not _is_categoryless_subcategory(raw_category, sub)
    ]


async def _load_cycle_scope_taken_elsewhere(
    location: str,
    cycle_version: int,
    inventory: dict[str, Any],
    db: AsyncSession,
    exclude_target_date: date | None = None,
) -> tuple[set[str], dict[str, set[str]]]:
    reports = (
        await db.scalars(
            select(Report)
            .where(Report.location == location)
            .where(Report.cycle_version == cycle_version)
            .where(Report.report_type == DAILY_REPORT_TYPE)
            .order_by(Report.report_date.asc(), Report.id.asc())
        )
    ).all()
    dated_rows = (
        await db.scalars(
            select(SelectionTargetDay)
            .where(SelectionTargetDay.location == location)
            .where(SelectionTargetDay.cycle_version == cycle_version)
            .order_by(SelectionTargetDay.target_date.asc(), SelectionTargetDay.id.asc())
        )
    ).all()

    dated_targets_by_date: dict[date, list[SelectionTargetDay]] = defaultdict(list)
    for row in dated_rows:
        dated_targets_by_date[row.target_date].append(row)

    legacy_targets: list[SelectionTarget] | None = None
    category_lookup = {category['id']: category for category in inventory.get('categories', [])}
    occupied_category_ids: set[str] = set()
    occupied_subcategory_ids: dict[str, set[str]] = defaultdict(set)
    consumed_dates: set[date] = set()

    def absorb(targets: list[SelectionTarget | SelectionTargetDay]) -> None:
        for row in targets:
            raw_category = category_lookup.get(row.category_id)
            if not raw_category or raw_category.get('name') == DEFAULT_CATEGORY_NAME:
                continue
            if row.target_type == 'category':
                occupied_category_ids.add(row.category_id)
                occupied_subcategory_ids[row.category_id].update(_iter_real_subcategory_ids(raw_category))
            elif row.target_type == 'subcategory' and row.subcategory_id:
                occupied_subcategory_ids[row.category_id].add(row.subcategory_id)

    for report in reports:
        if exclude_target_date is not None and report.report_date == exclude_target_date:
            continue
        consumed_dates.add(report.report_date)
        report_targets = dated_targets_by_date.get(report.report_date)
        if report_targets is None:
            if legacy_targets is None:
                legacy_targets = await _load_selection_targets(location, cycle_version, db)
            report_targets = legacy_targets
        absorb(report_targets)

    for target_date, targets in dated_targets_by_date.items():
        if exclude_target_date is not None and target_date == exclude_target_date:
            continue
        if target_date in consumed_dates:
            continue
        absorb(targets)

    return occupied_category_ids, {category_id: set(sub_ids) for category_id, sub_ids in occupied_subcategory_ids.items()}


async def _get_daily_report_for_cycle_date(location: str, cycle_version: int, target_date: date, db: AsyncSession) -> Report | None:
    report = await db.scalar(
        select(Report)
        .where(Report.location == location)
        .where(Report.cycle_version == cycle_version)
        .where(Report.report_type == DAILY_REPORT_TYPE)
        .where(Report.report_date == target_date)
        .order_by(Report.id.desc())
        .limit(1)
    )
    if report is not None:
        await _sync_report_status(report, db)
    return report


def _admin_report_subcategory_stats(
    categories: list[CategoryResult],
    source_inventory_by_name: dict[str, dict[str, Any]],
    completed_before_report: dict[str, set[str]] | None = None,
    report_type: str = DAILY_REPORT_TYPE,
) -> tuple[int, int, int, int]:
    completed_before_report = completed_before_report or {}
    total_subcategories = 0
    discrepancy_subcategories = 0
    no_discrepancy_subcategories = 0

    for category in categories:
        source_category = source_inventory_by_name.get(category.name)
        scoped_names: set[str] = set()
        if category.selected_on_cycle and source_category is not None:
            excluded_ids = completed_before_report.get(source_category['id'], set()) if report_type == DAILY_REPORT_TYPE else set()
            for sub in source_category.get('subcategories', []):
                if _is_categoryless_subcategory(source_category, sub):
                    continue
                if sub['id'] in excluded_ids:
                    continue
                scoped_names.add(sub['name'])

        scoped_names.update(name for name in category.selected_subcategories if name)
        scoped_names.update(name for name in category.remaining_subcategories if name)
        scoped_names.update(item.name for item in category.completed_subcategories if item.name)
        scoped_names.update(
            item.subcategory_name
            for item in category.problem_items
            if item.subcategory_name and item.subcategory_name != '-'
        )

        discrepancy_names = {
            item.subcategory_name
            for item in category.problem_items
            if item.subcategory_name and item.subcategory_name != '-'
        }
        no_discrepancy_names = {
            item.name
            for item in category.completed_subcategories
            if item.name and item.name not in discrepancy_names
        }

        total_subcategories += len(scoped_names)
        discrepancy_subcategories += len(discrepancy_names)
        no_discrepancy_subcategories += len(no_discrepancy_names)

    completed_subcategories = discrepancy_subcategories + no_discrepancy_subcategories
    return total_subcategories, completed_subcategories, discrepancy_subcategories, no_discrepancy_subcategories



def _report_selection_scope(
    report_snapshots: list[ReportTargetSnapshot],
    results: list[CheckResult],
) -> tuple[set[str], dict[str, set[str]], list[str], list[str]]:
    category_ids: set[str] = set()
    subcategory_ids: dict[str, set[str]] = defaultdict(set)
    category_names: set[str] = set()
    subcategory_labels: set[str] = set()

    def absorb(
        target_type: str,
        category_id: str,
        category_name: str | None,
        subcategory_id: str | None,
        subcategory_name: str | None,
    ) -> None:
        if category_name == DEFAULT_CATEGORY_NAME:
            return
        if target_type == 'category':
            category_ids.add(category_id)
            if category_name:
                category_names.add(category_name)
            return
        if not subcategory_id or not subcategory_name or subcategory_name == DEFAULT_SUBCATEGORY_NAME:
            return
        subcategory_ids[category_id].add(subcategory_id)
        if category_name:
            subcategory_labels.add(f"{category_name} → {subcategory_name}")

    for row in report_snapshots:
        absorb(row.target_type, row.category_id, row.category_name, row.subcategory_id, row.subcategory_name)

    for row in results:
        absorb(row.target_type, row.category_id, row.category_name, row.subcategory_id, row.subcategory_name)

    return category_ids, {category_id: set(sub_ids) for category_id, sub_ids in subcategory_ids.items()}, sorted(category_names), sorted(subcategory_labels)


def _report_history_target_maps(
    report_snapshots: list[ReportTargetSnapshot],
    results: list[CheckResult],
) -> tuple[set[str], dict[str, set[str]], dict[str, dict[str, set[str]]]]:
    category_ids: set[str] = set()
    subcategory_ids: dict[str, set[str]] = defaultdict(set)
    item_ids: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))

    def absorb(target_type: str, category_id: str, subcategory_id: str | None, target_id: str) -> None:
        if target_type == 'category':
            category_ids.add(category_id)
        elif target_type == 'subcategory' and subcategory_id:
            subcategory_ids[category_id].add(subcategory_id)
        elif target_type == 'item' and subcategory_id:
            item_ids[category_id][subcategory_id].add(target_id)

    for row in report_snapshots:
        absorb(row.target_type, row.category_id, row.subcategory_id, row.target_id)

    for row in results:
        absorb(row.target_type, row.category_id, row.subcategory_id, row.target_id)

    return category_ids, subcategory_ids, item_ids


def _is_categoryless_subcategory(category: dict[str, Any], subcategory: dict[str, Any]) -> bool:
    return category['name'] != DEFAULT_CATEGORY_NAME and subcategory['name'] == DEFAULT_SUBCATEGORY_NAME


def _filter_inventory_by_targets(
    inventory: dict[str, Any],
    selected_category_ids: set[str],
    selected_subcategory_ids: dict[str, set[str]],
    retained_category_ids: set[str] | None = None,
    retained_subcategory_ids: dict[str, set[str]] | None = None,
    retained_item_ids: dict[str, dict[str, set[str]]] | None = None,
    excluded_subcategory_ids: dict[str, set[str]] | None = None,
) -> dict[str, Any]:
    retained_category_ids = retained_category_ids or set()
    retained_subcategory_ids = retained_subcategory_ids or {}
    retained_item_ids = retained_item_ids or {}
    excluded_subcategory_ids = excluded_subcategory_ids or {}

    has_retained_scope = bool(retained_category_ids) or any(retained_subcategory_ids.values()) or any(
        retained_item_ids.get(category_id, {}) for category_id in retained_item_ids
    )
    if not selected_category_ids and not any(selected_subcategory_ids.values()) and not has_retained_scope:
        return inventory

    categories: list[dict[str, Any]] = []
    for category in inventory['categories']:
        include_full_category = category['id'] in selected_category_ids or category['id'] in retained_category_ids
        allowed_sub_ids = set(selected_subcategory_ids.get(category['id'], set())) | set(retained_subcategory_ids.get(category['id'], set()))
        retained_items_by_sub = retained_item_ids.get(category['id'], {})

        filtered_subcategories: list[dict[str, Any]] = []
        for sub in category['subcategories']:
            retained_sub_item_ids = set(retained_items_by_sub.get(sub['id'], set()))
            keep_subcategory = False
            keep_all_items = False
            excluded_for_cycle = sub['id'] in excluded_subcategory_ids.get(category['id'], set())
            preserve_excluded_subcategory = sub['id'] in retained_subcategory_ids.get(category['id'], set()) or bool(retained_sub_item_ids)

            if excluded_for_cycle and not preserve_excluded_subcategory:
                continue

            if include_full_category or sub['id'] in allowed_sub_ids:
                keep_subcategory = True
                keep_all_items = True
            elif retained_sub_item_ids:
                keep_subcategory = True

            if not keep_subcategory:
                continue

            items = [dict(item) for item in sub['items']]
            if not keep_all_items:
                items = [item for item in items if item['id'] in retained_sub_item_ids]
                if not items:
                    continue

            filtered_subcategories.append({
                'id': sub['id'],
                'name': sub['name'],
                'items': items,
            })

        if filtered_subcategories:
            categories.append({
                'id': category['id'],
                'name': category['name'],
                'subcategories': filtered_subcategories,
            })

    return {'location': inventory.get('location'), 'categories': categories}


async def list_locations(db: AsyncSession, current_user: User) -> LocationListResponse:
    await _ensure_default_location_points(db)
    if current_user.role == RoleEnum.SUPERADMIN.value:
        rows = (await db.scalars(select(LocationPoint).order_by(LocationPoint.name.asc()))).all()
    elif current_user.role == RoleEnum.ADMIN.value:
        rows = (
            await db.scalars(
                select(LocationPoint)
                .join(AdminLocationAccess, AdminLocationAccess.location_point_id == LocationPoint.id)
                .where(AdminLocationAccess.admin_user_id == current_user.id)
                .order_by(LocationPoint.name.asc())
            )
        ).all()
    else:
        rows = []
    locations = [LocationPointModel.model_validate(row) for row in rows]
    return LocationListResponse(locations=locations)


async def list_moysklad_stores_by_token(ms_token: str) -> StoreListResponse:
    headers = {
        'Authorization': f'Bearer {ms_token}',
        'Accept-Encoding': 'gzip',
        'Content-Type': 'application/json',
    }
    timeout = httpx.Timeout(20.0, connect=10.0)
    async with httpx.AsyncClient(timeout=timeout) as client:
        response = await client.get(f"{settings.ms_api_base_url.rstrip('/')}/entity/store", headers=headers)
        response.raise_for_status()
        data = response.json()
    stores = [
        StoreOption(id=row.get('id'), name=row.get('name') or 'Без названия')
        for row in (data.get('rows') or []) if row.get('id')
    ]
    stores.sort(key=lambda item: item.name.lower())
    return StoreListResponse(stores=stores)


async def create_location_point(payload: CreateLocationRequest, db: AsyncSession) -> CreateLocationResponse:
    name = _normalize_location(payload.name)
    existing = await db.scalar(select(LocationPoint).where(LocationPoint.name == name).limit(1))
    if existing:
        raise HTTPException(status_code=400, detail='Точка с таким названием уже существует.')

    point = LocationPoint(
        name=name,
        ms_token=payload.ms_token.strip(),
        ms_store_id=payload.ms_store_id.strip(),
        ms_store_name=payload.ms_store_name.strip(),
    )
    db.add(point)
    await db.commit()
    await db.refresh(point)
    ms_client.invalidate_inventory(point.name)
    _invalidate_runtime_inventory_cache(point.name)
    return CreateLocationResponse(success=True, message='Точка создана.', location=LocationPointModel.model_validate(point))


async def update_location_point(location_id: int, payload: UpdateLocationRequest, db: AsyncSession) -> UpdateLocationResponse:
    point = await db.get(LocationPoint, location_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка не найдена.')

    name = _normalize_location(payload.name)
    duplicate = await db.scalar(select(LocationPoint).where(LocationPoint.name == name, LocationPoint.id != location_id).limit(1))
    if duplicate:
        raise HTTPException(status_code=400, detail='Точка с таким названием уже существует.')

    old_name = point.name
    point.name = name
    point.ms_token = payload.ms_token.strip()
    point.ms_store_id = payload.ms_store_id.strip()
    point.ms_store_name = payload.ms_store_name.strip()

    if old_name != point.name:
        await db.execute(update(User).where(User.location == old_name).values(location=point.name))
        await db.execute(update(Report).where(Report.location == old_name).values(location=point.name))
        await db.execute(update(SelectionCycle).where(SelectionCycle.location == old_name).values(location=point.name))
        await db.execute(update(SelectionTarget).where(SelectionTarget.location == old_name).values(location=point.name))
        await db.execute(update(SelectionTargetDay).where(SelectionTargetDay.location == old_name).values(location=point.name))
        await db.execute(update(CategoryAssignment).where(CategoryAssignment.location == old_name).values(location=point.name))

    await db.commit()
    await db.refresh(point)
    ms_client.invalidate_inventory(old_name)
    ms_client.invalidate_inventory(point.name)
    _invalidate_runtime_inventory_cache(old_name)
    _invalidate_runtime_inventory_cache(point.name)
    return UpdateLocationResponse(success=True, message='Точка обновлена.', location=LocationPointModel.model_validate(point))


async def delete_location_point(location_id: int, db: AsyncSession) -> DeleteResponse:
    point = await db.get(LocationPoint, location_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка не найдена.')

    linked_entities: list[str] = []
    checks = [
        ('пользователи', User),
        ('ревизии', Report),
        ('циклы выбора', SelectionCycle),
        ('выбор категорий цикла', SelectionTarget),
        ('выбор подкатегорий по дням', SelectionTargetDay),
        ('закрепления сотрудников', CategoryAssignment),
    ]
    for label, model in checks:
        count = await db.scalar(select(func.count()).select_from(model).where(model.location == point.name))
        if (count or 0) > 0:
            linked_entities.append(label)

    access_count = await db.scalar(select(func.count()).select_from(AdminLocationAccess).where(AdminLocationAccess.location_point_id == point.id))
    if (access_count or 0) > 0:
        linked_entities.append('доступы управляющих')

    if linked_entities:
        raise HTTPException(
            status_code=400,
            detail='Нельзя удалить точку, пока с ней связаны: ' + ', '.join(linked_entities) + '.',
        )

    old_name = point.name
    await db.delete(point)
    await db.commit()
    ms_client.invalidate_inventory(old_name)
    _invalidate_runtime_inventory_cache(old_name)
    return DeleteResponse(success=True, message='Точка удалена.')


async def get_cycle_targets(location: str, db: AsyncSession, target_date: date | None = None) -> AdminCycleTargetsResponse:
    normalized = _normalize_location(location)
    cycle_context = await _resolve_cycle_context(normalized, db, target_date)
    resolved_target_date = cycle_context.target_date
    inventory = await _get_inventory_for(normalized, db=db)
    targets = await _resolve_selection_targets_for_date(normalized, cycle_context.cycle_version, resolved_target_date, db)
    previous_targets, previous_target_date = await _resolve_previous_selection_targets_for_date(
        normalized,
        cycle_context.cycle_version,
        resolved_target_date,
        db,
    )
    selected_category_ids, selected_subcategory_ids = _selection_target_maps(targets)
    previous_category_ids, previous_subcategory_ids = _selection_target_maps(previous_targets)
    completed_subcategory_ids = await _load_completed_subcategory_ids_for_cycle(
        normalized,
        cycle_context.cycle_version,
        inventory,
        db,
        before_report_date=resolved_target_date,
    )
    occupied_category_ids, occupied_subcategory_ids = await _load_cycle_scope_taken_elsewhere(
        normalized,
        cycle_context.cycle_version,
        inventory,
        db,
        exclude_target_date=resolved_target_date,
    )
    target_report = await _get_daily_report_for_cycle_date(normalized, cycle_context.cycle_version, resolved_target_date, db)
    today = get_moscow_today()
    is_past_date_locked = resolved_target_date < today
    is_completed_report_locked = bool(target_report and target_report.status == 'completed')
    is_locked = bool(is_past_date_locked or is_completed_report_locked)
    if is_past_date_locked:
        locked_message = (
            f'Дата {resolved_target_date.strftime("%d.%m.%Y")} уже прошла. Для прошлых дат выбор доступен только для просмотра.'
        )
    elif is_completed_report_locked:
        locked_message = (
            f'Ревизия за {resolved_target_date.strftime("%d.%m.%Y")} уже завершена. Для неё можно только посмотреть выбранные подкатегории.'
        )
    else:
        locked_message = None

    visible_previous_category_ids: set[str] = set()
    visible_previous_subcategory_ids: set[str] = set()
    categories: list[AdminCycleTargetCategory] = []
    for category in inventory['categories']:
        if category['name'] == DEFAULT_CATEGORY_NAME:
            continue

        category_id = category['id']
        selected_whole_category = category_id in selected_category_ids
        current_selected_sub_ids = selected_subcategory_ids.get(category_id, set())
        completed_ids_for_category = completed_subcategory_ids.get(category_id, set())
        occupied_ids_for_category = occupied_subcategory_ids.get(category_id, set())

        has_other_taken_remaining = any(
            sub['id'] not in completed_ids_for_category and sub['id'] in occupied_ids_for_category and sub['id'] not in current_selected_sub_ids
            for sub in category['subcategories']
            if not _is_categoryless_subcategory(category, sub)
        )
        whole_category_taken_elsewhere = category_id in occupied_category_ids and not selected_whole_category
        category_disabled = is_locked or (not selected_whole_category and (whole_category_taken_elsewhere or has_other_taken_remaining))

        subcategories: list[AdminCycleTargetItem] = []
        completed_subcategories: list[AdminCycleTargetItem] = []
        for sub in category['subcategories']:
            if _is_categoryless_subcategory(category, sub):
                continue

            is_selected = sub['id'] in current_selected_sub_ids
            if sub['id'] in completed_ids_for_category:
                completed_subcategories.append(AdminCycleTargetItem(
                    id=sub['id'],
                    name=sub['name'],
                    selected=is_selected,
                    disabled=True,
                ))
                continue

            occupied_elsewhere = sub['id'] in occupied_ids_for_category and not is_selected
            if occupied_elsewhere:
                continue

            subcategories.append(AdminCycleTargetItem(
                id=sub['id'],
                name=sub['name'],
                selected=is_selected,
                disabled=is_locked,
            ))

        if category_id in previous_category_ids and (selected_whole_category or not category_disabled):
            visible_previous_category_ids.add(category_id)
        for sub in subcategories:
            if sub.id in previous_subcategory_ids.get(category_id, set()):
                visible_previous_subcategory_ids.add(sub.id)

        if not selected_whole_category and not subcategories and not completed_subcategories:
            continue

        categories.append(AdminCycleTargetCategory(
            id=category_id,
            name=category['name'],
            selected=selected_whole_category,
            disabled=category_disabled,
            subcategories=subcategories,
            completed_subcategories=completed_subcategories,
        ))

    min_target_date, max_target_date = _cycle_date_bounds(cycle_context.started_at)
    return AdminCycleTargetsResponse(
        location=normalized,
        cycle_version=cycle_context.cycle_version,
        cycle_started_at=cycle_context.started_at.strftime('%d.%m.%Y'),
        target_date=resolved_target_date.isoformat(),
        min_target_date=min_target_date.isoformat(),
        max_target_date=max_target_date.isoformat(),
        previous_target_date=previous_target_date.isoformat() if previous_target_date else None,
        previous_category_ids=sorted(visible_previous_category_ids),
        previous_subcategory_ids=sorted(visible_previous_subcategory_ids),
        is_locked=is_locked,
        locked_message=locked_message,
        report_status=_report_status_label(target_report.status) if target_report is not None else None,
        categories=categories,
    )


async def save_cycle_targets(payload: SaveCycleTargetsRequest, db: AsyncSession) -> SaveCycleTargetsResponse:
    normalized = _normalize_location(payload.location)
    cycle = await _get_or_create_selection_cycle(normalized, db)

    if payload.cycle_started_at:
        cycle.started_at = payload.cycle_started_at
        cycle.updated_at = datetime.utcnow()

    target_date = _resolve_cycle_target_date(cycle.started_at, payload.target_date)
    min_target_date, max_target_date = _cycle_date_bounds(cycle.started_at)
    if target_date < min_target_date or target_date > max_target_date:
        raise HTTPException(status_code=400, detail='Дата выбора должна быть в пределах текущего полумесячного цикла.')
    if target_date < get_moscow_today():
        raise HTTPException(status_code=400, detail='Прошлые даты доступны только для просмотра. Менять категории можно только на текущую дату и вперёд.')

    target_report = await _get_daily_report_for_cycle_date(normalized, cycle.cycle_version, target_date, db)
    if target_report is not None and target_report.status == 'completed':
        raise HTTPException(status_code=400, detail=f'Ревизия за {target_date.strftime("%d.%m.%Y")} уже завершена, менять её подкатегории нельзя.')

    requested_category_ids = sorted(set(payload.category_ids))
    requested_subcategory_ids = sorted(set(payload.subcategory_ids))
    existing_targets = await _resolve_selection_targets_for_date(normalized, cycle.cycle_version, target_date, db)

    if not requested_category_ids and not requested_subcategory_ids:
        await db.commit()
        if existing_targets:
            return SaveCycleTargetsResponse(
                success=True,
                message='Пустой выбор не сохранён. Оставлен предыдущий выбор для выбранной даты.',
                cycle_started_at=cycle.started_at.strftime('%d.%m.%Y'),
                target_date=target_date.isoformat(),
            )
        return SaveCycleTargetsResponse(
            success=True,
            message='Нечего сохранять: категории и подкатегории не выбраны.',
            cycle_started_at=cycle.started_at.strftime('%d.%m.%Y'),
            target_date=target_date.isoformat(),
        )

    inventory = await _get_inventory_for(normalized, db=db)

    completed_subcategory_ids = await _load_completed_subcategory_ids_for_cycle(
        normalized,
        cycle.cycle_version,
        inventory,
        db,
        before_report_date=target_date,
    )
    occupied_category_ids, occupied_subcategory_ids = await _load_cycle_scope_taken_elsewhere(
        normalized,
        cycle.cycle_version,
        inventory,
        db,
        exclude_target_date=target_date,
    )

    category_by_id = {row['id']: row for row in inventory['categories']}
    sub_by_id: dict[str, tuple[dict[str, Any], dict[str, Any]]] = {}
    for category in inventory['categories']:
        for sub in category['subcategories']:
            sub_by_id[sub['id']] = (category, sub)

    await db.execute(
        delete(SelectionTargetDay)
        .where(SelectionTargetDay.location == normalized)
        .where(SelectionTargetDay.cycle_version == cycle.cycle_version)
        .where(SelectionTargetDay.target_date == target_date)
    )

    skipped_completed_subcategories = 0
    skipped_busy_subcategories = 0
    skipped_busy_categories = 0

    for category_id in requested_category_ids:
        category = category_by_id.get(category_id)
        if not category or category['name'] == DEFAULT_CATEGORY_NAME:
            continue
        remaining_sub_ids = [
            sub['id']
            for sub in category['subcategories']
            if not _is_categoryless_subcategory(category, sub) and sub['id'] not in completed_subcategory_ids.get(category_id, set())
        ]
        if category_id in occupied_category_ids or any(sub_id in occupied_subcategory_ids.get(category_id, set()) for sub_id in remaining_sub_ids):
            skipped_busy_categories += 1
            continue
        db.add(SelectionTargetDay(
            location=normalized,
            cycle_version=cycle.cycle_version,
            target_date=target_date,
            category_id=category_id,
            category_name=category['name'],
            subcategory_id=None,
            subcategory_name=None,
            target_type='category',
            target_id=category_id,
            target_name=category['name'],
        ))

    selected_categories = set(requested_category_ids)
    for subcategory_id in requested_subcategory_ids:
        pair = sub_by_id.get(subcategory_id)
        if not pair:
            continue
        category, sub = pair
        if category['id'] in selected_categories or category['name'] == DEFAULT_CATEGORY_NAME or _is_categoryless_subcategory(category, sub):
            continue
        if sub['id'] in completed_subcategory_ids.get(category['id'], set()):
            skipped_completed_subcategories += 1
            continue
        if category['id'] in occupied_category_ids or sub['id'] in occupied_subcategory_ids.get(category['id'], set()):
            skipped_busy_subcategories += 1
            continue
        db.add(SelectionTargetDay(
            location=normalized,
            cycle_version=cycle.cycle_version,
            target_date=target_date,
            category_id=category['id'],
            category_name=category['name'],
            subcategory_id=sub['id'],
            subcategory_name=sub['name'],
            target_type='subcategory',
            target_id=sub['id'],
            target_name=sub['name'],
        ))

    await db.commit()
    _invalidate_runtime_inventory_cache(normalized)
    message_parts = ['Изменения сохранены.']
    if skipped_completed_subcategories:
        message_parts.append(f'Уже пройденных подкатегорий до выбранной даты пропущено: {skipped_completed_subcategories}.')
    if skipped_busy_categories:
        message_parts.append(f'Категорий, уже занятых в других ревизиях цикла, пропущено: {skipped_busy_categories}.')
    if skipped_busy_subcategories:
        message_parts.append(f'Подкатегорий, уже занятых в других ревизиях цикла, пропущено: {skipped_busy_subcategories}.')
    return SaveCycleTargetsResponse(
        success=True,
        message=' '.join(message_parts),
        cycle_started_at=cycle.started_at.strftime('%d.%m.%Y'),
        target_date=target_date.isoformat(),
    )


def _category_assignments_map(assignments: list[CategoryAssignment]) -> tuple[
    dict[str, CategoryAssignment],
    dict[str, dict[str, CategoryAssignment]],
    dict[str, dict[str, dict[str, CategoryAssignment]]],
]:
    category_map: dict[str, CategoryAssignment] = {}
    subcategory_map: dict[str, dict[str, CategoryAssignment]] = defaultdict(dict)
    item_map: dict[str, dict[str, dict[str, CategoryAssignment]]] = defaultdict(lambda: defaultdict(dict))
    for assignment in assignments:
        if assignment.target_type == 'category':
            category_map[assignment.category_id] = assignment
        elif assignment.target_type == 'subcategory' and assignment.subcategory_id:
            subcategory_map[assignment.category_id][assignment.subcategory_id] = assignment
        elif assignment.target_type == 'item' and assignment.subcategory_id:
            item_map[assignment.category_id][assignment.subcategory_id][assignment.target_id] = assignment
    return category_map, subcategory_map, item_map


def _subcategories_user_can_work(raw_category: dict[str, Any], category_assignment: CategoryAssignment | None, sub_assignments: dict[str, CategoryAssignment], user_id: int) -> list[str]:
    if category_assignment and category_assignment.user_id == user_id:
        return [sub['id'] for sub in raw_category['subcategories']]
    return [sub['id'] for sub in raw_category['subcategories'] if sub_assignments.get(sub['id']) and sub_assignments[sub['id']].user_id == user_id]


async def _load_report_employee_completions(report_id: int, db: AsyncSession) -> list[ReportEmployeeCompletion]:
    return (
        await db.scalars(
            select(ReportEmployeeCompletion)
            .where(ReportEmployeeCompletion.report_id == report_id)
            .order_by(ReportEmployeeCompletion.finished_at.asc(), ReportEmployeeCompletion.id.asc())
        )
    ).all()


async def _load_report_employee_starts(report_id: int, db: AsyncSession) -> list[ReportEmployeeStart]:
    return (
        await db.scalars(
            select(ReportEmployeeStart)
            .where(ReportEmployeeStart.report_id == report_id)
            .order_by(ReportEmployeeStart.started_at.asc(), ReportEmployeeStart.id.asc())
        )
    ).all()


async def _load_report_employee_completions_for_report_ids(report_ids: list[int], db: AsyncSession) -> list[ReportEmployeeCompletion]:
    if not report_ids:
        return []
    return (
        await db.scalars(
            select(ReportEmployeeCompletion)
            .where(ReportEmployeeCompletion.report_id.in_(report_ids))
            .order_by(ReportEmployeeCompletion.finished_at.asc(), ReportEmployeeCompletion.id.asc())
        )
    ).all()


async def _load_report_employee_starts_for_report_ids(report_ids: list[int], db: AsyncSession) -> list[ReportEmployeeStart]:
    if not report_ids:
        return []
    return (
        await db.scalars(
            select(ReportEmployeeStart)
            .where(ReportEmployeeStart.report_id.in_(report_ids))
            .order_by(ReportEmployeeStart.started_at.asc(), ReportEmployeeStart.id.asc())
        )
    ).all()


def _aggregate_employee_activity(
    starts: list[ReportEmployeeStart],
    completions: list[ReportEmployeeCompletion],
    results: list[CheckResult],
) -> tuple[dict[int, datetime], dict[int, datetime], dict[str, datetime], dict[str, datetime]]:
    first_by_user_id: dict[int, datetime] = {}
    last_by_user_id: dict[int, datetime] = {}
    first_by_name: dict[str, datetime] = {}
    last_by_name: dict[str, datetime] = {}

    def absorb(user_id: int | None, full_name: str | None, value: datetime | None) -> None:
        if value is None:
            return
        if user_id is not None:
            current_first = first_by_user_id.get(int(user_id))
            current_last = last_by_user_id.get(int(user_id))
            if current_first is None or value < current_first:
                first_by_user_id[int(user_id)] = value
            if current_last is None or value > current_last:
                last_by_user_id[int(user_id)] = value
        if full_name:
            current_first_name = first_by_name.get(full_name)
            current_last_name = last_by_name.get(full_name)
            if current_first_name is None or value < current_first_name:
                first_by_name[full_name] = value
            if current_last_name is None or value > current_last_name:
                last_by_name[full_name] = value

    for row in starts:
        absorb(row.user_id, row.user_full_name_snapshot, row.started_at)
    for row in completions:
        absorb(row.user_id, row.user_full_name_snapshot, row.finished_at)
    for row in results:
        absorb(row.checked_by_user_id, row.checked_by_name_snapshot, row.created_at)

    return first_by_user_id, last_by_user_id, first_by_name, last_by_name


async def _is_employee_started_report(report_id: int, user_id: int, db: AsyncSession) -> bool:
    started = await db.scalar(
        select(ReportEmployeeStart)
        .where(ReportEmployeeStart.report_id == report_id)
        .where(ReportEmployeeStart.user_id == user_id)
        .limit(1)
    )
    return started is not None


async def _get_report_participant_user_ids(report_id: int, db: AsyncSession) -> set[int]:
    started_ids = {
        int(value)
        for value in (
            await db.scalars(
                select(ReportEmployeeStart.user_id).where(ReportEmployeeStart.report_id == report_id)
            )
        ).all()
        if value is not None
    }
    completion_ids = {
        int(value)
        for value in (
            await db.scalars(
                select(ReportEmployeeCompletion.user_id).where(ReportEmployeeCompletion.report_id == report_id)
            )
        ).all()
        if value is not None
    }
    result_user_ids = {
        int(value)
        for value in (
            await db.scalars(
                select(CheckResult.checked_by_user_id)
                .where(CheckResult.report_id == report_id)
                .where(CheckResult.checked_by_user_id.is_not(None))
            )
        ).all()
        if value is not None
    }
    return started_ids | completion_ids | result_user_ids


def _format_completion_datetime(value: datetime | None) -> str | None:
    if not value:
        return None
    return _format_moscow_datetime(value)


async def _active_employee_users_for_location(location: str, db: AsyncSession) -> list[User]:
    return (
        await db.scalars(
            select(User)
            .where(User.role == RoleEnum.EMPLOYEE.value)
            .where(User.is_active.is_(True))
            .where(User.location == location)
            .order_by(User.full_name.asc(), User.id.asc())
        )
    ).all()


async def _is_employee_finished_report(report_id: int, user_id: int, db: AsyncSession) -> bool:
    completion = await db.scalar(
        select(ReportEmployeeCompletion)
        .where(ReportEmployeeCompletion.report_id == report_id)
        .where(ReportEmployeeCompletion.user_id == user_id)
        .limit(1)
    )
    return completion is not None


async def _refresh_report_status(report: Report, db: AsyncSession) -> None:
    await _sync_report_status(report, db)
    await db.commit()


def _subcategory_belongs_to_current_user(subcategory: SubcategoryModel) -> bool:
    if subcategory.assigned_to_current_user or subcategory.taken_as_part_of_category or subcategory.has_my_items:
        return True
    return any(item.assigned_to_current_user for item in subcategory.items)


def _subcategory_has_pending_mine_work(subcategory: SubcategoryModel) -> bool:
    if not _subcategory_belongs_to_current_user(subcategory):
        return False

    has_pending_whole_subcategory = (subcategory.assigned_to_current_user or subcategory.taken_as_part_of_category) and not subcategory.is_completed
    has_pending_diagnostic_items = any(item.assigned_to_current_user and not item.is_final for item in subcategory.items)
    return has_pending_whole_subcategory or has_pending_diagnostic_items


def _category_has_pending_mine_work(category: CategoryModel) -> bool:
    if category.assigned_to_current_user and not category.is_completed:
        return True
    return any(_subcategory_has_pending_mine_work(subcategory) for subcategory in category.subcategories)


def _category_has_free_work(category: CategoryModel) -> bool:
    if category.can_take:
        return True
    if any(subcategory.can_take for subcategory in category.subcategories):
        return True
    return any(item.can_take for subcategory in category.subcategories for item in subcategory.items)


def _build_finish_block_message(my_pending_count: int, free_pending_count: int) -> str | None:
    parts: list[str] = []
    if my_pending_count > 0:
        parts.append('в разделе «Мои» ещё есть незавершённые выборы')
    if free_pending_count > 0:
        parts.append('в разделе «Свободные» ещё остались незакреплённые категории, подкатегории или товары')
    if not parts:
        return None
    return 'Нельзя завершить ревизию: ' + ' и '.join(parts) + '.'


def _evaluate_finish_readiness(categories: list[CategoryModel]) -> tuple[bool, int, int, str | None]:
    my_pending_count = sum(1 for category in categories if _category_has_pending_mine_work(category))
    free_pending_count = sum(1 for category in categories if _category_has_free_work(category))
    can_finish = my_pending_count == 0 and free_pending_count == 0
    return can_finish, my_pending_count, free_pending_count, _build_finish_block_message(my_pending_count, free_pending_count)


async def _build_inventory_structure_for_report(
    report: Report,
    *,
    db: AsyncSession,
    user: User,
    cycle_started_at: date | None = None,
    cycle_days_left: int | None = None,
    start_block_message: str | None = None,
) -> InventoryStructureResponse:
    normalized = _normalize_location(report.location)
    cycle_version = int(report.cycle_version or 1)
    assignments = await _load_assignments(normalized, cycle_version, db)
    results = [
        row for row in await _load_results(report.id, db)
        if row.category_name != DEFAULT_CATEGORY_NAME and (row.subcategory_name is None or row.subcategory_name != DEFAULT_SUBCATEGORY_NAME)
    ]
    targets = await _resolve_selection_targets_for_date(normalized, cycle_version, report.report_date, db)
    report_snapshots = await _bootstrap_report_target_snapshots(report, assignments, results, db)
    employee_started = await _is_employee_started_report(report.id, user.id, db)
    employee_finished = await _is_employee_finished_report(report.id, user.id, db)
    await _sync_report_status(report, db)
    report_started = report.status != 'created'
    report_completed = report.status == 'completed'

    category_assignments, subcategory_assignments, item_assignments = _category_assignments_map(assignments)
    selected_category_ids, selected_subcategory_ids = _selection_target_maps(targets)
    retained_category_ids, retained_subcategory_ids, retained_item_ids = _build_retained_scope_for_user(user.id, report_snapshots, results)
    (
        snapshot_category_user_ids,
        snapshot_subcategory_user_ids,
        snapshot_item_user_ids,
        snapshot_category_owner_names,
        snapshot_subcategory_owner_names,
        snapshot_item_owner_names,
    ) = _report_snapshot_maps(report_snapshots)

    rows_by_category_target: dict[str, dict[str, CheckResult]] = defaultdict(dict)
    for row in results:
        rows_by_category_target[row.category_id][row.target_id] = row

    has_selected_scope = bool(selected_category_ids) or any(selected_subcategory_ids.values())
    has_retained_scope = bool(retained_category_ids) or any(retained_subcategory_ids.values()) or any(
        retained_item_ids.get(category_id, {}) for category_id in retained_item_ids
    )
    if not has_selected_scope and not has_retained_scope and not assignments and not results and not report_snapshots:
        can_finish_report, _, _, finish_block_message = _evaluate_finish_readiness([])
        resolved_cycle_started_at = cycle_started_at or report.report_date
        resolved_cycle_days_left = cycle_days_left if cycle_days_left is not None else _cycle_days_left_for_date(report.report_date)
        return InventoryStructureResponse(
            report_id=report.id,
            location=normalized,
            report_date=report.report_date.strftime('%d.%m.%Y'),
            categories=[],
            cycle_version=cycle_version,
            cycle_started_at=resolved_cycle_started_at.strftime('%d.%m.%Y'),
            cycle_days_left=resolved_cycle_days_left,
            report_status=report.status,
            employee_started=employee_started,
            employee_finished=employee_finished,
            report_started=report_started,
            report_completed=report_completed,
            can_finish_report=can_finish_report,
            finish_block_message=finish_block_message,
            start_block_message=start_block_message,
        )

    categories: list[CategoryModel] = []
    inventory = await _get_inventory_for(normalized, db=db)
    completed_subcategory_ids = await _load_completed_subcategory_ids_for_cycle(
        normalized,
        cycle_version,
        inventory,
        db,
        before_report_date=report.report_date,
    )
    inventory = _filter_inventory_by_targets(
        inventory,
        selected_category_ids,
        selected_subcategory_ids,
        retained_category_ids,
        retained_subcategory_ids,
        retained_item_ids,
        excluded_subcategory_ids=completed_subcategory_ids,
    )

    for raw_category in inventory['categories']:
        category_is_diagnostic = raw_category['name'] == DEFAULT_CATEGORY_NAME
        category_assignment = category_assignments.get(raw_category['id'])
        sub_assignments = subcategory_assignments.get(raw_category['id'], {})
        item_assignments_by_sub = item_assignments.get(raw_category['id'], {})

        snapshot_category_user_set = snapshot_category_user_ids.get(raw_category['id'], set())
        snapshot_category_taken_by_user = user.id in snapshot_category_user_ids.get(raw_category['id'], set())
        snapshot_category_taken_by_other = any(owner_id != user.id for owner_id in snapshot_category_user_set)
        snapshot_has_subcategory_assignments = any(snapshot_subcategory_user_ids.get(raw_category['id'], {}).values())
        snapshot_has_item_assignments = any(snapshot_item_user_ids.get(raw_category['id'], {}).values())
        assigned_to_current_user = bool((category_assignment and category_assignment.user_id == user.id) or snapshot_category_taken_by_user)
        assigned_to_other = bool((category_assignment and category_assignment.user_id != user.id) or snapshot_category_taken_by_other)
        has_my_subcategories = any(a.user_id == user.id for a in sub_assignments.values()) or any(
            user.id in user_ids for user_ids in snapshot_subcategory_user_ids.get(raw_category['id'], {}).values()
        )
        has_other_subcategories = any(a.user_id != user.id for a in sub_assignments.values()) or any(
            any(owner_id != user.id for owner_id in user_ids)
            for user_ids in snapshot_subcategory_user_ids.get(raw_category['id'], {}).values()
        )
        has_my_items = any(a.user_id == user.id for sub_items in item_assignments_by_sub.values() for a in sub_items.values()) or any(
            user.id in user_ids
            for sub_items in snapshot_item_user_ids.get(raw_category['id'], {}).values()
            for user_ids in sub_items.values()
        )
        has_other_items = any(a.user_id != user.id for sub_items in item_assignments_by_sub.values() for a in sub_items.values()) or any(
            any(owner_id != user.id for owner_id in user_ids)
            for sub_items in snapshot_item_user_ids.get(raw_category['id'], {}).values()
            for user_ids in sub_items.values()
        )

        has_free_diag_items = False
        for raw_sub in raw_category['subcategories']:
            diagnostic_sub = category_is_diagnostic or raw_sub['name'] == DEFAULT_SUBCATEGORY_NAME
            if not diagnostic_sub:
                continue
            assigned_item_ids = set(item_assignments_by_sub.get(raw_sub['id'], {}).keys())
            if any(
                item['id'] not in assigned_item_ids
                and not snapshot_item_user_ids.get(raw_category['id'], {}).get(raw_sub['id'], {}).get(item['id'])
                for item in raw_sub['items']
            ):
                has_free_diag_items = True
                break

        has_diagnostic_subcategories = any(sub['name'] == DEFAULT_SUBCATEGORY_NAME for sub in raw_category['subcategories'])
        selected_whole_category = raw_category['id'] in selected_category_ids
        selected_subcategory_names = [
            sub['name']
            for sub in raw_category['subcategories']
            if sub['id'] in selected_subcategory_ids.get(raw_category['id'], set())
        ]
        can_take_category = (
            selected_whole_category
            and (not category_is_diagnostic)
            and (not has_diagnostic_subcategories)
            and category_assignment is None
            and not snapshot_category_user_set
            and not sub_assignments
            and not snapshot_has_subcategory_assignments
            and not item_assignments_by_sub
            and not snapshot_has_item_assignments
        )

        owner_names = {a.user_full_name_snapshot for a in sub_assignments.values() if a.user_full_name_snapshot}
        owner_names.update(a.user_full_name_snapshot for sub_items in item_assignments_by_sub.values() for a in sub_items.values() if a.user_full_name_snapshot)
        assigned_to = None
        mixed_assignment = False
        if category_assignment:
            assigned_to = category_assignment.user_full_name_snapshot
        elif snapshot_category_taken_by_user:
            assigned_to = snapshot_category_owner_names.get(raw_category['id']) or user.full_name
        elif owner_names:
            if len(owner_names) == 1:
                assigned_to = next(iter(owner_names))
            else:
                assigned_to = 'Несколько сотрудников'
                mixed_assignment = True
        elif snapshot_category_taken_by_other:
            assigned_to = snapshot_category_owner_names.get(raw_category['id']) or 'Другой сотрудник'

        category_results = rows_by_category_target.get(raw_category['id'], {})
        subcategories: list[SubcategoryModel] = []

        selected_sub_ids = _subcategories_user_can_work(raw_category, category_assignment, sub_assignments, user.id)
        first_incomplete_selected: str | None = None
        sub_states: dict[str, tuple[bool, StatusEnum]] = {}
        for raw_sub in raw_category['subcategories']:
            is_completed, status = _subcategory_is_complete(raw_sub, category_results)
            sub_states[raw_sub['id']] = (is_completed, status)
            if raw_sub['id'] in selected_sub_ids and not is_completed and first_incomplete_selected is None:
                first_incomplete_selected = raw_sub['id']

        for raw_sub in raw_category['subcategories']:
            diagnostic_sub = category_is_diagnostic or raw_sub['name'] == DEFAULT_SUBCATEGORY_NAME
            is_completed, status = sub_states[raw_sub['id']]
            sub_item_assignments = item_assignments_by_sub.get(raw_sub['id'], {})
            snapshot_sub_user_set = snapshot_subcategory_user_ids.get(raw_category['id'], {}).get(raw_sub['id'], set())
            snapshot_sub_taken_by_user = user.id in snapshot_sub_user_set
            snapshot_sub_taken_by_other = any(owner_id != user.id for owner_id in snapshot_sub_user_set)
            has_my_items_in_sub = any(a.user_id == user.id for a in sub_item_assignments.values()) or any(
                user.id in user_ids
                for user_ids in snapshot_item_user_ids.get(raw_category['id'], {}).get(raw_sub['id'], {}).values()
            )
            has_other_items_in_sub = any(a.user_id != user.id for a in sub_item_assignments.values()) or any(
                any(owner_id != user.id for owner_id in user_ids)
                for user_ids in snapshot_item_user_ids.get(raw_category['id'], {}).get(raw_sub['id'], {}).values()
            )

            item_rows = []
            for item in raw_sub['items']:
                item_row = category_results.get(item['id'])
                item_status = StatusEnum.GREY
                is_final = False
                if item_row:
                    if item_row.status == 'green':
                        item_status = StatusEnum.GREEN
                        is_final = True
                    elif item_row.status == 'red':
                        item_status = StatusEnum.RED
                        is_final = True
                    elif item_row.status == 'orange':
                        item_status = StatusEnum.ORANGE

                item_assignment = sub_item_assignments.get(item['id'])
                snapshot_item_user_set = snapshot_item_user_ids.get(raw_category['id'], {}).get(raw_sub['id'], {}).get(item['id'], set())
                item_rows.append(ItemModel(
                    id=item['id'],
                    name=item['name'],
                    status=item_status,
                    is_final=is_final,
                    assigned_to=(item_assignment.user_full_name_snapshot if item_assignment else snapshot_item_owner_names.get(raw_category['id'], {}).get(raw_sub['id'], {}).get(item['id'])),
                    assigned_to_current_user=bool((item_assignment and item_assignment.user_id == user.id) or user.id in snapshot_item_user_set),
                    can_take=(
                        diagnostic_sub
                        and category_assignment is None
                        and not snapshot_category_user_set
                        and sub_assignments.get(raw_sub['id']) is None
                        and not snapshot_sub_user_set
                        and item_assignment is None
                        and not snapshot_item_user_set
                    ),
                    is_blocked_by_other=bool((item_assignment and item_assignment.user_id != user.id) or any(owner_id != user.id for owner_id in snapshot_item_user_set)),
                    is_diagnostic=diagnostic_sub,
                ))

            sub_assignment = sub_assignments.get(raw_sub['id'])
            sub_assigned_to_current_user = bool((sub_assignment and sub_assignment.user_id == user.id) or snapshot_sub_taken_by_user)
            sub_assigned_to_other = bool((sub_assignment and sub_assignment.user_id != user.id) or snapshot_sub_taken_by_other)
            can_take_sub = (
                (not diagnostic_sub)
                and category_assignment is None
                and not snapshot_category_user_set
                and sub_assignment is None
                and not snapshot_sub_user_set
                and not sub_item_assignments
                and not snapshot_item_user_ids.get(raw_category['id'], {}).get(raw_sub['id'], {})
            )

            is_locked = True
            is_expanded = False
            if raw_sub['id'] in selected_sub_ids:
                if first_incomplete_selected is None:
                    is_locked = False
                else:
                    is_locked = raw_sub['id'] != first_incomplete_selected and not is_completed
                    is_expanded = raw_sub['id'] == first_incomplete_selected and not is_completed
                if status == StatusEnum.ORANGE:
                    is_expanded = True
                    is_locked = False

            sub_owner_names = {a.user_full_name_snapshot for a in sub_item_assignments.values() if a.user_full_name_snapshot}
            sub_assigned_to = None
            if sub_assignment:
                sub_assigned_to = sub_assignment.user_full_name_snapshot
            elif category_assignment:
                sub_assigned_to = category_assignment.user_full_name_snapshot
            elif snapshot_sub_taken_by_user:
                sub_assigned_to = snapshot_subcategory_owner_names.get(raw_category['id'], {}).get(raw_sub['id']) or user.full_name
            elif snapshot_category_taken_by_user:
                sub_assigned_to = snapshot_category_owner_names.get(raw_category['id']) or user.full_name
            elif sub_owner_names:
                sub_assigned_to = next(iter(sub_owner_names)) if len(sub_owner_names) == 1 else 'Несколько сотрудников'
            elif snapshot_sub_taken_by_other:
                sub_assigned_to = snapshot_subcategory_owner_names.get(raw_category['id'], {}).get(raw_sub['id']) or 'Другой сотрудник'

            subcategories.append(
                SubcategoryModel(
                    id=raw_sub['id'],
                    name=raw_sub['name'],
                    is_locked=is_locked,
                    is_completed=is_completed,
                    is_expanded=is_expanded,
                    status=status,
                    items=item_rows,
                    assigned_to=sub_assigned_to,
                    assigned_to_current_user=sub_assigned_to_current_user,
                    can_take=can_take_sub,
                    is_blocked_by_other=sub_assigned_to_other or assigned_to_other,
                    taken_as_part_of_category=assigned_to_current_user,
                    is_diagnostic=diagnostic_sub,
                    has_my_items=has_my_items_in_sub,
                    has_other_items=has_other_items_in_sub,
                )
            )

        category_is_completed = _category_is_complete(raw_category, category_results)
        categories.append(
            CategoryModel(
                id=raw_category['id'],
                name=raw_category['name'],
                is_available=assigned_to_current_user or has_my_subcategories or has_my_items or can_take_category or has_free_diag_items,
                is_completed=category_is_completed,
                is_open=(assigned_to_current_user or has_my_subcategories or has_my_items) and not category_is_completed,
                subcategories=subcategories,
                assigned_to=assigned_to,
                assigned_to_current_user=assigned_to_current_user,
                can_take=can_take_category,
                is_blocked_by_other=assigned_to_other,
                has_my_subcategories=has_my_subcategories,
                has_other_subcategories=has_other_subcategories,
                mixed_assignment=mixed_assignment,
                is_diagnostic=category_is_diagnostic,
                has_my_items=has_my_items,
                has_other_items=has_other_items,
                selected_whole_category=selected_whole_category,
                selected_subcategory_names=selected_subcategory_names,
            )
        )

    can_finish_report, _, _, finish_block_message = _evaluate_finish_readiness(categories)

    resolved_cycle_started_at = cycle_started_at or report.report_date
    resolved_cycle_days_left = cycle_days_left if cycle_days_left is not None else _cycle_days_left_for_date(report.report_date)

    return InventoryStructureResponse(
        report_id=report.id,
        location=normalized,
        report_date=report.report_date.strftime('%d.%m.%Y'),
        categories=categories,
        cycle_version=cycle_version,
        cycle_started_at=resolved_cycle_started_at.strftime('%d.%m.%Y'),
        cycle_days_left=resolved_cycle_days_left,
        report_status=report.status,
        employee_started=employee_started,
        employee_finished=employee_finished,
        report_started=report_started,
        report_completed=report_completed,
        can_finish_report=can_finish_report,
        finish_block_message=finish_block_message,
        start_block_message=start_block_message,
    )


async def _has_user_activity_in_report(report_id: int, user_id: int, db: AsyncSession) -> bool:
    if await _is_employee_started_report(report_id, user_id, db):
        return True
    result = await db.scalar(
        select(CheckResult.id)
        .where(CheckResult.report_id == report_id)
        .where(CheckResult.checked_by_user_id == user_id)
        .limit(1)
    )
    return result is not None


async def _complete_employee_report_access_if_ready(report: Report, user: User, db: AsyncSession) -> bool:
    if await _is_employee_finished_report(report.id, user.id, db):
        return True

    state = await _build_inventory_structure_for_report(report, db=db, user=user)
    if not state.can_finish_report:
        return False

    db.add(ReportEmployeeCompletion(
        report_id=report.id,
        user_id=user.id,
        user_full_name_snapshot=user.full_name,
        finished_at=datetime.utcnow(),
    ))
    await db.flush()
    await _sync_report_status(report, db)
    return True


async def _get_previous_unfinished_report_block_message(
    current_report: Report,
    user: User,
    db: AsyncSession,
    *,
    auto_complete_ready: bool,
) -> str | None:
    previous_reports = (
        await db.scalars(
            select(Report)
            .where(Report.location == current_report.location)
            .where(Report.cycle_version == current_report.cycle_version)
            .where(Report.report_type == DAILY_REPORT_TYPE)
            .where(Report.report_date < current_report.report_date)
            .order_by(Report.report_date.desc(), Report.id.desc())
        )
    ).all()

    for previous_report in previous_reports:
        if not await _has_user_activity_in_report(previous_report.id, user.id, db):
            continue
        if await _is_employee_finished_report(previous_report.id, user.id, db):
            continue

        if auto_complete_ready and await _complete_employee_report_access_if_ready(previous_report, user, db):
            continue

        previous_state = await _build_inventory_structure_for_report(previous_report, db=db, user=user)
        if previous_state.can_finish_report:
            continue

        blocking_reason = previous_state.finish_block_message or 'прошлая ревизия ещё не доведена до конца'
        return (
            f'Предыдущая ревизия за {previous_report.report_date.strftime("%d.%m.%Y")} не завершена. '
            f'{blocking_reason} Сначала завершите прошлую ревизию.'
        )

    return None


async def get_inventory_data(location: str, db: AsyncSession, user: User) -> InventoryStructureResponse:
    normalized = _normalize_location(location)
    cycle = await _get_or_create_selection_cycle(normalized, db)
    today = get_moscow_today()
    report = await db.scalar(
        select(Report).where(
            Report.location == normalized,
            Report.report_date == today,
            Report.report_type == DAILY_REPORT_TYPE,
        ).limit(1)
    )
    days_left = _cycle_days_left_for_date(today)

    if report is None:
        preview_report = Report(
            location=normalized,
            report_date=today,
            cycle_version=cycle.cycle_version,
            report_type=DAILY_REPORT_TYPE,
            status='created',
        )
        start_block_message = await _get_previous_unfinished_report_block_message(preview_report, user, db, auto_complete_ready=False)
        return InventoryStructureResponse(
            report_id=None,
            location=normalized,
            report_date=today.isoformat(),
            categories=[],
            cycle_version=cycle.cycle_version,
            cycle_started_at=cycle.started_at.strftime('%d.%m.%Y'),
            cycle_days_left=days_left,
            report_status='created',
            employee_started=False,
            employee_finished=False,
            report_started=False,
            report_completed=False,
            can_finish_report=False,
            finish_block_message=None,
            start_block_message=start_block_message,
        )

    start_block_message = await _get_previous_unfinished_report_block_message(report, user, db, auto_complete_ready=False)
    return await _build_inventory_structure_for_report(
        report,
        db=db,
        user=user,
        cycle_started_at=cycle.started_at,
        cycle_days_left=days_left,
        start_block_message=start_block_message,
    )


async def assign_selection_to_user(report_id: int, category_id: str, target_type: str, subcategory_id: str | None, item_id: str | None, db: AsyncSession, user: User) -> AssignSelectionResponse:
    if not user.location:
        raise HTTPException(status_code=403, detail='Сотруднику не назначена точка.')

    report = await db.get(Report, report_id)
    if not report or report.location != user.location:
        raise HTTPException(status_code=404, detail='Ревизия не найдена.')
    await _sync_report_status(report, db)
    if report.status == 'completed':
        raise HTTPException(status_code=409, detail='Ревизия по этой точке уже завершена на сегодня.')
    if not await _is_employee_started_report(report.id, user.id, db):
        raise HTTPException(status_code=409, detail='Сначала нажмите «Начать ревизию».')
    if await _is_employee_finished_report(report.id, user.id, db):
        raise HTTPException(status_code=409, detail='Вы уже завершили свою ревизию на сегодня. Продолжить работу нельзя.')

    cycle = await _get_or_create_selection_cycle(report.location, db)
    assignments = await _load_assignments(report.location, cycle.cycle_version, db)
    targets = await _resolve_selection_targets_for_date(report.location, cycle.cycle_version, report.report_date, db)
    inventory = await _get_inventory_for(report.location, db=db)
    completed_subcategory_ids = await _load_completed_subcategory_ids_for_cycle(
        report.location,
        cycle.cycle_version,
        inventory,
        db,
        before_report_date=report.report_date,
    )
    category_map, sub_map, item_map = _category_assignments_map(assignments)
    selected_category_ids, selected_subcategory_ids = _selection_target_maps(targets)

    category = await _find_category(report.location, category_id, db=db)
    category_assignment = category_map.get(category_id)
    sub_assignments = sub_map.get(category_id, {})
    item_assignments_by_sub = item_map.get(category_id, {})
    category_is_diagnostic = category['name'] == DEFAULT_CATEGORY_NAME

    if target_type == 'category':
        if category_is_diagnostic:
            raise HTTPException(status_code=400, detail='Категорию «Без категории» нельзя брать целиком.')
        if category_id not in selected_category_ids:
            partial_sub_ids = selected_subcategory_ids.get(category_id, set())
            if partial_sub_ids:
                raise HTTPException(
                    status_code=400,
                    detail='Управляющий выбрал в этой категории только отдельные подкатегории. Возьмите нужную подкатегорию ниже.',
                )
            raise HTTPException(status_code=400, detail='Эта категория не выбрана управляющим для текущего цикла.')
        if any(sub['name'] == DEFAULT_SUBCATEGORY_NAME for sub in category['subcategories']):
            raise HTTPException(status_code=400, detail='Категории со служебными ветками «Без категории/Без подкатегории» нельзя брать целиком. Выберите обычную подкатегорию или конкретные товары.')
        remaining_subcategories = [
            sub for sub in category['subcategories']
            if sub['id'] not in completed_subcategory_ids.get(category_id, set()) and not _is_categoryless_subcategory(category, sub)
        ]
        if not remaining_subcategories:
            raise HTTPException(status_code=400, detail='В этой категории на текущий цикл не осталось новых подкатегорий.')
        if category_assignment:
            if category_assignment.user_id == user.id:
                return AssignSelectionResponse(success=True, message='Категория уже закреплена за вами.')
            raise HTTPException(status_code=400, detail=f'Категория уже закреплена за сотрудником {category_assignment.user_full_name_snapshot}.')
        if sub_assignments or item_assignments_by_sub:
            raise HTTPException(status_code=400, detail='Внутри этой категории уже есть закреплённые подкатегории или товары. Возьмите свободную подкатегорию либо товар отдельно.')

        db.add(CategoryAssignment(
            location=report.location,
            cycle_version=cycle.cycle_version,
            category_id=category_id,
            category_name=category['name'],
            subcategory_id=None,
            subcategory_name=None,
            target_type='category',
            target_id=category_id,
            target_name=category['name'],
            user_id=user.id,
            user_full_name_snapshot=user.full_name,
        ))
        await _upsert_report_target_snapshot(
            report_id=report.id,
            category_id=category_id,
            category_name=category['name'],
            subcategory_id=None,
            subcategory_name=None,
            target_type='category',
            target_id=category_id,
            target_name=category['name'],
            assigned_user_id_snapshot=user.id,
            assigned_user_name_snapshot=user.full_name,
            db=db,
        )
        await db.commit()
        _invalidate_runtime_inventory_cache(report.location)
        return AssignSelectionResponse(success=True, message='Категория закреплена за вами на сегодня.')

    if target_type == 'subcategory':
        if not subcategory_id:
            raise HTTPException(status_code=400, detail='Не указана подкатегория.')
        if category_assignment:
            if category_assignment.user_id == user.id:
                return AssignSelectionResponse(success=True, message='Вся категория уже закреплена за вами.')
            raise HTTPException(status_code=400, detail=f'Вся категория уже закреплена за сотрудником {category_assignment.user_full_name_snapshot}.')

        if category_id not in selected_category_ids and subcategory_id not in selected_subcategory_ids.get(category_id, set()):
            raise HTTPException(status_code=400, detail='Эта подкатегория не выбрана управляющим для текущего цикла.')

        subcategory = await _find_subcategory(report.location, category_id, subcategory_id, db=db)
        diagnostic_sub = category_is_diagnostic or subcategory['name'] == DEFAULT_SUBCATEGORY_NAME
        if diagnostic_sub:
            raise HTTPException(status_code=400, detail='Служебные ветки «Без категории/Без подкатегории» нельзя брать целиком. Выберите конкретные товары.')
        if subcategory_id in completed_subcategory_ids.get(category_id, set()):
            raise HTTPException(status_code=400, detail='Эта подкатегория уже была пройдена в текущем цикле месяца.')

        existing = sub_assignments.get(subcategory_id)
        if existing:
            if existing.user_id == user.id:
                return AssignSelectionResponse(success=True, message='Подкатегория уже закреплена за вами.')
            raise HTTPException(status_code=400, detail=f'Подкатегория уже закреплена за сотрудником {existing.user_full_name_snapshot}.')
        if item_assignments_by_sub.get(subcategory_id):
            raise HTTPException(status_code=400, detail='Внутри этой подкатегории уже есть закреплённые товары. Выберите свободный товар отдельно.')

        db.add(CategoryAssignment(
            location=report.location,
            cycle_version=cycle.cycle_version,
            category_id=category_id,
            category_name=category['name'],
            subcategory_id=subcategory_id,
            subcategory_name=subcategory['name'],
            target_type='subcategory',
            target_id=subcategory_id,
            target_name=subcategory['name'],
            user_id=user.id,
            user_full_name_snapshot=user.full_name,
        ))
        await _upsert_report_target_snapshot(
            report_id=report.id,
            category_id=category_id,
            category_name=category['name'],
            subcategory_id=subcategory_id,
            subcategory_name=subcategory['name'],
            target_type='subcategory',
            target_id=subcategory_id,
            target_name=subcategory['name'],
            assigned_user_id_snapshot=user.id,
            assigned_user_name_snapshot=user.full_name,
            db=db,
        )
        await db.commit()
        _invalidate_runtime_inventory_cache(report.location)
        return AssignSelectionResponse(success=True, message='Подкатегория закреплена за вами на сегодня.')

    if target_type == 'item':
        if not subcategory_id or not item_id:
            raise HTTPException(status_code=400, detail='Для выбора товара нужно указать подкатегорию и товар.')
        if category_assignment:
            raise HTTPException(status_code=400, detail='Сейчас внутри этой категории нельзя закреплять отдельные товары, потому что уже есть выбор категории целиком.')

        if category_id not in selected_category_ids and subcategory_id not in selected_subcategory_ids.get(category_id, set()):
            raise HTTPException(status_code=400, detail='Эта подкатегория не выбрана управляющим для текущего цикла.')

        subcategory = await _find_subcategory(report.location, category_id, subcategory_id, db=db)
        diagnostic_sub = category_is_diagnostic or subcategory['name'] == DEFAULT_SUBCATEGORY_NAME
        if not diagnostic_sub:
            raise HTTPException(status_code=400, detail='Поштучный выбор доступен только для служебных веток «Без категории/Без подкатегории».')

        sub_assignment = sub_assignments.get(subcategory_id)
        if sub_assignment:
            if sub_assignment.user_id == user.id:
                raise HTTPException(status_code=400, detail='У вас уже закреплена вся подкатегория. Отдельный товар выбирать не нужно.')
            raise HTTPException(status_code=400, detail=f'Подкатегория уже закреплена за сотрудником {sub_assignment.user_full_name_snapshot}.')

        existing_item = item_assignments_by_sub.get(subcategory_id, {}).get(item_id)
        if existing_item:
            if existing_item.user_id == user.id:
                return AssignSelectionResponse(success=True, message='Товар уже закреплён за вами.')
            raise HTTPException(status_code=400, detail=f'Товар уже закреплён за сотрудником {existing_item.user_full_name_snapshot}.')

        item = next((row for row in subcategory['items'] if row['id'] == item_id), None)
        if not item:
            raise HTTPException(status_code=404, detail='Товар не найден в выбранной подкатегории.')

        db.add(CategoryAssignment(
            location=report.location,
            cycle_version=cycle.cycle_version,
            category_id=category_id,
            category_name=category['name'],
            subcategory_id=subcategory_id,
            subcategory_name=subcategory['name'],
            target_type='item',
            target_id=item_id,
            target_name=item['name'],
            user_id=user.id,
            user_full_name_snapshot=user.full_name,
        ))
        await _upsert_report_target_snapshot(
            report_id=report.id,
            category_id=category_id,
            category_name=category['name'],
            subcategory_id=subcategory_id,
            subcategory_name=subcategory['name'],
            target_type='item',
            target_id=item_id,
            target_name=item['name'],
            assigned_user_id_snapshot=user.id,
            assigned_user_name_snapshot=user.full_name,
            db=db,
        )
        await db.commit()
        _invalidate_runtime_inventory_cache(report.location)
        return AssignSelectionResponse(success=True, message='Товар закреплён за вами на сегодня.')

    raise HTTPException(status_code=400, detail='Некорректный тип выбора.')


def _user_can_verify_target(user: User, report: Report, category_id: str, subcategory_id: str | None, target_id: str, target_type: str, assignments: list[CategoryAssignment]) -> bool:
    category_map, sub_map, item_map = _category_assignments_map(assignments)
    cat_assignment = category_map.get(category_id)
    if cat_assignment and cat_assignment.user_id == user.id:
        return True
    if subcategory_id and sub_map.get(category_id, {}).get(subcategory_id) and sub_map[category_id][subcategory_id].user_id == user.id:
        return True
    if target_type == 'item' and subcategory_id:
        item_assignment = item_map.get(category_id, {}).get(subcategory_id, {}).get(target_id)
        if item_assignment and item_assignment.user_id == user.id:
            return True
    return False



async def _upsert_check_result(
    *,
    report_id: int,
    category_id: str,
    category_name: str,
    subcategory_id: str | None,
    subcategory_name: str | None,
    target_type: str,
    target_id: str,
    target_name: str,
    expected_qty: float,
    actual_qty: float,
    diff: float,
    status_value: str,
    attempts_used: int,
    checked_by_user_id: int | None,
    checked_by_name_snapshot: str | None,
    db: AsyncSession,
) -> None:
    existing = await db.scalar(select(CheckResult).where(CheckResult.report_id == report_id, CheckResult.target_id == target_id).limit(1))
    if existing:
        existing.category_id = category_id
        existing.category_name = category_name
        existing.subcategory_id = subcategory_id
        existing.subcategory_name = subcategory_name
        existing.target_type = target_type
        existing.target_name = target_name
        existing.expected_qty = expected_qty
        existing.actual_qty = actual_qty
        existing.diff = diff
        existing.status = status_value
        existing.attempts_used = attempts_used
        existing.checked_by_user_id = checked_by_user_id
        existing.checked_by_name_snapshot = checked_by_name_snapshot
    else:
        db.add(CheckResult(
            report_id=report_id,
            category_id=category_id,
            category_name=category_name,
            subcategory_id=subcategory_id,
            subcategory_name=subcategory_name,
            target_type=target_type,
            target_id=target_id,
            target_name=target_name,
            expected_qty=expected_qty,
            actual_qty=actual_qty,
            diff=diff,
            status=status_value,
            attempts_used=attempts_used,
            checked_by_user_id=checked_by_user_id,
            checked_by_name_snapshot=checked_by_name_snapshot,
        ))


async def _get_or_increment_attempt_count(
    *,
    report_id: int,
    target_type: str,
    target_id: str,
    checked_by_user_id: int,
    db: AsyncSession,
) -> int:
    progress = await db.scalar(
        select(VerifyAttemptProgress)
        .where(VerifyAttemptProgress.report_id == report_id)
        .where(VerifyAttemptProgress.target_type == target_type)
        .where(VerifyAttemptProgress.target_id == target_id)
        .where(VerifyAttemptProgress.checked_by_user_id == checked_by_user_id)
        .limit(1)
    )
    now = datetime.utcnow()
    if progress:
        progress.attempts_used += 1
        progress.updated_at = now
        return progress.attempts_used

    progress = VerifyAttemptProgress(
        report_id=report_id,
        target_type=target_type,
        target_id=target_id,
        checked_by_user_id=checked_by_user_id,
        attempts_used=1,
        created_at=now,
        updated_at=now,
    )
    db.add(progress)
    await db.flush()
    return progress.attempts_used


async def _clear_attempt_count(
    *,
    report_id: int,
    target_type: str,
    target_id: str,
    checked_by_user_id: int,
    db: AsyncSession,
) -> None:
    await db.execute(
        delete(VerifyAttemptProgress)
        .where(VerifyAttemptProgress.report_id == report_id)
        .where(VerifyAttemptProgress.target_type == target_type)
        .where(VerifyAttemptProgress.target_id == target_id)
        .where(VerifyAttemptProgress.checked_by_user_id == checked_by_user_id)
    )


async def verify_item_or_category(data: VerifyRequest, db: AsyncSession, checked_by_user: User) -> VerifyResponse:
    report = await db.get(Report, data.report_id)
    if not report:
        raise HTTPException(status_code=404, detail='Ревизия не найдена.')
    if checked_by_user.location != report.location:
        raise HTTPException(status_code=403, detail='Ревизия относится к другой точке.')
    await _sync_report_status(report, db)
    if report.status == 'completed':
        raise HTTPException(status_code=409, detail='Ревизия по этой точке уже завершена на сегодня.')
    if not await _is_employee_started_report(report.id, checked_by_user.id, db):
        raise HTTPException(status_code=409, detail='Сначала нажмите «Начать ревизию».')
    if await _is_employee_finished_report(report.id, checked_by_user.id, db):
        raise HTTPException(status_code=409, detail='Вы уже завершили свою ревизию на сегодня. Продолжить работу нельзя.')

    category_id, category_name, subcategory_id, subcategory_name, target_type, target_name, expected_qty = await _find_target(report.location, data.target_id, db=db)
    assignments = await _load_assignments(report.location, report.cycle_version, db)
    if not _user_can_verify_target(checked_by_user, report, category_id, subcategory_id, data.target_id, target_type, assignments):
        raise HTTPException(status_code=403, detail='Эта категория, подкатегория или товар не закреплены за вами.')

    attempt_number = await _get_or_increment_attempt_count(
        report_id=report.id,
        target_type=target_type,
        target_id=data.target_id,
        checked_by_user_id=checked_by_user.id,
        db=db,
    )

    is_correct = abs(data.quantity - expected_qty) < 1e-9
    if is_correct:
        await _upsert_report_target_snapshot(
            report_id=report.id,
            category_id=category_id,
            category_name=category_name,
            subcategory_id=subcategory_id,
            subcategory_name=subcategory_name,
            target_type=target_type,
            target_id=data.target_id,
            target_name=target_name,
            assigned_user_id_snapshot=checked_by_user.id,
            assigned_user_name_snapshot=checked_by_user.full_name,
            db=db,
        )
        await _upsert_check_result(
            report_id=report.id,
            category_id=category_id,
            category_name=category_name,
            subcategory_id=subcategory_id,
            subcategory_name=subcategory_name,
            target_type=target_type,
            target_id=data.target_id,
            target_name=target_name,
            expected_qty=expected_qty,
            actual_qty=data.quantity,
            diff=0.0,
            status_value='green',
            attempts_used=attempt_number,
            checked_by_user_id=checked_by_user.id,
            checked_by_name_snapshot=checked_by_user.full_name,
            db=db,
        )
        await _clear_attempt_count(
            report_id=report.id,
            target_type=target_type,
            target_id=data.target_id,
            checked_by_user_id=checked_by_user.id,
            db=db,
        )
        await db.commit()
        _invalidate_runtime_inventory_cache(report.location)
        await _refresh_report_status(report, db)
        return VerifyResponse(is_correct=True, attempts_left=0, message='Верно!', expand_category=False)

    attempts_left = max(0, 3 - attempt_number)
    if attempts_left > 0:
        await db.commit()
        return VerifyResponse(is_correct=False, attempts_left=attempts_left, message=f'Неверно. Осталось {attempts_left} попытк(и).', expand_category=False)

    status_value = 'orange' if data.is_category else 'red'
    await _upsert_report_target_snapshot(
        report_id=report.id,
        category_id=category_id,
        category_name=category_name,
        subcategory_id=subcategory_id,
        subcategory_name=subcategory_name,
        target_type=target_type,
        target_id=data.target_id,
        target_name=target_name,
        assigned_user_id_snapshot=checked_by_user.id,
        assigned_user_name_snapshot=checked_by_user.full_name,
        db=db,
    )
    await _upsert_check_result(
        report_id=report.id,
        category_id=category_id,
        category_name=category_name,
        subcategory_id=subcategory_id,
        subcategory_name=subcategory_name,
        target_type=target_type,
        target_id=data.target_id,
        target_name=target_name,
        expected_qty=expected_qty,
        actual_qty=data.quantity,
        diff=float(data.quantity - expected_qty),
        status_value=status_value,
        attempts_used=attempt_number,
        checked_by_user_id=checked_by_user.id,
        checked_by_name_snapshot=checked_by_user.full_name,
        db=db,
    )
    await _clear_attempt_count(
        report_id=report.id,
        target_type=target_type,
        target_id=data.target_id,
        checked_by_user_id=checked_by_user.id,
        db=db,
    )
    await db.commit()
    _invalidate_runtime_inventory_cache(report.location)
    await _refresh_report_status(report, db)
    return VerifyResponse(
        is_correct=False,
        attempts_left=0,
        message='Расхождение! Переходим к поштучной проверке...' if data.is_category else 'Расхождение зафиксировано.',
        expand_category=data.is_category,
    )


async def start_report(report_id: int | None, db: AsyncSession, user: User) -> StartReportResponse:
    if user.role != RoleEnum.EMPLOYEE.value or not user.location:
        raise HTTPException(status_code=403, detail='Можно начать только свою ревизию по назначенной точке.')

    report: Report | None = None
    if report_id is not None:
        report = await db.get(Report, report_id)
        if not report:
            raise HTTPException(status_code=404, detail='Ревизия не найдена.')
        if user.location != report.location:
            raise HTTPException(status_code=403, detail='Можно начать только свою ревизию по назначенной точке.')
    else:
        cycle = await _get_or_create_selection_cycle(user.location, db)
        report = await get_or_create_daily_report(user.location, cycle.cycle_version, db)

    await _sync_report_status(report, db)
    if report.status == 'completed':
        return StartReportResponse(success=True, message='Ревизия по этой точке уже завершена на сегодня.')

    previous_block_message = await _get_previous_unfinished_report_block_message(report, user, db, auto_complete_ready=True)
    if previous_block_message:
        await db.commit()
        raise HTTPException(status_code=409, detail=previous_block_message)

    existing = await db.scalar(
        select(ReportEmployeeStart)
        .where(ReportEmployeeStart.report_id == report.id)
        .where(ReportEmployeeStart.user_id == user.id)
        .limit(1)
    )
    if existing:
        await db.commit()
        return StartReportResponse(success=True, message='Ревизия уже начата. Можно продолжать работу.')

    db.add(ReportEmployeeStart(
        report_id=report.id,
        user_id=user.id,
        user_full_name_snapshot=user.full_name,
        started_at=datetime.utcnow(),
    ))
    await db.flush()
    await _sync_report_status(report, db)
    await db.commit()
    _invalidate_runtime_inventory_cache(report.location)
    return StartReportResponse(success=True, message='Ревизия начата. Можно приступать к работе.')


async def finish_report(report_id: int, db: AsyncSession, user: User) -> tuple[bool, str]:
    report = await db.get(Report, report_id)
    if not report:
        return False, 'Ревизия не найдена.'
    if user.role != RoleEnum.EMPLOYEE.value or user.location != report.location:
        raise HTTPException(status_code=403, detail='Можно завершать только свою ревизию по назначенной точке.')

    await _sync_report_status(report, db)
    if report.status == 'completed':
        return True, 'Ревизия по этой точке уже завершена на сегодня.'
    if not await _is_employee_started_report(report.id, user.id, db):
        raise HTTPException(status_code=409, detail='Сначала нажмите «Начать ревизию».')

    existing = await db.scalar(
        select(ReportEmployeeCompletion)
        .where(ReportEmployeeCompletion.report_id == report.id)
        .where(ReportEmployeeCompletion.user_id == user.id)
        .limit(1)
    )
    if existing:
        return True, 'Вы уже завершили свою ревизию на сегодня.'

    state = await _build_inventory_structure_for_report(report, db=db, user=user)
    if not state.can_finish_report:
        raise HTTPException(status_code=409, detail=state.finish_block_message or 'Сначала завершите все свои выборы и разберите свободные подкатегории.')

    db.add(ReportEmployeeCompletion(
        report_id=report.id,
        user_id=user.id,
        user_full_name_snapshot=user.full_name,
        finished_at=datetime.utcnow(),
    ))
    await db.flush()
    await _sync_report_status(report, db)
    await db.commit()
    _invalidate_runtime_inventory_cache(report.location)

    if report.status == 'completed':
        return True, 'Ревизия завершена.'
    return True, f'Ваша ревизия завершена.'


async def reopen_employee_report_access(report_id: int, employee_user_id: int, db: AsyncSession, current_user: User) -> ReopenEmployeeAccessResponse:
    report = await db.get(Report, report_id)
    if not report:
        raise HTTPException(status_code=404, detail='Ревизия не найдена.')

    await ensure_user_can_access_location(current_user, report.location, db)

    if (report.report_type or DAILY_REPORT_TYPE) != DAILY_REPORT_TYPE:
        raise HTTPException(status_code=400, detail='Вернуть доступ можно только для обычной дневной ревизии.')

    if report.report_date != get_moscow_today():
        raise HTTPException(status_code=400, detail='Вернуть доступ можно только для текущей ревизии за сегодня.')

    employee = await db.get(User, employee_user_id)
    if not employee or employee.role != RoleEnum.EMPLOYEE.value:
        raise HTTPException(status_code=404, detail='Сотрудник не найден.')

    employee_location = _normalize_location(employee.location or '') if employee.location else ''
    if not employee_location or employee_location != report.location:
        raise HTTPException(status_code=403, detail='Нельзя вернуть в ревизию сотрудника из другой точки.')

    completion = await db.scalar(
        select(ReportEmployeeCompletion)
        .where(ReportEmployeeCompletion.report_id == report.id)
        .where(ReportEmployeeCompletion.user_id == employee.id)
        .limit(1)
    )

    if completion is None:
        await _sync_report_status(report, db)
        await db.commit()
        return ReopenEmployeeAccessResponse(success=True, message=f'У сотрудника {employee.full_name} уже открыт доступ к ревизии.')

    await db.delete(completion)
    await db.flush()
    await _sync_report_status(report, db)
    await db.commit()
    _invalidate_runtime_inventory_cache(report.location)
    return ReopenEmployeeAccessResponse(success=True, message=f'Сотруднику {employee.full_name} снова открыт доступ к ревизии.')


async def get_reports_history(
    location: str,
    db: AsyncSession,
    *,
    year: int | None = None,
    month: int | None = None,
) -> ReportHistoryResponse:
    normalized = _normalize_location(location)
    query = select(Report).where(Report.location == normalized)

    if month is not None and not 1 <= month <= 12:
        raise HTTPException(status_code=400, detail='Некорректный месяц.')
    if year is not None and not 2000 <= year <= 2100:
        raise HTTPException(status_code=400, detail='Некорректный год.')

    if year is not None and month is not None:
        period_start = date(year, month, 1)
        period_end = date(year, month, calendar.monthrange(year, month)[1])
        query = query.where(Report.report_date >= period_start, Report.report_date <= period_end)
    elif year is not None:
        query = query.where(Report.report_date >= date(year, 1, 1), Report.report_date <= date(year, 12, 31))

    reports = (await db.scalars(query.order_by(Report.date_created.desc(), Report.id.desc()))).all()
    for report in reports:
        await _sync_report_status(report, db)
    await db.commit()

    report_numbers = _build_report_numbers(reports)

    history_items: list[ReportHistoryItem] = []
    for report in reports:
        report_type = report.report_type or DAILY_REPORT_TYPE
        report_number = report_numbers.get(report.id) if report_type != FINAL_REPORT_TYPE else None
        cycle_label = _cycle_label_for_date(report.report_date)
        if report_type == FINAL_REPORT_TYPE:
            label = (
                f"{cycle_label} · Итоговая · "
                f"{_format_moscow_datetime(report.date_created)} — {_report_status_label(report.status)}"
            )
        else:
            label = (
                f"{cycle_label} · №{report_number or '-'} · "
                f"{_format_moscow_datetime(report.date_created)} — {_report_status_label(report.status)}"
            )
        history_items.append(
            ReportHistoryItem(
                report_id=report.id,
                report_number=report_number,
                report_type=report_type,
                date=_format_moscow_datetime(report.date_created),
                status=report.status,
                label=label,
            )
        )

    return ReportHistoryResponse(location=normalized, reports=history_items)


def _category_assignment_label(category_id: str, report_assignments: list[CategoryAssignment]) -> str | None:
    category_level = [a for a in report_assignments if a.category_id == category_id and a.target_type == 'category']
    if category_level:
        return category_level[0].user_full_name_snapshot
    owners = sorted({a.user_full_name_snapshot for a in report_assignments if a.category_id == category_id and a.user_full_name_snapshot})
    if not owners:
        return None
    if len(owners) == 1:
        return owners[0]
    return 'Несколько сотрудников'


def _owner_label(owner_names: set[str]) -> str | None:
    owners = sorted({name for name in owner_names if name}, key=str.lower)
    if not owners:
        return None
    if len(owners) == 1:
        return owners[0]
    return 'Несколько сотрудников'


def _category_snapshot_assignment_label(category_id: str, report_snapshots: list[ReportTargetSnapshot]) -> str | None:
    category_level = {
        row.assigned_user_name_snapshot
        for row in report_snapshots
        if row.category_id == category_id and row.target_type == 'category' and row.assigned_user_name_snapshot
    }
    if category_level:
        return _owner_label(category_level)
    owners = {
        row.assigned_user_name_snapshot
        for row in report_snapshots
        if row.category_id == category_id and row.assigned_user_name_snapshot
    }
    return _owner_label(owners)


def _subcategory_snapshot_assignment_label(category_id: str, subcategory_id: str, report_snapshots: list[ReportTargetSnapshot]) -> str | None:
    category_level = {
        row.assigned_user_name_snapshot
        for row in report_snapshots
        if row.category_id == category_id and row.target_type == 'category' and row.assigned_user_name_snapshot
    }
    if category_level:
        return _owner_label(category_level)

    subcategory_level = {
        row.assigned_user_name_snapshot
        for row in report_snapshots
        if row.category_id == category_id and row.subcategory_id == subcategory_id and row.target_type == 'subcategory' and row.assigned_user_name_snapshot
    }
    if subcategory_level:
        return _owner_label(subcategory_level)

    item_level = {
        row.assigned_user_name_snapshot
        for row in report_snapshots
        if row.category_id == category_id and row.subcategory_id == subcategory_id and row.target_type == 'item' and row.assigned_user_name_snapshot
    }
    return _owner_label(item_level)


def _subcategory_taken_in_report(category_id: str, subcategory_id: str, report_snapshots: list[ReportTargetSnapshot]) -> bool:
    return any(
        row.category_id == category_id and (
            row.target_type == 'category'
            or (row.subcategory_id == subcategory_id and row.target_type in {'subcategory', 'item'})
        )
        for row in report_snapshots
    )


def _category_taken_in_report(category_id: str, report_snapshots: list[ReportTargetSnapshot]) -> bool:
    return any(
        row.category_id == category_id and row.target_type == 'category'
        for row in report_snapshots
    )


def _subcategory_name_lookup(
    category_id: str,
    source_category: dict[str, Any] | None,
    targets: list[SelectionTarget],
    sub_assignments: dict[str, CategoryAssignment],
    item_assignments_by_sub: dict[str, dict[str, CategoryAssignment]],
    report_snapshots: list[ReportTargetSnapshot],
    results: list[CheckResult],
) -> dict[str, str]:
    names: dict[str, str] = {}

    if source_category:
        for raw_sub in source_category.get('subcategories', []):
            if _is_categoryless_subcategory(source_category, raw_sub):
                continue
            names.setdefault(raw_sub['id'], raw_sub['name'])

    for row in targets:
        if row.category_id == category_id and row.subcategory_id and row.subcategory_name:
            names.setdefault(row.subcategory_id, row.subcategory_name)

    for assignment in sub_assignments.values():
        if assignment.subcategory_id and assignment.subcategory_name:
            names.setdefault(assignment.subcategory_id, assignment.subcategory_name)

    for sub_id, item_assignments in item_assignments_by_sub.items():
        for assignment in item_assignments.values():
            if assignment.subcategory_id and assignment.subcategory_name:
                names.setdefault(assignment.subcategory_id, assignment.subcategory_name)
                break
        else:
            if sub_id:
                names.setdefault(sub_id, sub_id)

    for row in report_snapshots:
        if row.category_id == category_id and row.subcategory_id and row.subcategory_name:
            names.setdefault(row.subcategory_id, row.subcategory_name)

    for row in results:
        if row.category_id == category_id and row.subcategory_id and row.subcategory_name:
            names.setdefault(row.subcategory_id, row.subcategory_name)

    return names


def _subcategory_completion_status_for_admin(
    raw_subcategory: dict[str, Any] | None,
    subcategory_id: str,
    subcategory_name: str,
    result_map: dict[str, CheckResult],
) -> tuple[bool, StatusEnum]:
    if raw_subcategory is not None:
        return _subcategory_is_complete(raw_subcategory, result_map)

    sub_row = result_map.get(subcategory_id)
    if sub_row and sub_row.status == 'green':
        return True, StatusEnum.GREEN
    if sub_row and sub_row.status == 'orange':
        return False, StatusEnum.ORANGE
    if sub_row and sub_row.status == 'red':
        return True, StatusEnum.RED
    return False, StatusEnum.GREY


def _build_report_numbers(reports: list[Report]) -> dict[int, int]:
    ordered = sorted((item for item in reports if (item.report_type or DAILY_REPORT_TYPE) != FINAL_REPORT_TYPE), key=lambda item: (item.cycle_version, item.date_created, item.id))
    counters: dict[int, int] = defaultdict(int)
    numbers: dict[int, int] = {}

    for report in ordered:
        cycle_key = int(report.cycle_version or 0)
        counters[cycle_key] += 1
        numbers[report.id] = counters[cycle_key]

    return numbers


async def _get_report_number(report: Report, db: AsyncSession) -> int:
    if (report.report_type or DAILY_REPORT_TYPE) == FINAL_REPORT_TYPE:
        return 0

    result = await db.scalar(
        select(func.count())
        .select_from(Report)
        .where(
            Report.location == report.location,
            Report.cycle_version == report.cycle_version,
            Report.report_type == DAILY_REPORT_TYPE,
            or_(
                Report.date_created < report.date_created,
                and_(Report.date_created == report.date_created, Report.id <= report.id),
            ),
        )
    )
    return int(result or 0)


def _logic_money_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return round(float(value) / 100.0, 2)
    except (TypeError, ValueError):
        return None



def _logic_first_money_value(container: dict[str, Any], fields: tuple[str, ...]) -> float | None:
    for field in fields:
        if field in container and container.get(field) is not None:
            amount = _logic_money_or_none(container.get(field))
            if amount is not None:
                return amount
    return None



def _logic_quantity(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0



def _logic_clean_text(value: str | None) -> str:
    return ' '.join(str(value or '').strip().lower().replace('ё', 'е').split())



def _logic_candidate_name_variants(value: str | None) -> set[str]:
    raw = _logic_clean_text(value)
    if not raw:
        return set()
    variants = {raw}
    separators = [':', '-', '—', '(', ')', '[', ']', '/', '\\']
    parts = {raw}
    for sep in separators:
        next_parts = set()
        for part in parts:
            next_parts.update(p.strip() for p in part.split(sep) if p.strip())
        parts |= next_parts
    variants |= parts
    return {item for item in variants if item}



def _logic_iter_point_reference_candidates(doc: dict[str, Any]) -> list[dict[str, Any]]:
    collected: list[dict[str, Any]] = []
    seen: set[int] = set()

    def add(candidate: Any) -> None:
        if not isinstance(candidate, dict):
            return
        marker = id(candidate)
        if marker in seen:
            return
        seen.add(marker)
        collected.append(candidate)

    def add_store_refs(container: dict[str, Any] | None) -> None:
        if not isinstance(container, dict):
            return
        add(container.get('store'))
        add(container.get('retailStore'))
        add(container.get('retailstore'))

    add_store_refs(doc)

    retail_shift = doc.get('retailShift')
    add(retail_shift)
    add_store_refs(retail_shift if isinstance(retail_shift, dict) else None)

    demand = doc.get('demand')
    add(demand)
    if isinstance(demand, dict):
        add_store_refs(demand)
        demand_shift = demand.get('retailShift')
        add(demand_shift)
        add_store_refs(demand_shift if isinstance(demand_shift, dict) else None)

    return collected



def _logic_doc_matches_point(doc: dict[str, Any], point: LocationPoint) -> bool:
    if doc.get('applicable') is False:
        return False

    point_ids = {str(point.ms_store_id or '').strip()} - {''}
    point_names = set()
    for raw in (point.name, point.ms_store_name):
        point_names |= _logic_candidate_name_variants(raw)

    if not point_ids and not point_names:
        return True

    candidate_names: set[str] = set()
    candidate_ids: set[str] = set()
    for candidate in _logic_iter_point_reference_candidates(doc):
        candidate_id = candidate.get('id') or ((candidate.get('meta') or {}).get('id') if isinstance(candidate.get('meta'), dict) else None)
        if not candidate_id and isinstance(candidate.get('meta'), dict):
            href = candidate['meta'].get('href')
            if href:
                candidate_id = str(href).rstrip('/').split('/')[-1]
        if candidate_id:
            candidate_ids.add(str(candidate_id))
        candidate_names |= _logic_candidate_name_variants(candidate.get('name'))
        meta = candidate.get('meta') if isinstance(candidate.get('meta'), dict) else None
        if meta:
            candidate_names |= _logic_candidate_name_variants(meta.get('name'))

    if point_ids & candidate_ids:
        return True

    if point_names and candidate_names:
        if point_names & candidate_names:
            return True
        for point_name in point_names:
            for candidate_name in candidate_names:
                if point_name in candidate_name or candidate_name in point_name:
                    return True

    return False



def _logic_iter_positions(doc: dict[str, Any]) -> list[dict[str, Any]]:
    positions = doc.get('positions')
    if isinstance(positions, dict):
        return positions.get('rows', []) or []
    if isinstance(positions, list):
        return positions
    return []



def _logic_extract_position_amount(position: dict[str, Any]) -> float:
    quantity = _logic_quantity(position.get('quantity'))
    amount = _logic_first_money_value(position, ('sum', 'amount', 'saleSum', 'retailSum'))
    if amount is not None:
        return round(amount, 2)
    unit_price = _logic_first_money_value(position, ('price', 'salePrice', 'sellingPrice'))
    if unit_price is not None and quantity > 0:
        return round(unit_price * quantity, 2)
    return 0.0



def _logic_extract_position_cost_amount(position: dict[str, Any]) -> float | None:
    quantity = _logic_quantity(position.get('quantity'))
    total_cost = _logic_first_money_value(position, ('costSum', 'buySum', 'buyPriceSum', 'purchaseCostSum', 'costAmount', 'purchaseSum', 'cost'))
    if total_cost is not None:
        return round(max(total_cost, 0.0), 2)
    unit_cost = _logic_first_money_value(position, ('buyPrice', 'costPrice', 'purchasePrice', 'purchaseCost', 'costValue'))
    if unit_cost is not None and quantity > 0:
        return round(max(unit_cost, 0.0) * quantity, 2)
    return None



def _logic_extract_document_item_id(position: dict[str, Any]) -> str | None:
    assortment = position.get('assortment')
    if not isinstance(assortment, dict):
        return None
    meta = assortment.get('meta') if isinstance(assortment.get('meta'), dict) else assortment
    item_id = meta.get('id')
    if item_id:
        return str(item_id)
    href = meta.get('href')
    if href:
        return str(href).rstrip('/').split('/')[-1]
    return None



async def _load_cached_discrepancy_financials(location: str, item_ids: set[str], db: AsyncSession | None = None) -> dict[str, dict[str, float | None | str]]:
    normalized = _normalize_location(location)
    unique_item_ids = {str(item_id).strip() for item_id in item_ids if str(item_id).strip()}
    if not unique_item_ids or db is None:
        return {}

    point = await db.scalar(select(LocationPoint).where(LocationPoint.name == normalized).limit(1))
    if point is None:
        return {}

    cache_rows = (
        await db.scalars(
            select(ProductFinancialCache)
            .where(ProductFinancialCache.location_point_id == point.id)
            .where(ProductFinancialCache.item_id.in_(sorted(unique_item_ids)))
        )
    ).all()
    financials: dict[str, dict[str, float | None | str]] = {
        row.item_id: {
            'cost_price': float(row.cost_price) if row.cost_price is not None else None,
            'retail_price': float(row.retail_price) if row.retail_price is not None else None,
            'cost_price_source': 'cache' if row.cost_price is not None else None,
            'cost_price_note': None,
            'cost_price_updated_at': row.updated_at.isoformat() if row.updated_at else None,
        }
        for row in cache_rows
    }

    override_rows = (
        await db.scalars(
            select(ProductCostOverride)
            .where(ProductCostOverride.location_point_id == point.id)
            .where(ProductCostOverride.item_id.in_(sorted(unique_item_ids)))
        )
    ).all()
    for row in override_rows:
        current = financials.get(row.item_id) or {
            'cost_price': None,
            'retail_price': None,
            'cost_price_source': None,
            'cost_price_note': None,
            'cost_price_updated_at': None,
        }
        current['cost_price'] = float(row.cost_price) if row.cost_price is not None else None
        current['cost_price_source'] = 'override'
        current['cost_price_note'] = row.note or None
        current['cost_price_updated_at'] = row.updated_at.isoformat() if row.updated_at else None
        financials[row.item_id] = current

    missing_ids = {item_id for item_id in unique_item_ids if item_id not in financials or financials[item_id].get('retail_price') is None}
    if not missing_ids:
        return financials

    token, store_id = await _get_location_ms_credentials(normalized, db)
    if not _ms_client_enabled(token, location=normalized):
        return financials

    seeds = ms_client.get_inventory_financial_seeds(normalized, store_id=store_id)
    for item_id in missing_ids:
        seed = seeds.get(item_id) or {}
        retail_price = seed.get('retail_price')
        if retail_price is None:
            continue
        current = financials.get(item_id) or {
            'cost_price': None,
            'retail_price': None,
            'cost_price_source': None,
            'cost_price_note': None,
            'cost_price_updated_at': None,
        }
        current['retail_price'] = float(retail_price)
        financials[item_id] = current
    return financials


async def _load_discrepancy_financials(location: str, results: list[CheckResult], db: AsyncSession | None = None, *, date_from: date | None = None, date_to: date | None = None) -> dict[str, dict[str, float | None | str]]:
    unique_item_ids = {
        str(row.target_id).strip()
        for row in results
        if row.target_type == 'item' and str(row.target_id or '').strip()
    }
    return await _load_cached_discrepancy_financials(location, unique_item_ids, db=db)


async def refresh_product_financial_cache(
    db: AsyncSession,
    *,
    location: str | None = None,
    force_refresh: bool = False,
) -> dict[str, Any]:
    location_query = select(LocationPoint).order_by(LocationPoint.name.asc())
    if location:
        location_query = location_query.where(LocationPoint.name == _normalize_location(location))
    points = (await db.scalars(location_query)).all()

    summary: dict[str, Any] = {
        'locations_total': len(points),
        'locations_processed': 0,
        'locations_skipped': 0,
        'items_upserted': 0,
        'items_deleted': 0,
        'locations': [],
    }
    refreshed_at = datetime.utcnow()

    for point in points:
        normalized = _normalize_location(point.name)
        token, store_id = await _get_location_ms_credentials(normalized, db)
        if not _ms_client_enabled(token, location=normalized):
            summary['locations_skipped'] += 1
            summary['locations'].append({
                'location': normalized,
                'status': 'skipped',
                'reason': 'moysklad_disabled',
                'items_upserted': 0,
                'items_deleted': 0,
            })
            continue

        inventory = await ms_client.get_inventory(normalized, token=token, store_id=store_id)
        inventory_items = _iter_inventory_items(inventory)
        item_ids = {item['id'] for item in inventory_items}
        seeds = ms_client.get_inventory_financial_seeds(normalized, store_id=store_id)
        assortment_rows = await ms_client.get_all_pages(
            'entity/assortment',
            params={'expand': 'product'},
            token=token,
            location=normalized,
            page_limit=1000,
        )

        assortment_by_id: dict[str, dict[str, Any]] = {}
        assortment_by_code: dict[str, dict[str, Any]] = {}
        for row in assortment_rows:
            row_id = _extract_meta_id(row)
            if row_id:
                assortment_by_id[row_id] = row
            row_code = str(row.get('code') or '').strip().lower()
            if row_code and row_code not in assortment_by_code:
                assortment_by_code[row_code] = row

        existing_rows = (
            await db.scalars(
                select(ProductFinancialCache)
                .where(ProductFinancialCache.location_point_id == point.id)
            )
        ).all()
        existing_by_item_id = {row.item_id: row for row in existing_rows}
        seen_item_ids: set[str] = set()
        items_upserted = 0

        for item in inventory_items:
            item_id = item['id']
            seen_item_ids.add(item_id)
            seed = seeds.get(item_id) or {}
            item_code = str(seed.get('code') or '').strip() or None
            assortment_row = assortment_by_id.get(item_id)
            if assortment_row is None and item_code:
                assortment_row = assortment_by_code.get(item_code.lower())
            product_row = assortment_row.get('product') if isinstance((assortment_row or {}).get('product'), dict) else None
            cost_price, retail_price = ms_client.extract_financials_from_sources(assortment_row, product_row)
            if retail_price in {None, 0.0}:
                retail_price = seed.get('retail_price')

            existing = existing_by_item_id.get(item_id)
            if existing is None:
                db.add(ProductFinancialCache(
                    location_point_id=point.id,
                    item_id=item_id,
                    item_name=item['name'],
                    item_code=item_code,
                    cost_price=cost_price,
                    retail_price=retail_price,
                    source_refreshed_at=refreshed_at,
                    updated_at=refreshed_at,
                ))
            else:
                if force_refresh or existing.cost_price != cost_price or existing.retail_price != retail_price or existing.item_name != item['name'] or existing.item_code != item_code:
                    existing.item_name = item['name']
                    existing.item_code = item_code
                    existing.cost_price = cost_price
                    existing.retail_price = retail_price
                    existing.source_refreshed_at = refreshed_at
                    existing.updated_at = refreshed_at
                else:
                    existing.source_refreshed_at = refreshed_at
                    existing.updated_at = refreshed_at
            items_upserted += 1

        stale_item_ids = {row.item_id for row in existing_rows if row.item_id not in seen_item_ids}
        items_deleted = 0
        if stale_item_ids:
            delete_result = await db.execute(
                delete(ProductFinancialCache)
                .where(ProductFinancialCache.location_point_id == point.id)
                .where(ProductFinancialCache.item_id.in_(sorted(stale_item_ids)))
            )
            items_deleted = int(delete_result.rowcount or 0)

        await db.commit()
        summary['locations_processed'] += 1
        summary['items_upserted'] += items_upserted
        summary['items_deleted'] += items_deleted
        summary['locations'].append({
            'location': normalized,
            'status': 'ok',
            'items_upserted': items_upserted,
            'items_deleted': items_deleted,
        })

    return summary


async def get_admin_period_report(location: str, date_from: date, date_to: date, db: AsyncSession) -> AdminReport:
    normalized = _normalize_location(location)
    if date_from > date_to:
        raise HTTPException(status_code=400, detail='Дата начала периода не может быть позже даты окончания.')

    reports = (
        await db.scalars(
            select(Report)
            .where(Report.location == normalized)
            .where(Report.report_type == DAILY_REPORT_TYPE)
            .where(Report.report_date >= date_from)
            .where(Report.report_date <= date_to)
            .order_by(Report.report_date.asc(), Report.id.asc())
        )
    ).all()

    for row in reports:
        await _sync_report_status(row, db)
    await db.commit()

    period_label = f"{date_from.strftime('%d.%m.%Y')} — {date_to.strftime('%d.%m.%Y')}"
    if not reports:
        return AdminReport(
            report_id=None,
            report_number=None,
            report_type=PERIOD_REPORT_TYPE,
            date=f'{period_label} · ревизий: 0',
            location=normalized,
            status='Нет данных за период',
            categories=[],
            selected_categories=[],
            selected_subcategories=[],
            total_plus=0.0,
            total_minus=0.0,
            total_cost=0.0,
            total_retail=0.0,
            total_lost_profit=0.0,
            can_manage_employee_completion=False,
            employees=[],
        )

    inventory = await _get_inventory_for(normalized, db=db)
    report_ids = [row.id for row in reports]
    report_snapshots = await _load_report_target_snapshots_for_report_ids(report_ids, db)
    raw_results = [
        row for row in await _load_results_for_report_ids(report_ids, db)
        if row.category_name != DEFAULT_CATEGORY_NAME and (row.subcategory_name is None or row.subcategory_name != DEFAULT_SUBCATEGORY_NAME)
    ]
    results_by_target_key: dict[tuple[str, str], CheckResult] = {}
    for row in raw_results:
        results_by_target_key[(row.target_type, row.target_id)] = row
    results = list(results_by_target_key.values())

    starts = await _load_report_employee_starts_for_report_ids(report_ids, db)
    completions = await _load_report_employee_completions_for_report_ids(report_ids, db)
    participant_user_ids = {
        int(row.user_id)
        for row in starts + completions
        if row.user_id is not None
    }
    participant_user_ids.update(
        int(row.checked_by_user_id)
        for row in results
        if row.checked_by_user_id is not None
    )

    discrepancy_financials = await _load_discrepancy_financials(normalized, results, db=db, date_from=date_from, date_to=date_to)
    historical_category_ids, historical_subcategory_ids, historical_item_ids = _report_history_target_maps(report_snapshots, results)
    report_scope_category_ids, report_scope_subcategory_ids, report_scope_category_names, report_scope_subcategory_labels = _report_selection_scope(
        report_snapshots,
        results,
    )
    selected_category_ids = report_scope_category_ids
    selected_subcategory_ids = report_scope_subcategory_ids
    target_category_names = report_scope_category_names
    target_subcategory_labels = report_scope_subcategory_labels

    active_employees = await _active_employee_users_for_location(normalized, db)
    active_employee_by_id = {employee.id: employee for employee in active_employees}
    active_employee_by_name = {employee.full_name: employee for employee in active_employees}
    report_employees = [active_employee_by_id[user_id] for user_id in sorted(participant_user_ids) if user_id in active_employee_by_id]
    can_manage_employee_completion = False
    first_activity_by_user_id, last_activity_by_user_id, first_activity_by_name, last_activity_by_name = _aggregate_employee_activity(
        starts,
        completions,
        results,
    )

    rows_by_category_target: dict[str, dict[str, CheckResult]] = defaultdict(dict)
    for row in results:
        rows_by_category_target[row.category_id][row.target_id] = row

    grouped_problem_items: dict[str, list[DiscrepancyItem]] = defaultdict(list)
    employee_bucket: dict[str, EmployeeReportSummary] = {}
    assignment_category_map, assignment_subcategory_map, assignment_item_map = _category_assignments_map([])

    def ensure_employee_bucket(full_name: str | None, user_id: int | None = None) -> EmployeeReportSummary | None:
        if not full_name:
            return None
        summary = employee_bucket.get(full_name)
        active_employee = active_employee_by_name.get(full_name)
        resolved_user_id = user_id or (active_employee.id if active_employee else None)
        first_activity = None
        last_activity = None
        if resolved_user_id is not None:
            first_activity = first_activity_by_user_id.get(resolved_user_id)
            last_activity = last_activity_by_user_id.get(resolved_user_id)
        if first_activity is None:
            first_activity = first_activity_by_name.get(full_name)
        if last_activity is None:
            last_activity = last_activity_by_name.get(full_name)

        if summary is None:
            summary = EmployeeReportSummary(
                user_id=resolved_user_id,
                full_name=full_name,
                categories=[],
                completed_categories=0,
                discrepancy_items=0,
                started_current_report=first_activity is not None,
                started_at=_format_completion_datetime(first_activity) if first_activity else None,
                finished_current_report=last_activity is not None,
                finished_at=_format_completion_datetime(last_activity) if last_activity else None,
                can_reopen_access=False,
            )
            employee_bucket[full_name] = summary
            return summary

        if summary.user_id is None and resolved_user_id is not None:
            summary.user_id = resolved_user_id
        if first_activity is not None:
            summary.started_current_report = True
            summary.started_at = _format_completion_datetime(first_activity)
        if last_activity is not None:
            summary.finished_current_report = True
            summary.finished_at = _format_completion_datetime(last_activity)
        summary.can_reopen_access = False
        return summary

    for employee in report_employees:
        ensure_employee_bucket(employee.full_name, employee.id)

    for row in starts:
        ensure_employee_bucket(row.user_full_name_snapshot, row.user_id)
    for row in completions:
        ensure_employee_bucket(row.user_full_name_snapshot, row.user_id)

    for row in results:
        if row.checked_by_name_snapshot:
            bucket = ensure_employee_bucket(row.checked_by_name_snapshot, row.checked_by_user_id)
            if bucket and row.category_name not in bucket.categories:
                bucket.categories.append(row.category_name)

        if row.target_type == 'item' and row.status == 'red':
            financials = discrepancy_financials.get(row.target_id, {})
            diff_qty = abs(float(row.diff or 0))
            cost_price = financials.get('cost_price')
            retail_price = financials.get('retail_price')
            cost_total = round(diff_qty * cost_price, 2) if cost_price is not None else None
            retail_total = round(diff_qty * retail_price, 2) if retail_price is not None else None
            lost_profit = None
            if cost_total is not None and retail_total is not None:
                lost_profit = round(retail_total - cost_total, 2)

            if row.checked_by_name_snapshot:
                bucket = ensure_employee_bucket(row.checked_by_name_snapshot, row.checked_by_user_id)
                if bucket:
                    bucket.discrepancy_items += 1
                    bucket.total_cost = float(round(float(bucket.total_cost or 0.0) + float(cost_total or 0.0), 2))
                    bucket.total_retail = float(round(float(bucket.total_retail or 0.0) + float(retail_total or 0.0), 2))
                    bucket.total_lost_profit = float(round(float(bucket.total_lost_profit or 0.0) + float(lost_profit or 0.0), 2))

            grouped_problem_items[row.category_name].append(
                DiscrepancyItem(
                    check_result_id=row.id,
                    category_name=row.category_name,
                    name=row.target_name,
                    expected=float(row.expected_qty),
                    actual=float(row.actual_qty or 0),
                    diff=float(row.diff or 0),
                    checked_by=row.checked_by_name_snapshot,
                    subcategory_name=row.subcategory_name,
                    cost_price=cost_price,
                    retail_price=retail_price,
                    cost_total=cost_total,
                    retail_total=retail_total,
                    lost_profit=lost_profit,
                    cost_price_source=str(financials.get('cost_price_source') or '') or None,
                    cost_price_note=str(financials.get('cost_price_note') or '') or None,
                    cost_price_updated_at=str(financials.get('cost_price_updated_at') or '') or None,
                )
            )

    categories: list[CategoryResult] = []
    full_inventory_by_category_id = {category['id']: category for category in inventory['categories']}
    full_inventory_by_category_name = {category['name']: category for category in inventory['categories']}
    inventory = _filter_inventory_by_targets(
        inventory,
        selected_category_ids,
        selected_subcategory_ids,
        retained_category_ids=historical_category_ids,
        retained_subcategory_ids=historical_subcategory_ids,
        retained_item_ids=historical_item_ids,
    )
    for raw_category in inventory['categories']:
        result_map = rows_by_category_target.get(raw_category['id'], {})
        is_completed, status = (True, StatusEnum.GREEN) if _category_is_complete(raw_category, result_map) else (False, StatusEnum.GREY)
        if not is_completed:
            statuses = [_subcategory_is_complete(sub, result_map)[1] for sub in raw_category['subcategories']]
            if any(s == StatusEnum.RED for s in statuses):
                status = StatusEnum.RED
            elif any(s == StatusEnum.ORANGE for s in statuses):
                status = StatusEnum.ORANGE
            elif any(s == StatusEnum.GREEN for s in statuses):
                status = StatusEnum.ORANGE
            else:
                status = StatusEnum.GREY
        elif grouped_problem_items.get(raw_category['name']):
            status = StatusEnum.RED

        selected_sub_ids_for_category = selected_subcategory_ids.get(raw_category['id'], set())
        selected_sub_names = sorted([
            sub['name']
            for sub in raw_category['subcategories']
            if sub['id'] in selected_sub_ids_for_category
        ])
        category_assignment = assignment_category_map.get(raw_category['id'])
        sub_assignments = assignment_subcategory_map.get(raw_category['id'], {})
        item_assignments_by_sub = assignment_item_map.get(raw_category['id'], {})
        completed_subcategories: list[CompletedSubcategoryInfo] = []
        in_progress_subcategories: list[InProgressSubcategoryInfo] = []
        source_category = full_inventory_by_category_id.get(raw_category['id'], raw_category)
        category_taken_whole = bool(category_assignment) or _category_taken_in_report(raw_category['id'], report_snapshots)
        inferred_category_owner: str | None = None

        detail_subcategories = raw_category['subcategories']
        category_snapshot_owner = _category_snapshot_assignment_label(raw_category['id'], report_snapshots) if category_taken_whole else None
        if not category_snapshot_owner and inferred_category_owner:
            category_snapshot_owner = inferred_category_owner

        for raw_sub in detail_subcategories:
            if _is_categoryless_subcategory(source_category, raw_sub):
                continue

            sub_completed, sub_status = _subcategory_is_complete(raw_sub, result_map)
            sub_assignment = sub_assignments.get(raw_sub['id'])
            item_assignments = item_assignments_by_sub.get(raw_sub['id'], {})
            sub_taken_in_report = (
                category_taken_whole
                or bool(category_assignment or sub_assignment or item_assignments)
                or _subcategory_taken_in_report(raw_category['id'], raw_sub['id'], report_snapshots)
            )

            if sub_completed:
                if sub_status == StatusEnum.GREEN:
                    sub_row = result_map.get(raw_sub['id'])
                    checked_by = sub_row.checked_by_name_snapshot if sub_row else None
                    if not checked_by:
                        item_rows = [result_map.get(item['id']) for item in raw_sub['items']]
                        item_rows = [row for row in item_rows if row and row.checked_by_name_snapshot]
                        if item_rows:
                            owners = {row.checked_by_name_snapshot for row in item_rows if row.checked_by_name_snapshot}
                            checked_by = _owner_label(owners)
                    completed_subcategories.append(CompletedSubcategoryInfo(
                        name=raw_sub['name'],
                        checked_by=checked_by,
                        status=sub_status,
                    ))
                continue

            show_as_taken_for_detail = category_taken_whole or sub_taken_in_report
            if not show_as_taken_for_detail:
                continue

            assigned_to_label = (
                category_assignment.user_full_name_snapshot
                if category_assignment and category_assignment.user_full_name_snapshot
                else (category_snapshot_owner or (sub_assignment.user_full_name_snapshot if sub_assignment and sub_assignment.user_full_name_snapshot else None))
            )
            if not assigned_to_label and item_assignments:
                assigned_to_label = _owner_label({
                    assignment.user_full_name_snapshot
                    for assignment in item_assignments.values()
                    if assignment.user_full_name_snapshot
                })
            if not assigned_to_label:
                assigned_to_label = _subcategory_snapshot_assignment_label(raw_category['id'], raw_sub['id'], report_snapshots)

            in_progress_subcategories.append(InProgressSubcategoryInfo(
                name=raw_sub['name'],
                assigned_to=assigned_to_label,
            ))

        completed_subcategories.sort(key=lambda item: item.name.lower())
        in_progress_subcategories.sort(key=lambda item: item.name.lower())
        remaining_subcategories = [item.name for item in in_progress_subcategories]

        categories.append(CategoryResult(
            name=raw_category['name'],
            status=status,
            assigned_to=(
                _category_assignment_label(raw_category['id'], [])
                or _category_snapshot_assignment_label(raw_category['id'], report_snapshots)
                or inferred_category_owner
            ),
            selected_on_cycle=raw_category['id'] in selected_category_ids,
            selected_subcategories=selected_sub_names,
            remaining_subcategories=remaining_subcategories,
            in_progress_subcategories=in_progress_subcategories,
            completed_subcategories=completed_subcategories,
            problem_items=grouped_problem_items.get(raw_category['name'], []),
        ))

    for category in categories:
        owners = {item.checked_by for item in category.problem_items if item.checked_by}
        if len(owners) == 1:
            owner = next(iter(owners))
            owner_bucket = ensure_employee_bucket(owner)
            if owner_bucket and category.name not in owner_bucket.categories:
                owner_bucket.categories.append(category.name)

    for summary in employee_bucket.values():
        summary.categories = sorted(summary.categories, key=str.lower)
        summary.completed_categories = sum(1 for category in categories if category.assigned_to == summary.full_name and category.status in {StatusEnum.GREEN, StatusEnum.RED})
        summary.total_cost = float(round(float(summary.total_cost or 0.0), 2))
        summary.total_retail = float(round(float(summary.total_retail or 0.0), 2))
        summary.total_lost_profit = float(round(float(summary.total_lost_profit or 0.0), 2))

    total_plus = sum(max(float(item.diff), 0.0) for items in grouped_problem_items.values() for item in items)
    total_minus = abs(sum(min(float(item.diff), 0.0) for items in grouped_problem_items.values() for item in items))
    total_cost = sum(float(item.cost_total or 0.0) for items in grouped_problem_items.values() for item in items)
    total_retail = sum(float(item.retail_total or 0.0) for items in grouped_problem_items.values() for item in items)
    total_lost_profit = sum(float(item.lost_profit or 0.0) for items in grouped_problem_items.values() for item in items)
    total_subcategories, completed_subcategories_count, discrepancy_subcategories_count, no_discrepancy_subcategories_count = _admin_report_subcategory_stats(
        categories,
        full_inventory_by_category_name,
        report_type=PERIOD_REPORT_TYPE,
    )

    return AdminReport(
        report_id=None,
        report_number=None,
        report_type=PERIOD_REPORT_TYPE,
        date=f'{period_label} · ревизий: {len(reports)}',
        location=normalized,
        status='Период',
        categories=categories,
        selected_categories=target_category_names,
        selected_subcategories=target_subcategory_labels,
        total_subcategories=total_subcategories,
        completed_subcategories_count=completed_subcategories_count,
        discrepancy_subcategories_count=discrepancy_subcategories_count,
        no_discrepancy_subcategories_count=no_discrepancy_subcategories_count,
        total_plus=float(total_plus),
        total_minus=float(total_minus),
        total_cost=float(round(total_cost, 2)),
        total_retail=float(round(total_retail, 2)),
        total_lost_profit=float(round(total_lost_profit, 2)),
        can_manage_employee_completion=can_manage_employee_completion,
        employees=sorted(employee_bucket.values(), key=lambda item: item.full_name.lower()),
    )


async def get_admin_report(location: str, db: AsyncSession, report_id: int | None = None) -> AdminReport:
    normalized = _normalize_location(location)
    report: Report | None = None
    if report_id is not None:
        report = await db.get(Report, report_id)
        if report and report.location != normalized:
            report = None
    if report is None:
        report = await db.scalar(select(Report).where(Report.location == normalized).order_by(Report.date_created.desc(), Report.id.desc()).limit(1))

    if not report:
        return AdminReport(report_id=None, report_number=None, report_type=DAILY_REPORT_TYPE, date='-', location=normalized, status='-', categories=[], total_plus=0.0, total_minus=0.0, can_manage_employee_completion=False, employees=[])

    await _sync_report_status(report, db)
    await db.commit()

    targets = await _resolve_selection_targets_for_date(normalized, report.cycle_version, report.report_date, db)
    inventory = await _get_inventory_for(normalized, db=db)

    report_type = report.report_type or DAILY_REPORT_TYPE
    assignments: list[CategoryAssignment] = []
    is_latest_daily_report = False
    if report_type == FINAL_REPORT_TYPE:
        cycle_reports = (
            await db.scalars(
                select(Report).where(
                    Report.location == normalized,
                    Report.cycle_version == report.cycle_version,
                    Report.report_type == DAILY_REPORT_TYPE,
                ).order_by(Report.date_created.asc(), Report.id.asc())
            )
        ).all()
        cycle_report_ids = [row.id for row in cycle_reports]
        report_snapshots = await _load_report_target_snapshots_for_report_ids(cycle_report_ids, db)
        raw_results = [
            row for row in await _load_results_for_report_ids(cycle_report_ids, db)
            if row.category_name != DEFAULT_CATEGORY_NAME and (row.subcategory_name is None or row.subcategory_name != DEFAULT_SUBCATEGORY_NAME)
        ]
        results_by_target_key: dict[tuple[str, str], CheckResult] = {}
        for row in raw_results:
            results_by_target_key[(row.target_type, row.target_id)] = row
        results = list(results_by_target_key.values())
    else:
        latest_daily_report_id = await db.scalar(
            select(Report.id)
            .where(Report.location == normalized)
            .where(Report.report_type == DAILY_REPORT_TYPE)
            .order_by(Report.date_created.desc(), Report.id.desc())
            .limit(1)
        )
        is_latest_daily_report = latest_daily_report_id == report.id
        results = [
            row for row in await _load_results(report.id, db)
            if row.category_name != DEFAULT_CATEGORY_NAME and (row.subcategory_name is None or row.subcategory_name != DEFAULT_SUBCATEGORY_NAME)
        ]
        if is_latest_daily_report:
            assignments = await _load_assignments(report.location, report.cycle_version, db)
            report_snapshots = await _bootstrap_report_target_snapshots(report, assignments, results, db)
        else:
            assignments = []
            report_snapshots = _sanitize_historical_daily_report_snapshots(
                report,
                await _load_report_target_snapshots(report.id, db),
                results,
            )

    discrepancy_financials = await _load_discrepancy_financials(normalized, results, db=db, date_from=report.report_date, date_to=report.report_date)
    completed_before_report: dict[str, set[str]] = {}
    if report_type == DAILY_REPORT_TYPE:
        completed_before_report = await _load_completed_subcategory_ids_for_cycle(
            normalized,
            report.cycle_version,
            inventory,
            db,
            before_report_date=report.report_date,
        )
    participant_user_ids = await _get_report_participant_user_ids(report.id, db) if report_type == DAILY_REPORT_TYPE else set()
    if participant_user_ids and report_type == DAILY_REPORT_TYPE:
        report_snapshots = [
            row for row in report_snapshots
            if row.assigned_user_id_snapshot is None or row.assigned_user_id_snapshot in participant_user_ids
        ]
    historical_category_ids, historical_subcategory_ids, historical_item_ids = _report_history_target_maps(report_snapshots, results)
    current_selected_category_ids, current_selected_subcategory_ids = _selection_target_maps(targets)
    current_target_category_names = sorted({target.category_name for target in targets if target.target_type == 'category'})
    current_target_subcategory_labels = sorted({
        f"{target.category_name} → {target.subcategory_name}"
        for target in targets
        if target.target_type == 'subcategory' and target.subcategory_name
    })
    report_scope_category_ids, report_scope_subcategory_ids, report_scope_category_names, report_scope_subcategory_labels = _report_selection_scope(
        report_snapshots,
        results,
    )
    use_report_scope_for_selection = report_type == DAILY_REPORT_TYPE and not is_latest_daily_report
    selected_category_ids = report_scope_category_ids if use_report_scope_for_selection else current_selected_category_ids
    selected_subcategory_ids = report_scope_subcategory_ids if use_report_scope_for_selection else current_selected_subcategory_ids
    target_category_names = report_scope_category_names if use_report_scope_for_selection else current_target_category_names
    target_subcategory_labels = report_scope_subcategory_labels if use_report_scope_for_selection else current_target_subcategory_labels
    active_employees = await _active_employee_users_for_location(report.location, db)
    active_employee_by_id = {employee.id: employee for employee in active_employees}
    report_employees = [active_employee_by_id[user_id] for user_id in sorted(participant_user_ids) if user_id in active_employee_by_id]
    completions = await _load_report_employee_completions(report.id, db)
    starts = await _load_report_employee_starts(report.id, db)
    can_manage_employee_completion = (report.report_type or DAILY_REPORT_TYPE) == DAILY_REPORT_TYPE and report.report_date == get_moscow_today()

    rows_by_category_target: dict[str, dict[str, CheckResult]] = defaultdict(dict)
    for row in results:
        rows_by_category_target[row.category_id][row.target_id] = row

    completion_by_user_id = {completion.user_id: completion for completion in completions}
    start_by_user_id = {start.user_id: start for start in starts}
    active_employee_by_name = {employee.full_name: employee for employee in active_employees}
    grouped_problem_items: dict[str, list[DiscrepancyItem]] = defaultdict(list)
    employee_bucket: dict[str, EmployeeReportSummary] = {}
    assignment_category_map, assignment_subcategory_map, assignment_item_map = _category_assignments_map(assignments)

    def ensure_employee_bucket(full_name: str | None, user_id: int | None = None) -> EmployeeReportSummary | None:
        if not full_name:
            return None
        summary = employee_bucket.get(full_name)
        active_employee = active_employee_by_name.get(full_name)
        resolved_user_id = user_id or (active_employee.id if active_employee else None)
        completion = completion_by_user_id.get(resolved_user_id) if resolved_user_id else None
        start = start_by_user_id.get(resolved_user_id) if resolved_user_id else None
        if summary is None:
            summary = EmployeeReportSummary(
                user_id=resolved_user_id,
                full_name=full_name,
                categories=[],
                completed_categories=0,
                discrepancy_items=0,
                started_current_report=start is not None,
                started_at=_format_completion_datetime(start.started_at) if start else None,
                finished_current_report=completion is not None,
                finished_at=_format_completion_datetime(completion.finished_at) if completion else None,
                can_reopen_access=bool(can_manage_employee_completion and completion is not None and resolved_user_id),
            )
            employee_bucket[full_name] = summary
            return summary

        if summary.user_id is None and resolved_user_id is not None:
            summary.user_id = resolved_user_id
        if start is not None:
            summary.started_current_report = True
            summary.started_at = _format_completion_datetime(start.started_at)
        if completion is not None:
            summary.finished_current_report = True
            summary.finished_at = _format_completion_datetime(completion.finished_at)
        summary.can_reopen_access = bool(can_manage_employee_completion and summary.finished_current_report and summary.user_id)
        return summary

    for employee in report_employees:
        ensure_employee_bucket(employee.full_name, employee.id)

    for row in results:
        if row.checked_by_name_snapshot:
            bucket = ensure_employee_bucket(row.checked_by_name_snapshot, row.checked_by_user_id)
            if bucket and row.category_name not in bucket.categories:
                bucket.categories.append(row.category_name)

        if row.target_type == 'item' and row.status == 'red':
            financials = discrepancy_financials.get(row.target_id, {})
            diff_qty = abs(float(row.diff or 0))
            cost_price = financials.get('cost_price')
            retail_price = financials.get('retail_price')
            cost_total = round(diff_qty * cost_price, 2) if cost_price is not None else None
            retail_total = round(diff_qty * retail_price, 2) if retail_price is not None else None
            lost_profit = None
            if cost_total is not None and retail_total is not None:
                lost_profit = round(retail_total - cost_total, 2)

            if row.checked_by_name_snapshot:
                bucket = ensure_employee_bucket(row.checked_by_name_snapshot, row.checked_by_user_id)
                if bucket:
                    bucket.discrepancy_items += 1
                    bucket.total_cost = float(round(float(bucket.total_cost or 0.0) + float(cost_total or 0.0), 2))
                    bucket.total_retail = float(round(float(bucket.total_retail or 0.0) + float(retail_total or 0.0), 2))
                    bucket.total_lost_profit = float(round(float(bucket.total_lost_profit or 0.0) + float(lost_profit or 0.0), 2))

            grouped_problem_items[row.category_name].append(
                DiscrepancyItem(
                    check_result_id=row.id,
                    category_name=row.category_name,
                    name=row.target_name,
                    expected=float(row.expected_qty),
                    actual=float(row.actual_qty or 0),
                    diff=float(row.diff or 0),
                    checked_by=row.checked_by_name_snapshot,
                    subcategory_name=row.subcategory_name,
                    cost_price=cost_price,
                    retail_price=retail_price,
                    cost_total=cost_total,
                    retail_total=retail_total,
                    lost_profit=lost_profit,
                    cost_price_source=str(financials.get('cost_price_source') or '') or None,
                    cost_price_note=str(financials.get('cost_price_note') or '') or None,
                    cost_price_updated_at=str(financials.get('cost_price_updated_at') or '') or None,
                )
            )

    categories: list[CategoryResult] = []
    full_inventory_by_category_id = {category['id']: category for category in inventory['categories']}
    full_inventory_by_category_name = {category['name']: category for category in inventory['categories']}
    inventory = _filter_inventory_by_targets(
        inventory,
        selected_category_ids,
        selected_subcategory_ids,
        retained_category_ids=historical_category_ids,
        retained_subcategory_ids=historical_subcategory_ids,
        retained_item_ids=historical_item_ids,
    )
    for raw_category in inventory['categories']:
        result_map = rows_by_category_target.get(raw_category['id'], {})
        is_completed, status = (True, StatusEnum.GREEN) if _category_is_complete(raw_category, result_map) else (False, StatusEnum.GREY)
        if not is_completed:
            statuses = [_subcategory_is_complete(sub, result_map)[1] for sub in raw_category['subcategories']]
            if any(s == StatusEnum.RED for s in statuses):
                status = StatusEnum.RED
            elif any(s == StatusEnum.ORANGE for s in statuses):
                status = StatusEnum.ORANGE
            elif any(s == StatusEnum.GREEN for s in statuses):
                status = StatusEnum.ORANGE
            else:
                status = StatusEnum.GREY
        elif grouped_problem_items.get(raw_category['name']):
            status = StatusEnum.RED

        selected_sub_ids_for_category = selected_subcategory_ids.get(raw_category['id'], set())
        report_scope_sub_ids_for_category = report_scope_subcategory_ids.get(raw_category['id'], set())
        selected_sub_names = sorted([
            sub['name']
            for sub in raw_category['subcategories']
            if sub['id'] in selected_sub_ids_for_category
        ])
        category_assignment = assignment_category_map.get(raw_category['id'])
        sub_assignments = assignment_subcategory_map.get(raw_category['id'], {})
        item_assignments_by_sub = assignment_item_map.get(raw_category['id'], {})
        completed_subcategories: list[CompletedSubcategoryInfo] = []
        in_progress_subcategories: list[InProgressSubcategoryInfo] = []
        source_category = full_inventory_by_category_id.get(raw_category['id'], raw_category)
        category_selected_on_cycle = raw_category['id'] in selected_category_ids
        category_selected_in_report_scope = raw_category['id'] in report_scope_category_ids
        category_taken_whole = bool(category_assignment) or _category_taken_in_report(raw_category['id'], report_snapshots)
        inferred_category_owner: str | None = None

        if report_type == DAILY_REPORT_TYPE and category_selected_on_cycle and not category_taken_whole:
            category_owner_candidates: set[str] = set()
            if category_assignment and category_assignment.user_full_name_snapshot:
                category_owner_candidates.add(category_assignment.user_full_name_snapshot)
            category_owner_candidates.update(
                assignment.user_full_name_snapshot
                for assignment in sub_assignments.values()
                if assignment.user_full_name_snapshot
            )
            category_owner_candidates.update(
                assignment.user_full_name_snapshot
                for assignments_by_item in item_assignments_by_sub.values()
                for assignment in assignments_by_item.values()
                if assignment.user_full_name_snapshot
            )
            category_owner_candidates.update(
                row.assigned_user_name_snapshot
                for row in report_snapshots
                if row.category_id == raw_category['id'] and row.assigned_user_name_snapshot
            )
            category_owner_candidates.update(
                row.checked_by_name_snapshot
                for row in result_map.values()
                if row and row.checked_by_name_snapshot
            )
            category_owner_candidates = {name for name in category_owner_candidates if name}
            if len(category_owner_candidates) == 1:
                inferred_category_owner = next(iter(category_owner_candidates))
                category_taken_whole = True

        detail_subcategories = raw_category['subcategories']
        if report_type == DAILY_REPORT_TYPE and category_taken_whole:
            historical_completed_ids = completed_before_report.get(raw_category['id'], set())
            detail_subcategories = [
                sub
                for sub in source_category.get('subcategories', [])
                if (
                    not _is_categoryless_subcategory(source_category, sub)
                    and sub['id'] not in historical_completed_ids
                    and (
                        category_selected_on_cycle
                        or category_selected_in_report_scope
                        or sub['id'] in selected_sub_ids_for_category
                        or sub['id'] in report_scope_sub_ids_for_category
                    )
                )
            ]

        category_snapshot_owner = _category_snapshot_assignment_label(raw_category['id'], report_snapshots) if category_taken_whole else None
        if not category_snapshot_owner and inferred_category_owner:
            category_snapshot_owner = inferred_category_owner

        problem_subcategory_ids = {
            row.subcategory_id
            for row in results
            if row.category_id == raw_category['id'] and row.subcategory_id and row.status in {'red', 'orange'}
        }
        completed_green_subcategory_ids: set[str] = set()
        taken_owner_by_subcategory_id: dict[str, str | None] = {}
        source_sub_by_id = {
            sub['id']: sub
            for sub in source_category.get('subcategories', [])
            if not _is_categoryless_subcategory(source_category, sub)
        }
        subcategory_name_by_id = _subcategory_name_lookup(
            raw_category['id'],
            source_category,
            targets,
            sub_assignments,
            item_assignments_by_sub,
            report_snapshots,
            results,
        )
        detail_subcategory_ids = {
            sub['id']
            for sub in detail_subcategories
            if not _is_categoryless_subcategory(source_category, sub)
        }
        relevant_subcategory_ids = set(detail_subcategory_ids)
        relevant_subcategory_ids.update(selected_sub_ids_for_category)
        relevant_subcategory_ids.update(report_scope_sub_ids_for_category)
        relevant_subcategory_ids.update(sub_assignments.keys())
        relevant_subcategory_ids.update(item_assignments_by_sub.keys())
        relevant_subcategory_ids.update(
            row.subcategory_id
            for row in report_snapshots
            if row.category_id == raw_category['id'] and row.subcategory_id
        )
        relevant_subcategory_ids.update(
            row.subcategory_id
            for row in results
            if row.category_id == raw_category['id'] and row.subcategory_id
        )
        relevant_subcategory_ids = {
            sub_id
            for sub_id in relevant_subcategory_ids
            if sub_id and (sub_id in subcategory_name_by_id or sub_id in source_sub_by_id)
        }

        if report_type == DAILY_REPORT_TYPE and category_taken_whole:
            historical_completed_ids = completed_before_report.get(raw_category['id'], set())
            relevant_subcategory_ids.difference_update(historical_completed_ids)

        def remember_taken_subcategory(sub_id: str, owner_label: str | None) -> None:
            if sub_id in taken_owner_by_subcategory_id:
                return
            taken_owner_by_subcategory_id[sub_id] = owner_label

        if category_taken_whole:
            category_level_owner = (
                category_assignment.user_full_name_snapshot
                if category_assignment and category_assignment.user_full_name_snapshot
                else category_snapshot_owner
            )
            for sub_id in sorted(detail_subcategory_ids, key=lambda value: (subcategory_name_by_id.get(value) or value).lower()):
                remember_taken_subcategory(sub_id, category_level_owner)

        for sub_id in sorted(relevant_subcategory_ids, key=lambda value: (subcategory_name_by_id.get(value) or value).lower()):
            raw_sub = source_sub_by_id.get(sub_id)
            sub_name = subcategory_name_by_id.get(sub_id) or (raw_sub['name'] if raw_sub else None)
            if not sub_name:
                continue

            sub_completed, sub_status = _subcategory_completion_status_for_admin(raw_sub, sub_id, sub_name, result_map)
            sub_assignment = sub_assignments.get(sub_id)
            item_assignments = item_assignments_by_sub.get(sub_id, {})
            sub_snapshot_owner = _subcategory_snapshot_assignment_label(raw_category['id'], sub_id, report_snapshots)
            sub_taken_in_report = (
                category_taken_whole
                or bool(category_assignment or sub_assignment or item_assignments)
                or _subcategory_taken_in_report(raw_category['id'], sub_id, report_snapshots)
            )

            if sub_taken_in_report:
                assigned_to_label = (
                    category_assignment.user_full_name_snapshot
                    if category_assignment and category_assignment.user_full_name_snapshot
                    else (category_snapshot_owner or (sub_assignment.user_full_name_snapshot if sub_assignment and sub_assignment.user_full_name_snapshot else None))
                )
                if not assigned_to_label and item_assignments:
                    assigned_to_label = _owner_label({
                        assignment.user_full_name_snapshot
                        for assignment in item_assignments.values()
                        if assignment.user_full_name_snapshot
                    })
                if not assigned_to_label:
                    assigned_to_label = sub_snapshot_owner
                remember_taken_subcategory(sub_id, assigned_to_label)

            if sub_completed and sub_status == StatusEnum.GREEN:
                sub_row = result_map.get(sub_id)
                checked_by = sub_row.checked_by_name_snapshot if sub_row else None
                if not checked_by and raw_sub is not None:
                    item_rows = [result_map.get(item['id']) for item in raw_sub['items']]
                    item_rows = [row for row in item_rows if row and row.checked_by_name_snapshot]
                    if item_rows:
                        owners = {row.checked_by_name_snapshot for row in item_rows if row.checked_by_name_snapshot}
                        checked_by = _owner_label(owners)
                completed_green_subcategory_ids.add(sub_id)
                completed_subcategories.append(CompletedSubcategoryInfo(
                    name=sub_name,
                    checked_by=checked_by,
                    status=sub_status,
                ))

        for sub_id in sorted(relevant_subcategory_ids, key=lambda value: (subcategory_name_by_id.get(value) or value).lower()):
            if sub_id not in taken_owner_by_subcategory_id:
                continue
            if sub_id in completed_green_subcategory_ids or sub_id in problem_subcategory_ids:
                continue
            sub_name = subcategory_name_by_id.get(sub_id) or (source_sub_by_id.get(sub_id) or {}).get('name')
            if not sub_name:
                continue
            in_progress_subcategories.append(InProgressSubcategoryInfo(
                name=sub_name,
                assigned_to=taken_owner_by_subcategory_id.get(sub_id),
            ))

        completed_subcategories.sort(key=lambda item: item.name.lower())
        in_progress_subcategories.sort(key=lambda item: item.name.lower())
        remaining_subcategories = [item.name for item in in_progress_subcategories]

        categories.append(CategoryResult(
            name=raw_category['name'],
            status=status,
            assigned_to=(
                _category_assignment_label(raw_category['id'], assignments)
                or _category_snapshot_assignment_label(raw_category['id'], report_snapshots)
                or inferred_category_owner
            ),
            selected_on_cycle=raw_category['id'] in selected_category_ids,
            selected_subcategories=selected_sub_names,
            remaining_subcategories=remaining_subcategories,
            in_progress_subcategories=in_progress_subcategories,
            completed_subcategories=completed_subcategories,
            problem_items=grouped_problem_items.get(raw_category['name'], []),
        ))

    for category in categories:
        owners = {item.checked_by for item in category.problem_items if item.checked_by}
        if len(owners) == 1:
            owner = next(iter(owners))
            owner_bucket = ensure_employee_bucket(owner)
            if owner_bucket and category.name not in owner_bucket.categories:
                owner_bucket.categories.append(category.name)

    for summary in employee_bucket.values():
        summary.categories = sorted(summary.categories, key=str.lower)
        summary.completed_categories = sum(1 for category in categories if category.assigned_to == summary.full_name and category.status in {StatusEnum.GREEN, StatusEnum.RED})
        summary.total_cost = float(round(float(summary.total_cost or 0.0), 2))
        summary.total_retail = float(round(float(summary.total_retail or 0.0), 2))
        summary.total_lost_profit = float(round(float(summary.total_lost_profit or 0.0), 2))

    total_plus = sum(max(float(item.diff), 0.0) for items in grouped_problem_items.values() for item in items)
    total_minus = abs(sum(min(float(item.diff), 0.0) for items in grouped_problem_items.values() for item in items))
    total_cost = sum(float(item.cost_total or 0.0) for items in grouped_problem_items.values() for item in items)
    total_retail = sum(float(item.retail_total or 0.0) for items in grouped_problem_items.values() for item in items)
    total_lost_profit = sum(float(item.lost_profit or 0.0) for items in grouped_problem_items.values() for item in items)
    total_subcategories, completed_subcategories_count, discrepancy_subcategories_count, no_discrepancy_subcategories_count = _admin_report_subcategory_stats(
        categories,
        full_inventory_by_category_name,
        completed_before_report=completed_before_report,
        report_type=report_type,
    )

    report_number = await _get_report_number(report, db)

    return AdminReport(
        report_id=report.id,
        report_number=report_number or None,
        report_type=report_type,
        date=_format_moscow_datetime(report.date_created),
        location=report.location,
        status=_report_status_label(report.status),
        categories=categories,
        selected_categories=target_category_names,
        selected_subcategories=target_subcategory_labels,
        total_subcategories=total_subcategories,
        completed_subcategories_count=completed_subcategories_count,
        discrepancy_subcategories_count=discrepancy_subcategories_count,
        no_discrepancy_subcategories_count=no_discrepancy_subcategories_count,
        total_plus=float(total_plus),
        total_minus=float(total_minus),
        total_cost=float(round(total_cost, 2)),
        total_retail=float(round(total_retail, 2)),
        total_lost_profit=float(round(total_lost_profit, 2)),
        can_manage_employee_completion=can_manage_employee_completion,
        employees=sorted(employee_bucket.values(), key=lambda item: item.full_name.lower()),
    )

def _excel_yes_no(value: bool) -> str:
    return 'Да' if value else 'Нет'


def _safe_excel_float(value: Any) -> float:
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        return 0.0
    if abs(number) < 1e-9:
        return 0.0
    return float(round(number, 2))


def _safe_excel_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _safe_excel_text(value: Any, fallback: str = '—') -> str:
    if value is None:
        return fallback
    if hasattr(value, 'value') and isinstance(getattr(value, 'value'), str):
        value = getattr(value, 'value')
    text = str(value).strip()
    return text or fallback


def _safe_filename_part(value: Any, fallback: str) -> str:
    text = ' '.join(str(value or '').strip().split())
    for bad in '\\/:*?"<>|':
        text = text.replace(bad, '-')
    text = text.strip(' .-_')
    return text or fallback


def _build_export_filename(report: AdminReport) -> str:
    location_part = _safe_filename_part(report.location, 'Точка')
    date_part = _safe_filename_part(report.date, 'Дата')
    if report.report_type == PERIOD_REPORT_TYPE:
        suffix = 'период'
    elif report.report_type == FINAL_REPORT_TYPE:
        suffix = 'итоговая'
    else:
        suffix = f"ревизия {report.report_number or report.report_id}"
    return f"{location_part} {date_part} {suffix}.xlsx"


def _init_export_sheet(ws, title: str, headers: list[str]) -> None:
    ws.title = title
    ws.sheet_view.showGridLines = True
    ws.append(headers)

    header_fill = PatternFill(fill_type='solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(bottom=Side(style='thin', color='D0D7DE'))

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    ws.freeze_panes = 'A2'


def _finalize_export_sheet(
    ws,
    currency_columns: set[int] | None = None,
    quantity_columns: set[int] | None = None,
    *,
    left_align_columns: set[int] | None = None,
    width_overrides: dict[str, float] | None = None,
    max_width: int = 40,
    default_horizontal: str = 'center',
    default_vertical: str = 'center',
) -> None:
    currency_columns = currency_columns or set()
    quantity_columns = quantity_columns or set()
    left_align_columns = left_align_columns or set()
    width_overrides = {str(key).upper(): value for key, value in (width_overrides or {}).items()}
    clear_fill = PatternFill(fill_type=None)
    wrap_alignment = Alignment(horizontal=default_horizontal, vertical=default_vertical, wrap_text=True)
    wrap_alignment_left = Alignment(horizontal='left', vertical=default_vertical, wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.fill = clear_fill
            cell.alignment = wrap_alignment_left if cell.column in left_align_columns else wrap_alignment
            if cell.column in currency_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00 "₽"'
            elif cell.column in quantity_columns and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.###'

    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            if len(value) > length:
                length = len(value)
        auto_width = min(max(length + 2, 12), max_width)
        ws.column_dimensions[column_letter].width = float(width_overrides.get(column_letter, auto_width))


def _status_fill(value: Any) -> PatternFill | None:
    normalized = _safe_excel_text(value, '').lower()
    if normalized in {'green', 'успешно', 'завершена', 'завершен', 'completed'}:
        return PatternFill(fill_type='solid', fgColor='EAF7EE')
    if normalized in {'red', 'с расхождением', 'расхождения'}:
        return PatternFill(fill_type='solid', fgColor='FCEBEC')
    if normalized in {'orange', 'в работе', 'active', 'in_progress'}:
        return PatternFill(fill_type='solid', fgColor='FFF4DB')
    if normalized in {'gray', 'grey', 'не начата', 'not_started'}:
        return PatternFill(fill_type='solid', fgColor='F3F4F6')
    return None


def _style_summary_sheet(ws) -> None:
    label_fill = PatternFill(fill_type='solid', fgColor='F8FAFC')
    emphasis_fills = {
        'Излишки': PatternFill(fill_type='solid', fgColor='EAF7EE'),
        'Недостача': PatternFill(fill_type='solid', fgColor='FCEBEC'),
        'Себестоимость': PatternFill(fill_type='solid', fgColor='EEF4FB'),
        'Розница': PatternFill(fill_type='solid', fgColor='F3EEFF'),
        'Утерянная прибыль': PatternFill(fill_type='solid', fgColor='FFF4DB'),
    }
    info_rows = {'Точка', 'Дата отчёта', 'Тип ревизии', 'Номер ревизии', 'Статус'}
    money_rows = set(emphasis_fills)

    for row_index in range(2, ws.max_row + 1):
        label_cell = ws.cell(row=row_index, column=1)
        value_cell = ws.cell(row=row_index, column=2)
        label = _safe_excel_text(label_cell.value, '')

        label_cell.fill = label_fill
        label_cell.font = Font(bold=True, color='243B53')
        if label in info_rows:
            value_cell.font = Font(bold=True, color='102A43')
        if label in money_rows:
            fill = emphasis_fills[label]
            label_cell.fill = fill
            value_cell.fill = fill
            value_cell.font = Font(bold=True, color='102A43')


def build_admin_report_excel(report: AdminReport) -> tuple[str, bytes]:
    if report.report_type != PERIOD_REPORT_TYPE and not report.report_id:
        raise HTTPException(status_code=404, detail='Ревизия для выгрузки не найдена.')

    categories = [category for category in (report.categories or []) if category.name != DEFAULT_CATEGORY_NAME]
    discrepancy_rows: list[list[Any]] = []
    successful_rows: list[list[Any]] = []
    category_rows: list[list[Any]] = []
    subcategory_rows: list[list[Any]] = []
    employee_extra: dict[str, dict[str, set[str]]] = defaultdict(lambda: {
        'in_progress_subcategories': set(),
        'completed_subcategories': set(),
        'discrepancy_subcategories': set(),
    })

    for category in categories:
        problem_items = list(category.problem_items or [])
        completed_subcategories = list(category.completed_subcategories or [])
        in_progress_subcategories = list(category.in_progress_subcategories or [])
        selected_subcategories = list(category.selected_subcategories or [])
        remaining_subcategories = list(category.remaining_subcategories or [])
        discrepancy_subcategories = sorted({_safe_excel_text(item.subcategory_name, '—') for item in problem_items})

        category_cost = sum(_safe_excel_float(item.cost_total) for item in problem_items)
        category_retail = sum(_safe_excel_float(item.retail_total) for item in problem_items)
        category_lost_profit = sum(_safe_excel_float(item.lost_profit) for item in problem_items)

        category_rows.append([
            _safe_excel_text(category.name),
            _safe_excel_text(category.status),
            _safe_excel_text(category.assigned_to),
            'Категория целиком' if category.selected_on_cycle else 'Подкатегории',
            _excel_yes_no(bool(category.selected_on_cycle)),
            len(selected_subcategories),
            ', '.join(selected_subcategories) if selected_subcategories else '—',
            len(remaining_subcategories),
            ', '.join(remaining_subcategories) if remaining_subcategories else '—',
            len(completed_subcategories),
            len(discrepancy_subcategories),
            len(problem_items),
            category_cost,
            category_retail,
            category_lost_profit,
        ])

        subcategory_tracker: dict[str, dict[str, Any]] = {}

        def ensure_subcategory_row(subcategory_name: str) -> dict[str, Any]:
            key = _safe_excel_text(subcategory_name, '—')
            row = subcategory_tracker.get(key)
            if row is None:
                row = {
                    'category_name': _safe_excel_text(category.name),
                    'subcategory_name': key,
                    'selected_on_cycle': bool(category.selected_on_cycle) or key in selected_subcategories,
                    'assigned_to': set(),
                    'status': 'Не начата',
                    'completed_by': set(),
                    'discrepancy_items': 0,
                    'cost_total': 0.0,
                    'retail_total': 0.0,
                    'lost_profit': 0.0,
                }
                subcategory_tracker[key] = row
            return row

        for name in selected_subcategories:
            ensure_subcategory_row(name)
        for name in remaining_subcategories:
            sub_row = ensure_subcategory_row(name)
            sub_row['status'] = 'В работе'
        for sub in in_progress_subcategories:
            sub_row = ensure_subcategory_row(sub.name)
            sub_row['status'] = 'В работе'
            if sub.assigned_to:
                sub_row['assigned_to'].add(sub.assigned_to)
                employee_extra[sub.assigned_to]['in_progress_subcategories'].add(f"{category.name} → {sub.name}")
        for sub in completed_subcategories:
            sub_row = ensure_subcategory_row(sub.name)
            if sub_row['status'] != 'С расхождением':
                sub_row['status'] = 'Успешно'
            if sub.checked_by:
                sub_row['completed_by'].add(sub.checked_by)
                employee_extra[sub.checked_by]['completed_subcategories'].add(f"{category.name} → {sub.name}")
            successful_rows.append([
                _safe_excel_text(category.name),
                _safe_excel_text(sub.name),
                _safe_excel_text(sub.checked_by),
                _safe_excel_text(getattr(sub, 'status', 'green')),
                _safe_excel_text(report.date),
                _safe_excel_text(report.location),
            ])

        for item in problem_items:
            subcategory_name = _safe_excel_text(item.subcategory_name, '—')
            sub_row = ensure_subcategory_row(subcategory_name)
            sub_row['status'] = 'С расхождением'
            if item.checked_by:
                sub_row['completed_by'].add(item.checked_by)
                employee_extra[item.checked_by]['discrepancy_subcategories'].add(f"{category.name} → {subcategory_name}")
            sub_row['discrepancy_items'] += 1
            sub_row['cost_total'] += _safe_excel_float(item.cost_total)
            sub_row['retail_total'] += _safe_excel_float(item.retail_total)
            sub_row['lost_profit'] += _safe_excel_float(item.lost_profit)

            discrepancy_rows.append([
                _safe_excel_text(category.name),
                subcategory_name,
                _safe_excel_text(item.name),
                _safe_excel_text(item.checked_by),
                _safe_excel_text(report.date),
                _safe_excel_float(item.expected),
                _safe_excel_float(item.actual),
                _safe_excel_float(item.diff),
                _safe_excel_float(item.cost_total),
                _safe_excel_float(item.lost_profit),
                _safe_excel_float(item.retail_total),
                _safe_excel_float(item.cost_price),
                _safe_excel_float(item.retail_price),
                _safe_excel_text(report.location),
            ])

        for sub_row in subcategory_tracker.values():
            subcategory_rows.append([
                sub_row['category_name'],
                sub_row['subcategory_name'],
                _excel_yes_no(bool(sub_row['selected_on_cycle'])),
                ', '.join(sorted(sub_row['assigned_to'], key=str.lower)) if sub_row['assigned_to'] else '—',
                sub_row['status'],
                ', '.join(sorted(sub_row['completed_by'], key=str.lower)) if sub_row['completed_by'] else '—',
                sub_row['discrepancy_items'],
                round(float(sub_row['cost_total']), 2),
                round(float(sub_row['retail_total']), 2),
                round(float(sub_row['lost_profit']), 2),
            ])

    employees_sheet_rows: list[list[Any]] = []
    for employee in report.employees or []:
        extra = employee_extra.get(employee.full_name, {})
        in_progress = sorted(extra.get('in_progress_subcategories', set()), key=str.lower)
        completed = sorted(extra.get('completed_subcategories', set()), key=str.lower)
        discrepancy = sorted(extra.get('discrepancy_subcategories', set()), key=str.lower)
        employees_sheet_rows.append([
            _safe_excel_text(employee.full_name),
            _safe_excel_int(employee.user_id),
            len(employee.categories or []),
            ', '.join(employee.categories or []) if employee.categories else '—',
            _safe_excel_int(employee.completed_categories),
            len(completed),
            ', '.join(completed) if completed else '—',
            len(in_progress),
            ', '.join(in_progress) if in_progress else '—',
            len(discrepancy),
            ', '.join(discrepancy) if discrepancy else '—',
            _safe_excel_int(employee.discrepancy_items),
            _safe_excel_float(employee.total_cost),
            _safe_excel_float(employee.total_retail),
            _safe_excel_float(employee.total_lost_profit),
            _excel_yes_no(bool(employee.started_current_report)),
            _safe_excel_text(employee.started_at),
            _excel_yes_no(bool(employee.finished_current_report)),
            _safe_excel_text(employee.finished_at),
        ])

    workbook = Workbook()
    summary_ws = workbook.active
    _init_export_sheet(summary_ws, 'Сводка', ['Показатель', 'Значение'])

    counted_categories = categories
    completed_categories_count = sum(1 for category in counted_categories if _safe_excel_text(category.status, '').lower() in {'green', 'red'})
    discrepancy_categories_count = sum(1 for category in counted_categories if category.problem_items)
    no_discrepancy_categories_count = sum(1 for category in counted_categories if _safe_excel_text(category.status, '').lower() == 'green')
    discrepancy_items_count = sum(len(category.problem_items or []) for category in counted_categories)

    summary_rows = [
        ['Точка', _safe_excel_text(report.location)],
        ['Дата отчёта', _safe_excel_text(report.date)],
        ['Тип ревизии', 'Период' if report.report_type == PERIOD_REPORT_TYPE else ('Итоговая' if report.report_type == FINAL_REPORT_TYPE else 'Дневная')],
        ['Номер ревизии', '—' if report.report_type == PERIOD_REPORT_TYPE else _safe_excel_text(report.report_number or report.report_id)],
        ['Статус', _safe_excel_text(report.status)],
        ['Сотрудников в ревизии', len(report.employees or [])],
        ['Категорий всего', len(counted_categories)],
        ['Категорий завершено', completed_categories_count],
        ['Категорий с расхождениями', discrepancy_categories_count],
        ['Категорий без расхождений', no_discrepancy_categories_count],
        ['Проблемных товаров', discrepancy_items_count],
        ['Излишки', _safe_excel_float(report.total_plus)],
        ['Недостача', _safe_excel_float(report.total_minus)],
        ['Себестоимость', _safe_excel_float(report.total_cost)],
        ['Розница', _safe_excel_float(report.total_retail)],
        ['Утерянная прибыль', _safe_excel_float(report.total_lost_profit)],
        ['Сотрудники', ', '.join(employee.full_name for employee in (report.employees or [])) or '—'],
        ['Категории цикла', ', '.join(report.selected_categories or []) if report.selected_categories else '—'],
        ['Подкатегории цикла', ', '.join(report.selected_subcategories or []) if report.selected_subcategories else '—'],
    ]
    for row in summary_rows:
        summary_ws.append(row)

    summary_currency_labels = {'Излишки', 'Недостача', 'Себестоимость', 'Розница', 'Утерянная прибыль'}
    summary_count_labels = {'Сотрудников в ревизии', 'Категорий всего', 'Категорий завершено', 'Категорий с расхождениями', 'Категорий без расхождений', 'Проблемных товаров'}
    _finalize_export_sheet(
        summary_ws,
        left_align_columns={1, 2},
        width_overrides={'A': 34, 'B': 62},
        max_width=62,
        default_horizontal='left',
        default_vertical='top',
    )
    _style_summary_sheet(summary_ws)
    for row_index in range(2, summary_ws.max_row + 1):
        label = _safe_excel_text(summary_ws.cell(row=row_index, column=1).value, '')
        value_cell = summary_ws.cell(row=row_index, column=2)
        if not isinstance(value_cell.value, (int, float)):
            continue
        if label in summary_currency_labels:
            value_cell.number_format = '#,##0.00 "₽"'
        elif label in summary_count_labels:
            value_cell.number_format = '#,##0'

    discrepancies_ws = workbook.create_sheet()
    _init_export_sheet(discrepancies_ws, 'Расхождения', ['Категория', 'Подкатегория', 'Товар', 'Сотрудник', 'Дата', 'Учётное кол-во', 'Фактическое кол-во', 'Разница', 'Себестоимость', 'Утерянная прибыль', 'Розница', 'Себестоимость за шт.', 'Розница за шт.', 'Точка'])
    if discrepancy_rows:
        for row in discrepancy_rows:
            discrepancies_ws.append(row)
    else:
        discrepancies_ws.append(['Нет расхождений'])
        discrepancies_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=14)
    _finalize_export_sheet(discrepancies_ws, currency_columns={9, 10, 11, 12, 13}, quantity_columns={6, 7, 8}, width_overrides={'A': 24, 'B': 28, 'C': 38, 'D': 24, 'E': 20, 'F': 16, 'G': 16, 'H': 14, 'I': 18, 'J': 18, 'K': 18, 'L': 20, 'M': 18, 'N': 18}, max_width=38)
    discrepancy_negative_fill = PatternFill(fill_type='solid', fgColor='FCEBEC')
    discrepancy_positive_fill = PatternFill(fill_type='solid', fgColor='EAF7EE')
    for row_index in range(2, discrepancies_ws.max_row + 1):
        diff_cell = discrepancies_ws.cell(row=row_index, column=8)
        diff_value = diff_cell.value
        if isinstance(diff_value, (int, float)):
            if diff_value < 0:
                diff_cell.fill = discrepancy_negative_fill
            elif diff_value > 0:
                diff_cell.fill = discrepancy_positive_fill

    successful_ws = workbook.create_sheet()
    _init_export_sheet(successful_ws, 'Успешно пройдены', ['Категория', 'Подкатегория', 'Сотрудник', 'Статус', 'Дата', 'Точка'])
    if successful_rows:
        for row in successful_rows:
            successful_ws.append(row)
    else:
        successful_ws.append(['Нет успешно завершённых подкатегорий'])
        successful_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    _finalize_export_sheet(successful_ws, width_overrides={'A': 24, 'B': 28, 'C': 24, 'D': 16, 'E': 20, 'F': 18}, max_width=32)
    for row_index in range(2, successful_ws.max_row + 1):
        status_cell = successful_ws.cell(row=row_index, column=4)
        fill = _status_fill(status_cell.value)
        if fill:
            status_cell.fill = fill

    employees_ws = workbook.create_sheet()
    _init_export_sheet(employees_ws, 'Сотрудники', ['Сотрудник', 'ID', 'Категорий взял', 'Категории', 'Категорий завершил', 'Успешных подкатегорий', 'Какие успешные', 'Подкатегорий в работе', 'Какие в работе', 'Подкатегорий с расхождениями', 'Какие с расхождениями', 'Проблемных товаров', 'Себестоимость', 'Розница', 'Утерянная прибыль', 'Начинал ревизию', 'Время старта', 'Завершил ревизию', 'Время завершения'])
    if employees_sheet_rows:
        for row in employees_sheet_rows:
            employees_ws.append(row)
    else:
        employees_ws.append(['В ревизии нет сотрудников'])
        employees_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=19)
    _finalize_export_sheet(employees_ws, currency_columns={13, 14, 15}, quantity_columns={2, 3, 5, 6, 8, 10, 12}, width_overrides={'A': 14, 'B': 8, 'C': 14, 'D': 26, 'E': 16, 'F': 18, 'G': 42, 'H': 18, 'I': 42, 'J': 20, 'K': 42, 'L': 16, 'M': 18, 'N': 18, 'O': 18, 'P': 14, 'Q': 18, 'R': 16, 'S': 18}, max_width=42)

    categories_ws = workbook.create_sheet()
    _init_export_sheet(categories_ws, 'Категории', ['Категория', 'Статус', 'Сотрудник', 'Тип выбора', 'Взята целиком', 'Подкатегорий выбрано', 'Какие выбраны', 'Подкатегорий осталось', 'Какие остались', 'Подкатегорий завершено', 'Подкатегорий с расхождениями', 'Проблемных товаров', 'Себестоимость', 'Розница', 'Утерянная прибыль'])
    if category_rows:
        for row in category_rows:
            categories_ws.append(row)
    else:
        categories_ws.append(['Нет категорий для отображения'])
        categories_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=15)
    _finalize_export_sheet(categories_ws, currency_columns={13, 14, 15}, quantity_columns={6, 8, 10, 11, 12}, width_overrides={'A': 18, 'B': 12, 'C': 14, 'D': 16, 'E': 14, 'F': 18, 'G': 44, 'H': 18, 'I': 44, 'J': 18, 'K': 20, 'L': 18, 'M': 18, 'N': 18, 'O': 18}, max_width=44)
    for row_index in range(2, categories_ws.max_row + 1):
        status_cell = categories_ws.cell(row=row_index, column=2)
        fill = _status_fill(status_cell.value)
        if fill:
            status_cell.fill = fill

    subcategories_ws = workbook.create_sheet()
    _init_export_sheet(subcategories_ws, 'Подкатегории', ['Категория', 'Подкатегория', 'Выбрана на цикл', 'Сотрудник', 'Статус', 'Проверил', 'Проблемных товаров', 'Себестоимость', 'Розница', 'Утерянная прибыль'])
    if subcategory_rows:
        for row in subcategory_rows:
            subcategories_ws.append(row)
    else:
        subcategories_ws.append(['Нет подкатегорий для отображения'])
        subcategories_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
    _finalize_export_sheet(subcategories_ws, currency_columns={8, 9, 10}, quantity_columns={7}, width_overrides={'A': 24, 'B': 30, 'C': 18, 'D': 24, 'E': 18, 'F': 28, 'G': 18, 'H': 18, 'I': 18, 'J': 18}, max_width=32)
    for row_index in range(2, subcategories_ws.max_row + 1):
        status_cell = subcategories_ws.cell(row=row_index, column=5)
        fill = _status_fill(status_cell.value)
        if fill:
            status_cell.fill = fill

    output = BytesIO()
    workbook.save(output)
    filename = _build_export_filename(report)
    return filename, output.getvalue()


async def update_discrepancy_actual_qty(
    check_result_id: int,
    payload: UpdateDiscrepancyRequest,
    db: AsyncSession,
    current_user: User,
) -> UpdateDiscrepancyResponse:
    check_result = await db.get(CheckResult, check_result_id)
    if not check_result:
        raise HTTPException(status_code=404, detail='Строка расхождения не найдена.')
    if check_result.target_type != 'item':
        raise HTTPException(status_code=400, detail='Редактировать можно только товарные расхождения.')

    report = await db.get(Report, check_result.report_id)
    if not report:
        raise HTTPException(status_code=404, detail='Ревизия для строки расхождения не найдена.')
    if (report.report_type or DAILY_REPORT_TYPE) == FINAL_REPORT_TYPE:
        raise HTTPException(status_code=400, detail='Итоговую ревизию редактировать нельзя.')

    await ensure_user_can_access_location(current_user, report.location, db)

    password = str(payload.password or '').strip()
    if not password or not verify_password(password, current_user.password_hash):
        raise HTTPException(status_code=403, detail='Неверный пароль текущего управляющего.')

    actual_quantity = float(payload.actual_quantity)
    expected_quantity = float(check_result.expected_qty or 0.0)
    diff = float(actual_quantity - expected_quantity)
    if abs(diff) < 1e-9:
        diff = 0.0

    check_result.actual_qty = actual_quantity
    check_result.diff = diff
    check_result.status = 'green' if diff == 0.0 else 'red'

    await _refresh_report_status(report, db)
    await db.commit()

    return UpdateDiscrepancyResponse(
        success=True,
        message='Расхождение обновлено.',
        check_result_id=check_result.id,
        actual_quantity=float(actual_quantity),
        diff=float(diff),
        status=StatusEnum.GREEN if diff == 0.0 else StatusEnum.RED,
    )




async def update_discrepancy_cost_override(
    check_result_id: int,
    payload: UpdateDiscrepancyCostOverrideRequest,
    db: AsyncSession,
    current_user: User,
) -> UpdateDiscrepancyCostOverrideResponse:
    check_result = await db.get(CheckResult, check_result_id)
    if not check_result:
        raise HTTPException(status_code=404, detail='Строка расхождения не найдена.')
    if check_result.target_type != 'item':
        raise HTTPException(status_code=400, detail='Локальную себестоимость можно задать только для товарного расхождения.')

    report = await db.get(Report, check_result.report_id)
    if not report:
        raise HTTPException(status_code=404, detail='Ревизия для строки расхождения не найдена.')

    await ensure_user_can_access_location(current_user, report.location, db)

    point = await db.scalar(select(LocationPoint).where(LocationPoint.name == report.location).limit(1))
    if point is None:
        raise HTTPException(status_code=404, detail='Точка для расхождения не найдена.')

    item_id = str(check_result.target_id or '').strip()
    if not item_id:
        raise HTTPException(status_code=400, detail='У строки расхождения отсутствует идентификатор товара.')

    existing = await db.scalar(
        select(ProductCostOverride)
        .where(ProductCostOverride.location_point_id == point.id)
        .where(ProductCostOverride.item_id == item_id)
        .limit(1)
    )

    normalized_note = str(payload.note or '').strip() or None
    if payload.cost_price is None:
        if existing is not None:
            await db.delete(existing)
            await db.commit()
        return UpdateDiscrepancyCostOverrideResponse(
            success=True,
            message='Локальная себестоимость удалена. Теперь будет использоваться общий кеш точки.',
            check_result_id=check_result.id,
            item_id=item_id,
            cost_price=None,
            note=None,
            source='cache',
        )

    cost_price = round(float(payload.cost_price), 2)
    now = datetime.utcnow()
    if existing is None:
        db.add(ProductCostOverride(
            location_point_id=point.id,
            item_id=item_id,
            item_name=check_result.target_name or 'Товар',
            cost_price=cost_price,
            note=normalized_note,
            created_by_user_id=current_user.id,
            updated_by_user_id=current_user.id,
            created_at=now,
            updated_at=now,
        ))
    else:
        existing.item_name = check_result.target_name or existing.item_name
        existing.cost_price = cost_price
        existing.note = normalized_note
        existing.updated_by_user_id = current_user.id
        existing.updated_at = now

    await db.commit()
    return UpdateDiscrepancyCostOverrideResponse(
        success=True,
        message='Локальная себестоимость сохранена для этой точки.',
        check_result_id=check_result.id,
        item_id=item_id,
        cost_price=cost_price,
        note=normalized_note,
        source='override',
    )


async def delete_report(report_id: int, db: AsyncSession, current_user: User | None = None) -> DeleteResponse:
    report = await db.get(Report, report_id)
    if not report:
        raise HTTPException(status_code=404, detail='Ревизия не найдена.')
    if current_user is not None:
        await ensure_user_can_access_location(current_user, report.location, db)
    location = report.location
    await db.delete(report)
    await db.commit()
    _invalidate_runtime_inventory_cache(location)
    return DeleteResponse(success=True, message='Ревизия удалена.')
