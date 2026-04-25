from __future__ import annotations

import asyncio
import json
import logging
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta, timezone
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from pydantic import BaseModel, Field
from sqlalchemy import and_, delete, func, or_, select
from sqlalchemy.exc import DatabaseError, OperationalError
from sqlalchemy.ext.asyncio import AsyncSession
from fastapi import HTTPException

from app.config import settings
from app.models import (
    AdminLocationAccess,
    CategoryAssignment,
    CheckResult,
    ExpenseTemplate,
    LocationPoint,
    MonthlyExpenseEntry,
    PayrollAuditLog,
    PayrollCategoryRateVersion,
    PayrollDailyMetricCache,
    PayrollRecalcJob,
    PayrollSettingsVersion,
    Report,
    ReportTargetSnapshot,
    SelectionTarget,
    SelectionTargetDay,
    ShiftPayrollCategorySnapshot,
    ShiftPayrollSnapshot,
    User,
    WorkShift,
)
from app.database import AsyncSessionLocal
from app.moysklad import DEFAULT_CATEGORY_NAME, ms_client

logger = logging.getLogger(__name__)

MSK_TZ = timezone(timedelta(hours=3))
DEFAULT_EXIT_AMOUNT = 2000.0
PAYROLL_RETURNS_CATEGORY_NAME = 'Возвраты'
DEFAULT_BONUS_THRESHOLD = 40000.0
DEFAULT_BONUS_AMOUNT = 500.0
DEFAULT_OTHER_RATE_PERCENT = 3.0
CATEGORY_SALES_ALLOCATION_VERSION = 2
EXPENSE_MODE_SPREAD = 'spread'
EXPENSE_MODE_SINGLE_DAY = 'single_day'
DEFAULT_MANAGER_SALARY_BRACKETS = [
    {'threshold': 200000.0, 'rate_percent': 25.0},
    {'threshold': 125000.0, 'rate_percent': 20.0},
    {'threshold': 100000.0, 'rate_percent': 15.0},
    {'threshold': 50000.0, 'rate_percent': 10.0},
]

TOBACCO_KEYWORDS = ('сигарет', 'сигарилл', 'стик')

SALES_METRICS_TTL_SECONDS = 90
PAYROLL_DAILY_CACHE_RECENT_TTL_SECONDS = 900
PAYROLL_SHIFT_AUTO_CLOSE_TIME = time(hour=3, minute=55)
_sales_metrics_cache: dict[tuple[int, str, str], tuple[float, dict[date, dict[str, Any]]]] = {}
_category_lookup_cache: dict[str, tuple[float, dict[str, dict[str, str]]]] = {}
_payroll_recalc_tasks: dict[int, asyncio.Task[Any]] = {}


def _is_locked_database_error(exc: Exception) -> bool:
    message = str(exc).lower()
    return 'database is locked' in message or 'database table is locked' in message


async def _sleep_for_locked_retry(attempt: int) -> None:
    await asyncio.sleep(min(0.25 * (attempt + 1), 1.5))


def _sanitize_uncategorized_net(category_name: Any, net_sales_amount: float) -> float:
    return round(float(net_sales_amount or 0.0), 2)


def _is_uncategorized_category(category_name: Any = None, category_id: Any = None) -> bool:
    normalized_name = str(category_name or '').strip()
    normalized_id = str(category_id or '').strip()
    return normalized_name == DEFAULT_CATEGORY_NAME or normalized_id == '__other__'


def _is_uncategorized_return_adjustment_values(sales_amount: Any, return_amount: Any, net_sales_amount: Any) -> bool:
    sales = round(float(sales_amount or 0.0), 2)
    returns = round(float(return_amount or 0.0), 2)
    net = round(float(net_sales_amount or 0.0), 2)
    return returns > 0.009 or sales < -0.009 or net < -0.009


def _is_uncategorized_return_adjustment_row(row: Any) -> bool:
    return _is_uncategorized_category(
        _row_value(row, 'category_name'),
        _row_value(row, 'category_id'),
    ) and _is_uncategorized_return_adjustment_values(
        _row_value(row, 'sales_amount'),
        _row_value(row, 'return_amount'),
        _row_value(row, 'net_sales_amount'),
    )


def _get_payroll_display_category_name(row: Any) -> str:
    category_name = str(_row_value(row, 'category_name') or '').strip()
    if _is_uncategorized_return_adjustment_row(row):
        return PAYROLL_RETURNS_CATEGORY_NAME
    return category_name or DEFAULT_CATEGORY_NAME


def _sanitize_uncategorized_row(row: dict[str, Any]) -> dict[str, Any]:
    category_name = row.get('category_name')
    net_sales_amount = _sanitize_uncategorized_net(category_name, row.get('net_sales_amount') or 0.0)
    row['net_sales_amount'] = net_sales_amount
    if _is_uncategorized_category(category_name, row.get('category_id')):
        row['earning_amount'] = 0.0
        row['rate_percent'] = 0.0
    return row


def _resolve_calculation_sales_base(sales_amount: Any, net_sales_amount: Any = None) -> float:
    net = round(float(net_sales_amount or 0.0), 2)
    if net_sales_amount is not None:
        return net
    sales = round(float(sales_amount or 0.0), 2)
    return sales


def _resolve_category_calculation_base(row: dict[str, Any]) -> float:
    return _resolve_calculation_sales_base(row.get('sales_amount'), row.get('net_sales_amount'))


def _calculate_category_earning_amount(base_amount: Any, rate_percent: Any) -> float:
    return round(float(base_amount or 0.0) * (float(rate_percent or 0.0) / 100.0), 2)


def _calculate_gross_profit_from_revenue(revenue_amount: Any, cost_amount: Any) -> float:
    return round(float(revenue_amount or 0.0) - float(cost_amount or 0.0), 2)


def _recalculate_summary_totals_from_categories(day_payload: dict[str, Any]) -> None:
    categories = day_payload.get('categories') or []
    if not categories:
        return
    day_payload['category_earnings_total'] = round(sum(float(item.get('earning_amount') or 0.0) for item in categories), 2)
    bonus_base_sales_amount = day_payload.get('bonus_base_sales_amount')
    if bonus_base_sales_amount is None:
        return
    if float(day_payload.get('bonus_amount') or 0.0) > 0 and float(day_payload.get('bonus_threshold') or 0.0) > 0:
        day_payload['bonus_base_sales_amount'] = round(max(float(bonus_base_sales_amount or 0.0), 0.0), 2)
    day_payload['gross_salary_amount'] = round(
        float(day_payload.get('exit_amount') or 0.0)
        + float(day_payload.get('bonus_amount') or 0.0)
        + float(day_payload.get('category_earnings_total') or 0.0),
        2,
    )


def _should_backfill_category_costs(categories: list[dict[str, Any]] | None, total_cost_amount: float) -> bool:
    rows = categories or []
    if not rows:
        return False
    visible_rows = [row for row in rows if str(row.get('category_name') or '').strip() != DEFAULT_CATEGORY_NAME]
    has_visible_sales = any(
        abs(float(row.get('sales_amount') or 0.0)) > 0.009
        or abs(float(row.get('net_sales_amount') or 0.0)) > 0.009
        for row in visible_rows
    )
    if not has_visible_sales:
        return False

    total_category_cost_sum = round(sum(float(row.get('cost_amount') or 0.0) for row in rows), 2)
    if abs(total_category_cost_sum) <= 0.009:
        return True

    visible_cost_sum = round(sum(float(row.get('cost_amount') or 0.0) for row in visible_rows), 2)
    if abs(visible_cost_sum) > 0.009:
        return False

    normalized_total_cost = round(float(total_cost_amount or 0.0), 2)
    if abs(normalized_total_cost) <= 0.009:
        return True

    other_cost_sum = round(sum(float(row.get('cost_amount') or 0.0) for row in rows if str(row.get('category_name') or '').strip() == DEFAULT_CATEGORY_NAME), 2)
    return abs(other_cost_sum - normalized_total_cost) <= 0.01


async def _backfill_category_costs_from_day_metrics(
    point: LocationPoint,
    shift_date: date,
    categories: list[dict[str, Any]],
    total_cost_amount: float,
    share_ratio: float,
    db: AsyncSession,
) -> list[dict[str, Any]]:
    if not _should_backfill_category_costs(categories, total_cost_amount):
        return categories

    day_metrics = (await _load_point_sales_metrics(point, shift_date, shift_date, db)).get(shift_date, _empty_day_metrics())
    source_rows = day_metrics.get('categories') or []
    by_id = {str(row.get('category_id') or '').strip(): row for row in source_rows if str(row.get('category_id') or '').strip()}
    by_name = {_normalize_category_name_key(row.get('category_name')): row for row in source_rows if _normalize_category_name_key(row.get('category_name'))}

    patched: list[dict[str, Any]] = []
    for row in categories:
        cloned = dict(row)
        key = str(cloned.get('category_id') or '').strip()
        matched = by_id.get(key)
        if matched is None:
            matched = by_name.get(_normalize_category_name_key(cloned.get('category_name')))
        if matched is not None:
            cloned['cost_amount'] = round(float(matched.get('cost_amount') or 0.0) * float(share_ratio or 0.0), 2)
        patched.append(_sanitize_uncategorized_row(cloned))

    visible_cost_sum = round(sum(float(row.get('cost_amount') or 0.0) for row in patched if str(row.get('category_name') or '').strip() != DEFAULT_CATEGORY_NAME), 2)
    source_total_cost_amount = round(float(day_metrics.get('cost_amount') or 0.0) * float(share_ratio or 0.0), 2)
    target_total_cost_amount = round(float(total_cost_amount or 0.0), 2)
    if abs(target_total_cost_amount) <= 0.009 and abs(source_total_cost_amount) > 0.009:
        target_total_cost_amount = source_total_cost_amount
    uncategorized_row = next((row for row in patched if str(row.get('category_name') or '').strip() == DEFAULT_CATEGORY_NAME), None)
    if uncategorized_row is not None:
        uncategorized_row['cost_amount'] = round(target_total_cost_amount - visible_cost_sum, 2)
    return patched


def _ms_client_enabled(token: str | None = None, location: str | None = None) -> bool:
    enabled_attr = getattr(ms_client, "enabled", None)
    if callable(enabled_attr):
        try:
            return bool(enabled_attr(token, location=location))
        except TypeError:
            try:
                return bool(enabled_attr(token))
            except TypeError:
                return bool(enabled_attr())
    return bool(enabled_attr)



def bootstrap_payroll_schema(connection) -> None:
    tables = {
        str(row[0])
        for row in connection.exec_driver_sql("SELECT name FROM sqlite_master WHERE type='table'").fetchall()
    }
    columns = {
        str(row[1])
        for row in connection.exec_driver_sql("PRAGMA table_info(payroll_settings_versions)").fetchall()
    }
    if 'bonus_category_ids_json' not in columns:
        connection.exec_driver_sql(
            "ALTER TABLE payroll_settings_versions ADD COLUMN bonus_category_ids_json TEXT NOT NULL DEFAULT '[]'"
        )
    if 'manager_salary_brackets_json' not in columns:
        default_json = json.dumps(DEFAULT_MANAGER_SALARY_BRACKETS, ensure_ascii=False).replace("'", "''")
        connection.exec_driver_sql(
            "ALTER TABLE payroll_settings_versions "
            f"ADD COLUMN manager_salary_brackets_json TEXT NOT NULL DEFAULT '{default_json}'"
        )

    if 'monthly_expense_entries' in tables:
        expense_columns = {
            str(row[1])
            for row in connection.exec_driver_sql("PRAGMA table_info(monthly_expense_entries)").fetchall()
        }
        if 'expense_date' not in expense_columns:
            connection.exec_driver_sql("ALTER TABLE monthly_expense_entries ADD COLUMN expense_date DATE")
            connection.exec_driver_sql(
                "UPDATE monthly_expense_entries SET expense_date = month_start WHERE expense_date IS NULL"
            )
        if 'distribution_mode' not in expense_columns:
            connection.exec_driver_sql(
                f"ALTER TABLE monthly_expense_entries ADD COLUMN distribution_mode VARCHAR(20) NOT NULL DEFAULT '{EXPENSE_MODE_SPREAD}'"
            )
            connection.exec_driver_sql(
                f"UPDATE monthly_expense_entries SET distribution_mode = '{EXPENSE_MODE_SINGLE_DAY}' WHERE template_id IS NULL"
            )

    if 'shift_payroll_category_snapshots' in tables:
        shift_category_columns = {
            str(row[1])
            for row in connection.exec_driver_sql("PRAGMA table_info(shift_payroll_category_snapshots)").fetchall()
        }
        if 'cost_amount' not in shift_category_columns:
            connection.exec_driver_sql(
                "ALTER TABLE shift_payroll_category_snapshots ADD COLUMN cost_amount FLOAT NOT NULL DEFAULT 0"
            )


class PayrollSettingsUpdateRequest(BaseModel):
    location: str
    effective_from: date
    exit_amount: float = Field(default=DEFAULT_EXIT_AMOUNT, ge=0)
    bonus_threshold: float = Field(default=DEFAULT_BONUS_THRESHOLD, ge=0)
    bonus_amount: float = Field(default=DEFAULT_BONUS_AMOUNT, ge=0)
    other_rate_percent: float = Field(default=DEFAULT_OTHER_RATE_PERCENT, ge=0)
    responsible_admin_user_id: int | None = Field(default=None, ge=1)
    bonus_category_ids: list[str] = Field(default_factory=list)
    manager_salary_brackets: list[dict[str, Any]] = Field(default_factory=list)
    category_rates: list[dict[str, Any]] = Field(default_factory=list)


class WorkShiftUpsertRequest(BaseModel):
    location: str
    shift_date: date
    employee_user_id: int = Field(..., ge=1)


class WorkShiftDeleteRequest(BaseModel):
    hard: bool = False


class ExpenseTemplateCreateRequest(BaseModel):
    location: str
    name: str = Field(..., min_length=1, max_length=255)
    amount_type: str = Field(default='dynamic')
    default_amount: float | None = Field(default=None, ge=0)
    assign_to_employee_by_default: bool = False


class ExpenseTemplateUpdateRequest(BaseModel):
    name: str = Field(..., min_length=1, max_length=255)
    amount_type: str = Field(default='dynamic')
    default_amount: float | None = Field(default=None, ge=0)
    assign_to_employee_by_default: bool = False
    is_active: bool = True


class MonthlyExpenseEntryUpdateRequest(BaseModel):
    amount: float = Field(..., ge=0)
    is_paid: bool = False
    assigned_employee_user_id: int | None = Field(default=None, ge=1)
    apply_to_employee_salary: bool = False
    distribution_mode: str = Field(default=EXPENSE_MODE_SPREAD)
    expense_date: date | None = None
    comment: str | None = Field(default=None, max_length=2000)


class ManualMonthlyExpenseCreateRequest(BaseModel):
    location: str
    month_start: date
    name: str = Field(..., min_length=1, max_length=255)
    amount: float = Field(..., ge=0)
    is_paid: bool = False
    assigned_employee_user_id: int | None = Field(default=None, ge=1)
    apply_to_employee_salary: bool = False
    distribution_mode: str = Field(default=EXPENSE_MODE_SINGLE_DAY)
    expense_date: date | None = None
    comment: str | None = Field(default=None, max_length=2000)


@dataclass(slots=True)
class CategoryDocMetrics:
    sales: float = 0.0
    returns: float = 0.0
    sales_cost: float = 0.0
    return_cost: float = 0.0


@dataclass(slots=True)
class ProfitReportDayMetrics:
    sales_amount: float = 0.0
    return_amount: float = 0.0
    cost_amount: float = 0.0
    gross_profit_amount: float = 0.0
    has_rows: bool = False
    category_costs_by_id: dict[str, float] = field(default_factory=dict)
    category_costs_by_name: dict[str, float] = field(default_factory=dict)


@dataclass(slots=True)
class ShiftComputedPayroll:
    shift: WorkShift
    location_point: LocationPoint
    employee: User
    settings: PayrollSettingsVersion
    split_count: int
    share_ratio: float
    categories: list[dict[str, Any]]
    exit_amount: float
    bonus_threshold: float
    bonus_amount: float
    bonus_base_sales_amount: float
    bonus_category_ids: list[str]
    other_rate_percent: float
    gross_sales_amount: float
    return_amount: float
    net_sales_amount: float
    cost_amount: float
    gross_profit_amount: float
    non_tobacco_net_sales_for_bonus: float
    category_earnings_total: float
    gross_salary_amount: float
    snapshot_id: int | None = None
    is_closed: bool = False
    closed_at: str | None = None
    is_auto_closed: bool = False



def _now_msk() -> datetime:
    return datetime.now(MSK_TZ)


def _normalize_msk_datetime(value: datetime | None = None) -> datetime:
    current = value or _now_msk()
    if current.tzinfo is None:
        return current.replace(tzinfo=MSK_TZ)
    return current.astimezone(MSK_TZ)


def get_moscow_today() -> date:
    return _now_msk().date()


def get_payroll_operational_today(value: datetime | None = None) -> date:
    current = _normalize_msk_datetime(value)
    if current.timetz().replace(tzinfo=None) < PAYROLL_SHIFT_AUTO_CLOSE_TIME:
        return current.date() - timedelta(days=1)
    return current.date()





def _serialize_recalc_job(job: PayrollRecalcJob, point: LocationPoint | None = None) -> dict[str, Any]:
    result: dict[str, Any]
    try:
        result = json.loads(job.result_json or '{}')
        if not isinstance(result, dict):
            result = {}
    except Exception:
        result = {}
    return {
        'job_id': job.id,
        'location': point.name if point else None,
        'settings_version_id': job.settings_version_id,
        'date_from': job.date_from.isoformat(),
        'date_to': job.date_to.isoformat(),
        'status': job.status,
        'progress_current': int(job.progress_current or 0),
        'progress_total': int(job.progress_total or 0),
        'message': job.message,
        'error_text': job.error_text,
        'result': result,
        'created_at': _datetime_to_str(job.created_at),
        'started_at': _datetime_to_str(job.started_at),
        'finished_at': _datetime_to_str(job.finished_at),
    }


async def _update_recalc_job_progress(job_id: int, *, status: str | None = None, current: int | None = None, total: int | None = None, message: str | None = None, error_text: str | None = None, result: dict[str, Any] | None = None) -> None:
    for attempt in range(8):
        try:
            async with AsyncSessionLocal() as db:
                job = await db.get(PayrollRecalcJob, job_id)
                if job is None:
                    return
                now = datetime.utcnow()
                if status is not None:
                    job.status = status
                    if status == 'running' and job.started_at is None:
                        job.started_at = now
                    if status in {'done', 'failed', 'cancelled'}:
                        job.finished_at = now
                if current is not None:
                    job.progress_current = max(int(current), 0)
                if total is not None:
                    job.progress_total = max(int(total), 0)
                if message is not None:
                    job.message = message
                if error_text is not None:
                    job.error_text = error_text
                if result is not None:
                    job.result_json = json.dumps(result, ensure_ascii=False)
                job.updated_at = now
                await db.commit()
                return
        except (OperationalError, DatabaseError) as exc:
            if _is_locked_database_error(exc) and attempt < 7:
                logger.warning('SQLite занят при обновлении прогресса payroll_recalc_jobs. job_id=%s attempt=%s', job_id, attempt + 1)
                await _sleep_for_locked_retry(attempt)
                continue
            raise


async def _run_payroll_recalc_job(job_id: int) -> None:
    try:
        await _update_recalc_job_progress(job_id, status='running', current=0, message='Запущен пересчёт закрытых смен...')
        async with AsyncSessionLocal() as db:
            job = await db.get(PayrollRecalcJob, job_id)
            if job is None:
                return
            point = await db.get(LocationPoint, job.location_point_id)
            if point is None:
                await _update_recalc_job_progress(job_id, status='failed', error_text='Точка для пересчёта не найдена.')
                return

            async def progress(current: int, total: int) -> None:
                await _update_recalc_job_progress(job_id, current=current, total=total, message=f'Пересчитываем смены: {current} из {total}.')

            result = await rebuild_closed_shift_snapshots(
                job.date_from,
                job.date_to,
                db,
                location=point.name,
                force_refresh_metrics=True,
                progress_callback=progress,
            )
            await _update_recalc_job_progress(
                job_id,
                status='done',
                current=int(result.get('processed') or 0),
                total=int(result.get('processed') or 0),
                message='Пересчёт завершён.',
                result=result,
            )
    except Exception as exc:
        logger.exception('Фоновый пересчёт зарплаты завершился ошибкой. job_id=%s', job_id)
        await _update_recalc_job_progress(job_id, status='failed', message='Ошибка пересчёта.', error_text=str(exc))
    finally:
        _payroll_recalc_tasks.pop(job_id, None)


async def _enqueue_payroll_recalc_job(point: LocationPoint, settings_version_id: int | None, date_from: date, date_to: date, requested_by_user_id: int | None) -> PayrollRecalcJob:
    async with AsyncSessionLocal() as db:
        active_job = await db.scalar(
            select(PayrollRecalcJob)
            .where(
                PayrollRecalcJob.location_point_id == point.id,
                PayrollRecalcJob.date_from == date_from,
                PayrollRecalcJob.date_to == date_to,
                PayrollRecalcJob.status.in_(['queued', 'running']),
            )
            .order_by(PayrollRecalcJob.id.desc())
            .limit(1)
        )
        if active_job is None:
            active_job = PayrollRecalcJob(
                location_point_id=point.id,
                settings_version_id=settings_version_id,
                requested_by_user_id=requested_by_user_id,
                date_from=date_from,
                date_to=date_to,
                status='queued',
                progress_current=0,
                progress_total=0,
                message='Задача поставлена в очередь на пересчёт.',
                result_json='{}',
                updated_at=datetime.utcnow(),
            )
            db.add(active_job)
            await db.commit()
            await db.refresh(active_job)
        elif settings_version_id and active_job.settings_version_id != settings_version_id:
            active_job.settings_version_id = settings_version_id
            active_job.updated_at = datetime.utcnow()
            await db.commit()
            await db.refresh(active_job)
    if active_job.id not in _payroll_recalc_tasks or _payroll_recalc_tasks[active_job.id].done():
        _payroll_recalc_tasks[active_job.id] = asyncio.create_task(_run_payroll_recalc_job(active_job.id))
    return active_job


async def resume_pending_payroll_recalc_jobs() -> None:
    async with AsyncSessionLocal() as db:
        pending_jobs = (
            await db.scalars(
                select(PayrollRecalcJob)
                .where(PayrollRecalcJob.status.in_(['queued', 'running']))
                .order_by(PayrollRecalcJob.id.asc())
            )
        ).all()
    for job in pending_jobs:
        if job.id in _payroll_recalc_tasks and not _payroll_recalc_tasks[job.id].done():
            continue
        _payroll_recalc_tasks[job.id] = asyncio.create_task(_run_payroll_recalc_job(job.id))


async def get_payroll_recalc_status(location: str, db: AsyncSession, current_user: User, job_id: int | None = None) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    query = select(PayrollRecalcJob).where(PayrollRecalcJob.location_point_id == point.id)
    if job_id is not None:
        query = query.where(PayrollRecalcJob.id == job_id)
    query = query.order_by(PayrollRecalcJob.id.desc()).limit(1)
    job = await db.scalar(query)
    if job is None:
        return {'location': point.name, 'job': None}
    return {'location': point.name, 'job': _serialize_recalc_job(job, point)}

def _normalize_location(location: str) -> str:
    return location.strip().title()



def _month_start(value: date) -> date:
    return date(value.year, value.month, 1)



def _month_end(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1) - timedelta(days=1)
    return date(value.year, value.month + 1, 1) - timedelta(days=1)



def _normalize_expense_distribution_mode(value: Any, *, default: str = EXPENSE_MODE_SPREAD) -> str:
    normalized_default = EXPENSE_MODE_SINGLE_DAY if default == EXPENSE_MODE_SINGLE_DAY else EXPENSE_MODE_SPREAD
    normalized = str(value or '').strip().lower()
    if normalized in {EXPENSE_MODE_SINGLE_DAY, 'single-day', 'one_day', 'oneday', 'day'}:
        return EXPENSE_MODE_SINGLE_DAY
    if normalized in {EXPENSE_MODE_SPREAD, 'month', 'monthly', 'spread_month'}:
        return EXPENSE_MODE_SPREAD
    return normalized_default

def _clip_expense_date_to_month(expense_date: date, month_start: date) -> date:
    month_key = _month_start(month_start)
    month_last = _month_end(month_key)
    if expense_date < month_key:
        return month_key
    if expense_date > month_last:
        return month_last
    return expense_date

def _default_expense_date_for_month(month_start: date, *, day_of_month: int = 1) -> date:
    month_key = _month_start(month_start)
    month_last = _month_end(month_key)
    safe_day = min(max(int(day_of_month or 1), 1), month_last.day)
    return date(month_key.year, month_key.month, safe_day)

def _resolve_entry_expense_date(entry: MonthlyExpenseEntry, template: ExpenseTemplate | None = None) -> date:
    month_key = _month_start(entry.month_start)
    existing_date = getattr(entry, 'expense_date', None)
    if isinstance(existing_date, date):
        return _clip_expense_date_to_month(existing_date, month_key)
    template_day = int(template.day_of_month or 1) if template is not None else 1
    return _default_expense_date_for_month(month_key, day_of_month=template_day)

def _expense_entry_mode(entry: MonthlyExpenseEntry) -> str:
    default_mode = EXPENSE_MODE_SINGLE_DAY if entry.template_id is None else EXPENSE_MODE_SPREAD
    return _normalize_expense_distribution_mode(getattr(entry, 'distribution_mode', None), default=default_mode)

def _expense_entry_amount_for_period(
    entry: MonthlyExpenseEntry,
    *,
    date_from: date,
    date_to: date,
    template: ExpenseTemplate | None = None,
) -> float:
    if not entry.is_paid:
        return 0.0

    amount = round(float(entry.amount or 0.0), 2)
    if amount <= 0:
        return 0.0

    month_key = _month_start(entry.month_start)
    month_last = _month_end(month_key)
    overlap_from = max(date_from, month_key)
    overlap_to = min(date_to, month_last)
    if overlap_from > overlap_to:
        return 0.0

    mode = _expense_entry_mode(entry)
    if mode == EXPENSE_MODE_SINGLE_DAY:
        expense_day = _resolve_entry_expense_date(entry, template)
        return amount if overlap_from <= expense_day <= overlap_to else 0.0

    days_in_month = (month_last - month_key).days + 1
    overlap_days = (overlap_to - overlap_from).days + 1
    if days_in_month <= 0 or overlap_days <= 0:
        return 0.0
    return round(amount * (overlap_days / days_in_month), 2)



def _daterange(start: date, end: date) -> list[date]:
    days = (end - start).days
    return [start + timedelta(days=offset) for offset in range(days + 1)]



def _extract_id(meta_or_obj: dict[str, Any] | None) -> str | None:
    if not meta_or_obj:
        return None
    if isinstance(meta_or_obj.get('meta'), dict):
        meta_or_obj = meta_or_obj['meta']
    obj_id = meta_or_obj.get('id')
    if obj_id:
        return str(obj_id)
    href = meta_or_obj.get('href')
    if not href:
        return None
    return href.rstrip('/').split('/')[-1]



def _money(value: Any) -> float:
    if value is None:
        return 0.0
    try:
        return round(float(value) / 100.0, 2)
    except (TypeError, ValueError):
        return 0.0



def _quantity(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0



def _money_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return round(float(value) / 100.0, 2)
    except (TypeError, ValueError):
        return None


def _number_or_none(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None



def _first_money_value(container: dict[str, Any], fields: tuple[str, ...]) -> float | None:
    for field in fields:
        if field in container and container.get(field) is not None:
            amount = _money_or_none(container.get(field))
            if amount is not None:
                return amount
    return None



def _extract_cost_from_stock_payload(payload: Any, quantity: float) -> float | None:
    total_fields = (
        'costSum',
        'costPriceSum',
        'buyPriceSum',
        'purchaseCostSum',
        'costAmount',
        'buySum',
        'purchaseSum',
        'cost',
        'stockCost',
        'stockCostSum',
    )
    unit_fields = (
        'buyPrice',
        'costPrice',
        'purchasePrice',
        'purchaseCost',
        'costValue',
        'price',
    )

    if isinstance(payload, list):
        total = 0.0
        found = False
        for item in payload:
            amount = _extract_cost_from_stock_payload(item, quantity)
            if amount is None:
                continue
            total += amount
            found = True
        return round(max(total, 0.0), 2) if found else None

    if not isinstance(payload, dict):
        return None

    total_cost = _first_money_value(payload, total_fields)
    if total_cost is not None:
        return round(max(total_cost, 0.0), 2)

    unit_cost = _first_money_value(payload, unit_fields)
    if unit_cost is not None:
        nested_quantity = _quantity(
            payload.get('quantity')
            or payload.get('qty')
            or payload.get('stock')
            or payload.get('available')
        )
        effective_quantity = nested_quantity if nested_quantity > 0 else quantity
        if effective_quantity > 0:
            return round(max(unit_cost, 0.0) * effective_quantity, 2)
        return round(max(unit_cost, 0.0), 2)

    nested_total = 0.0
    found_nested = False
    for key in ('rows', 'items', 'stocks', 'stores', 'stockByOperations', 'stockByStore', 'data'):
        amount = _extract_cost_from_stock_payload(payload.get(key), quantity)
        if amount is None:
            continue
        nested_total += amount
        found_nested = True
    if found_nested:
        return round(max(nested_total, 0.0), 2)

    for nested_value in payload.values():
        if not isinstance(nested_value, (dict, list)):
            continue
        amount = _extract_cost_from_stock_payload(nested_value, quantity)
        if amount is not None:
            return amount

    return None



def _extract_position_amount(position: dict[str, Any]) -> float:
    quantity = _quantity(position.get('quantity'))
    amount = _first_money_value(position, ('sum', 'amount', 'saleSum', 'retailSum'))
    if amount is not None:
        return round(amount, 2)

    unit_price = _first_money_value(position, ('price', 'salePrice', 'sellingPrice'))
    if unit_price is not None and quantity > 0:
        gross_amount = round(unit_price * quantity, 2)
        discount_sum = _first_money_value(position, ('discountSum', 'discountsum', 'discountAmount', 'discountamount'))
        if discount_sum is not None:
            return round(max(gross_amount - discount_sum, 0.0), 2)

        discount_percent = _number_or_none(position.get('discount'))
        if discount_percent is not None:
            normalized_discount = min(max(discount_percent, 0.0), 100.0)
            return round(max(gross_amount * (1.0 - normalized_discount / 100.0), 0.0), 2)

        return gross_amount
    return 0.0


def _row_value(row: Any, field: str, default: Any = None) -> Any:
    if isinstance(row, dict):
        return row.get(field, default)
    return getattr(row, field, default)


def _is_uncategorized_row(row: Any) -> bool:
    return _is_uncategorized_category(
        _row_value(row, 'category_name'),
        _row_value(row, 'category_id'),
    )


def _has_legacy_uncategorized_adjustment(rows: list[Any] | None) -> bool:
    categories = list(rows or [])
    if not categories:
        return False

    visible_rows = [row for row in categories if not _is_uncategorized_row(row)]
    has_visible_turnover = any(
        abs(float(_row_value(row, 'sales_amount') or 0.0)) > 0.009
        or abs(float(_row_value(row, 'return_amount') or 0.0)) > 0.009
        or abs(float(_row_value(row, 'net_sales_amount') or 0.0)) > 0.009
        for row in visible_rows
    )
    if not has_visible_turnover:
        return False

    for row in categories:
        if not _is_uncategorized_row(row):
            continue
        adjustment_version = int(_row_value(row, '_category_sales_allocation_version') or 0)
        if adjustment_version >= CATEGORY_SALES_ALLOCATION_VERSION:
            continue
        sales_amount = round(float(_row_value(row, 'sales_amount') or 0.0), 2)
        return_amount = round(float(_row_value(row, 'return_amount') or 0.0), 2)
        if sales_amount < -0.009 or return_amount > 0.009:
            return True
    return False


def _mark_category_sales_allocation_version(categories: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    marked: list[dict[str, Any]] = []
    for row in categories or []:
        cloned = dict(row)
        cloned['_category_sales_allocation_version'] = CATEGORY_SALES_ALLOCATION_VERSION
        marked.append(cloned)
    return marked



def _extract_position_cost_amount(position: dict[str, Any]) -> float | None:
    quantity = _quantity(position.get('quantity'))
    total_cost = _first_money_value(position, ('costSum', 'buySum', 'buyPriceSum', 'purchaseCostSum', 'costAmount', 'purchaseSum', 'cost'))
    if total_cost is not None:
        return round(max(total_cost, 0.0), 2)
    unit_cost = _first_money_value(position, ('buyPrice', 'costPrice', 'purchasePrice', 'purchaseCost', 'costValue'))
    if unit_cost is not None and quantity > 0:
        return round(max(unit_cost, 0.0) * quantity, 2)
    stock_cost = _extract_cost_from_stock_payload(position.get('stock'), quantity)
    if stock_cost is not None:
        return stock_cost
    return None



def _extract_position_unit_retail_price(position: dict[str, Any]) -> float | None:
    quantity = _quantity(position.get('quantity'))
    amount = _first_money_value(position, ('sum', 'amount', 'saleSum', 'retailSum'))
    if amount is not None and quantity > 0:
        return round(amount / quantity, 2)
    return _first_money_value(position, ('price', 'salePrice', 'sellingPrice'))



def _extract_shift_cost_amount(doc: dict[str, Any]) -> float | None:
    return _first_money_value(doc, ('costSum', 'costPriceSum', 'purchaseCostSum', 'buySum', 'buyPriceSum', 'costAmount', 'purchaseSum', 'cost'))



def _extract_shift_profit_amount(doc: dict[str, Any]) -> float | None:
    return _first_money_value(doc, ('profit', 'profitSum', 'grossProfit', 'grossProfitSum', 'margin', 'marginSum', 'pnl', 'pnlSum'))



def _datetime_to_str(value: datetime | None) -> str | None:
    if value is None:
        return None
    if value.tzinfo is None:
        value = value.replace(tzinfo=timezone.utc).astimezone(MSK_TZ)
    else:
        value = value.astimezone(MSK_TZ)
    return value.strftime('%Y-%m-%d %H:%M')



def _is_tobacco_category(name: str | None) -> bool:
    normalized = (name or '').strip().lower()
    return any(keyword in normalized for keyword in TOBACCO_KEYWORDS)



def _default_manager_salary_brackets() -> list[dict[str, float]]:
    return [dict(item) for item in DEFAULT_MANAGER_SALARY_BRACKETS]



def _normalize_manager_salary_brackets(raw_rows: Any) -> list[dict[str, float]]:
    normalized: list[dict[str, float]] = []
    seen_thresholds: set[float] = set()
    for row in raw_rows or []:
        try:
            threshold = round(max(float((row or {}).get('threshold') or 0), 0.0), 2)
            rate_percent = round(max(float((row or {}).get('rate_percent') or (row or {}).get('rate') or 0), 0.0), 2)
        except (TypeError, ValueError, AttributeError):
            continue
        if threshold in seen_thresholds:
            continue
        seen_thresholds.add(threshold)
        normalized.append({'threshold': threshold, 'rate_percent': rate_percent})
    normalized.sort(key=lambda item: item['threshold'], reverse=True)
    return normalized or _default_manager_salary_brackets()



def _load_manager_salary_brackets(settings_version: PayrollSettingsVersion | None) -> list[dict[str, float]]:
    raw_value = getattr(settings_version, 'manager_salary_brackets_json', None) if settings_version else None
    try:
        parsed = json.loads(raw_value or '[]')
    except (TypeError, ValueError, json.JSONDecodeError):
        parsed = []
    return _normalize_manager_salary_brackets(parsed)



def _load_bonus_category_ids(settings_version: PayrollSettingsVersion | None) -> list[str]:
    raw_value = getattr(settings_version, 'bonus_category_ids_json', None) if settings_version else None
    try:
        parsed = json.loads(raw_value or '[]')
    except (TypeError, ValueError, json.JSONDecodeError):
        parsed = []
    return _normalize_bonus_category_ids(parsed)



def _normalize_bonus_category_ids(raw_rows: Any) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in raw_rows or []:
        category_id = str(value or '').strip()
        if not category_id or category_id in seen:
            continue
        seen.add(category_id)
        result.append(category_id)
    return result



def _normalize_category_name_key(value: Any) -> str:
    return ' '.join(str(value or '').strip().lower().replace('ё', 'е').split())



def _build_category_rate_name_map(rate_map: dict[str, dict[str, Any]]) -> dict[str, dict[str, Any]]:
    output: dict[str, dict[str, Any]] = {}
    for row in rate_map.values():
        key = _normalize_category_name_key((row or {}).get('category_name'))
        if key and key not in output:
            output[key] = row
    return output



def _get_rate_info_for_category(
    category_id: str | None,
    category_name: str | None,
    rate_map: dict[str, dict[str, Any]],
    rate_name_map: dict[str, dict[str, Any]] | None = None,
) -> dict[str, Any] | None:
    normalized_id = str(category_id or '').strip()
    if normalized_id and normalized_id in rate_map:
        return rate_map[normalized_id]
    name_key = _normalize_category_name_key(category_name)
    if not name_key:
        return None
    lookup = rate_name_map or _build_category_rate_name_map(rate_map)
    return lookup.get(name_key)



def _manager_rate_for_profit(net_profit: float, brackets: list[dict[str, float]] | None = None) -> float:
    for row in _normalize_manager_salary_brackets(brackets):
        threshold = float(row.get('threshold') or 0)
        rate_percent = float(row.get('rate_percent') or 0)
        if net_profit >= threshold:
            return rate_percent
    return 0.0



def _cache_is_fresh(created_at: float, ttl: float) -> bool:
    return (asyncio.get_running_loop().time() - created_at) <= ttl




def _normalize_optional_ms_value(value: Any) -> str | None:
    raw = str(value or '').strip()
    if not raw or raw.lower() in {'none', 'null', 'undefined'}:
        return None
    return raw


def _point_ms_token(point: LocationPoint | None) -> str | None:
    token = _normalize_optional_ms_value(point.ms_token if point else None)
    if token:
        return token
    return _normalize_optional_ms_value(settings.moysklad_token)


def _point_store_id(point: LocationPoint | None) -> str | None:
    store_id = _normalize_optional_ms_value(point.ms_store_id if point else None)
    if store_id:
        return store_id
    normalized = _normalize_location(point.name) if point and point.name else ''
    if normalized.lower() == (settings.store_dmitrov or '').strip().lower():
        return _normalize_optional_ms_value(settings.store_dmitrov_id)
    if normalized.lower() == (settings.store_dubna or '').strip().lower():
        return _normalize_optional_ms_value(settings.store_dubna_id)
    return None


def _point_ms_kwargs(point: LocationPoint | None) -> dict[str, str | None]:
    return {
        'token': _point_ms_token(point),
        'store_id': _point_store_id(point),
    }

def _clean_text(value: str | None) -> str:
    return ' '.join(str(value or '').strip().lower().replace('ё', 'е').split())


def _candidate_name_variants(value: str | None) -> set[str]:
    raw = _clean_text(value)
    if not raw:
        return set()
    variants = {raw}
    separators = [':', '-', '—', '(', ')', '[', ']', '/', '\\']
    parts = {raw}
    for sep in separators:
        next_parts = set()
        for part in parts:
            next_parts.update(p.strip() for p in part.split(sep) if p.strip())
        parts |= next_parts
    variants |= parts
    return {item for item in variants if item}


async def _log_payroll_action(
    db: AsyncSession,
    *,
    actor_user_id: int | None,
    location_point_id: int | None,
    entity_type: str,
    entity_id: str | None,
    action_type: str,
    details: dict[str, Any],
) -> None:
    db.add(PayrollAuditLog(
        actor_user_id=actor_user_id,
        location_point_id=location_point_id,
        entity_type=entity_type,
        entity_id=entity_id,
        action_type=action_type,
        details_json=json.dumps(details, ensure_ascii=False),
    ))


async def _get_location_point_by_name(location: str, db: AsyncSession) -> LocationPoint:
    normalized = _normalize_location(location)
    point = await db.scalar(select(LocationPoint).where(LocationPoint.name == normalized).limit(1))
    if not point:
        raise HTTPException(status_code=404, detail='Точка не найдена.')
    return point


async def get_user_accessible_locations(user: User, db: AsyncSession) -> list[str]:
    fallback_location = _normalize_location(user.location) if user.location else ''

    def _merge_locations(rows: list[LocationPoint]) -> list[str]:
        result: list[str] = []
        seen: set[str] = set()
        for row in rows:
            name = _normalize_location(getattr(row, 'name', '') or '')
            if name and name not in seen:
                seen.add(name)
                result.append(name)
        if fallback_location and fallback_location not in seen:
            result.append(fallback_location)
        return result

    if user.role == 'superadmin':
        rows = (await db.scalars(select(LocationPoint).order_by(LocationPoint.name.asc()))).all()
        return _merge_locations(rows)
    if user.role == 'admin':
        rows = (
            await db.scalars(
                select(LocationPoint)
                .join(AdminLocationAccess, AdminLocationAccess.location_point_id == LocationPoint.id)
                .where(AdminLocationAccess.admin_user_id == user.id)
                .order_by(LocationPoint.name.asc())
            )
        ).all()
        return _merge_locations(rows)
    return [fallback_location] if fallback_location else []


async def ensure_user_can_access_location(user: User, location: str, db: AsyncSession) -> None:
    normalized = _normalize_location(location)
    accessible = set(await get_user_accessible_locations(user, db))
    if normalized not in accessible:
        raise HTTPException(status_code=403, detail='Нет доступа к выбранной точке.')


async def _ensure_default_payroll_settings(point: LocationPoint, db: AsyncSession) -> PayrollSettingsVersion:
    version = await db.scalar(
        select(PayrollSettingsVersion)
        .where(PayrollSettingsVersion.location_point_id == point.id)
        .order_by(PayrollSettingsVersion.effective_from.desc(), PayrollSettingsVersion.id.desc())
        .limit(1)
    )
    if version:
        return version

    responsible_admin_user_id = await db.scalar(
        select(AdminLocationAccess.admin_user_id)
        .where(AdminLocationAccess.location_point_id == point.id)
        .order_by(AdminLocationAccess.id.asc())
        .limit(1)
    )
    version = PayrollSettingsVersion(
        location_point_id=point.id,
        effective_from=get_moscow_today(),
        exit_amount=DEFAULT_EXIT_AMOUNT,
        bonus_threshold=DEFAULT_BONUS_THRESHOLD,
        bonus_amount=DEFAULT_BONUS_AMOUNT,
        other_rate_percent=DEFAULT_OTHER_RATE_PERCENT,
        bonus_category_ids_json='[]',
        manager_salary_brackets_json=json.dumps(_default_manager_salary_brackets(), ensure_ascii=False),
        responsible_admin_user_id=responsible_admin_user_id,
    )
    db.add(version)
    await db.flush()
    await _log_payroll_action(
        db,
        actor_user_id=None,
        location_point_id=point.id,
        entity_type='payroll_settings',
        entity_id=str(version.id),
        action_type='auto_create',
        details={
            'effective_from': version.effective_from.isoformat(),
            'exit_amount': version.exit_amount,
            'bonus_threshold': version.bonus_threshold,
            'bonus_amount': version.bonus_amount,
            'other_rate_percent': version.other_rate_percent,
            'bonus_category_ids': [],
            'manager_salary_brackets': _default_manager_salary_brackets(),
            'responsible_admin_user_id': version.responsible_admin_user_id,
        },
    )
    await db.commit()
    await db.refresh(version)
    return version


async def _get_settings_for_date(point: LocationPoint, target_date: date, db: AsyncSession) -> PayrollSettingsVersion:
    version = await db.scalar(
        select(PayrollSettingsVersion)
        .where(
            PayrollSettingsVersion.location_point_id == point.id,
            PayrollSettingsVersion.effective_from <= target_date,
        )
        .order_by(PayrollSettingsVersion.effective_from.desc(), PayrollSettingsVersion.id.desc())
        .limit(1)
    )
    if version:
        return version
    version = await _ensure_default_payroll_settings(point, db)
    return version


async def _get_settings_rates(settings_version_id: int, db: AsyncSession) -> dict[str, dict[str, Any]]:
    rows = (
        await db.scalars(
            select(PayrollCategoryRateVersion)
            .where(PayrollCategoryRateVersion.settings_version_id == settings_version_id)
            .order_by(PayrollCategoryRateVersion.category_name.asc())
        )
    ).all()
    return {
        row.category_id: {
            'category_id': row.category_id,
            'category_name': row.category_name,
            'rate_percent': float(row.rate_percent or 0),
        }
        for row in rows
    }


def _serialize_settings_version_payload(settings: PayrollSettingsVersion, rates: dict[str, dict[str, Any]]) -> dict[str, Any]:
    return {
        'id': settings.id,
        'effective_from': settings.effective_from.isoformat(),
        'exit_amount': round(float(settings.exit_amount or 0), 2),
        'bonus_threshold': round(float(settings.bonus_threshold or 0), 2),
        'bonus_amount': round(float(settings.bonus_amount or 0), 2),
        'other_rate_percent': round(float(settings.other_rate_percent or 0), 2),
        'bonus_category_ids': _load_bonus_category_ids(settings),
        'manager_salary_brackets': _load_manager_salary_brackets(settings),
        'responsible_admin_user_id': settings.responsible_admin_user_id,
        'category_rates': sorted(rates.values(), key=lambda item: item['category_name'].lower()),
    }


def _resolve_category_rate_percent(rate_info: dict[str, Any] | None, other_rate_percent: float | None) -> tuple[float, bool]:
    try:
        explicit_rate = float((rate_info or {}).get('rate_percent') or 0)
    except (TypeError, ValueError, AttributeError):
        explicit_rate = 0.0
    fallback_rate = max(float(other_rate_percent or 0), 0.0)
    if explicit_rate > 0:
        return round(explicit_rate, 2), False
    return round(fallback_rate, 2), True


async def _list_location_employees(point: LocationPoint, db: AsyncSession) -> list[User]:
    return (
        await db.scalars(
            select(User)
            .where(
                User.role == 'employee',
                User.location == point.name,
                User.is_active.is_(True),
            )
            .order_by(User.full_name.asc())
        )
    ).all()


async def _list_location_admins(point: LocationPoint, db: AsyncSession) -> list[User]:
    return (
        await db.scalars(
            select(User)
            .join(AdminLocationAccess, AdminLocationAccess.admin_user_id == User.id)
            .where(
                User.role.in_(['admin', 'superadmin']),
                User.is_active.is_(True),
                AdminLocationAccess.location_point_id == point.id,
            )
            .order_by(User.full_name.asc())
        )
    ).all()


async def get_location_payroll_setup(location: str, db: AsyncSession, current_user: User, effective_from: date | None = None) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    requested_effective_from = effective_from.isoformat() if effective_from else None
    settings = await _get_settings_for_date(point, effective_from, db) if effective_from else await _ensure_default_payroll_settings(point, db)
    rates = await _get_settings_rates(settings.id, db)
    employees = await _list_location_employees(point, db)
    admins = await _list_location_admins(point, db)
    category_catalog = await get_payroll_category_catalog(point.name, db, current_user)
    return {
        'location': point.name,
        'location_id': point.id,
        'requested_effective_from': requested_effective_from or settings.effective_from.isoformat(),
        'settings': _serialize_settings_version_payload(settings, rates),
        'category_catalog': category_catalog.get('categories', []),
        'employees': [{'id': item.id, 'full_name': item.full_name} for item in employees],
        'admins': [{'id': item.id, 'full_name': item.full_name, 'role': item.role} for item in admins],
    }


async def get_location_shift_setup(location: str, db: AsyncSession, current_user: User) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    employees = await _list_location_employees(point, db)
    admins = await _list_location_admins(point, db)
    return {
        'location': point.name,
        'location_id': point.id,
        'employees': [{'id': item.id, 'full_name': item.full_name} for item in employees],
        'admins': [{'id': item.id, 'full_name': item.full_name, 'role': item.role} for item in admins],
    }


async def get_payroll_category_catalog(location: str, db: AsyncSession, current_user: User) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    normalized = _normalize_location(location)
    categories: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    seen_names: dict[str, str] = {}

    def _append_category(category_id: Any, category_name: Any) -> None:
        normalized_id = str(category_id or '').strip()
        normalized_name = str(category_name or '').strip()
        if not normalized_id or not normalized_name or normalized_id == '__other__' or normalized_name == DEFAULT_CATEGORY_NAME:
            return
        name_key = _normalize_category_name_key(normalized_name)
        if normalized_id in seen_ids:
            return
        # Если уже есть та же категория по имени с осмысленным id, не плодим дубликаты с другим id.
        if name_key and name_key in seen_names:
            return
        seen_ids.add(normalized_id)
        if name_key:
            seen_names[name_key] = normalized_id
        categories.append({'id': normalized_id, 'name': normalized_name})

    # 1) Категории из всех сохраненных версий правил по точке.
    historical_rate_rows = (
        await db.execute(
            select(PayrollCategoryRateVersion.category_id, PayrollCategoryRateVersion.category_name)
            .join(PayrollSettingsVersion, PayrollCategoryRateVersion.settings_version_id == PayrollSettingsVersion.id)
            .where(PayrollSettingsVersion.location_point_id == point.id)
            .order_by(PayrollCategoryRateVersion.category_name.asc())
        )
    ).all()
    for category_id, category_name in historical_rate_rows:
        _append_category(category_id, category_name)

    # 1.1) Категории из истории сохранений правил, включая нулевые проценты.
    audit_rows = (
        await db.scalars(
            select(PayrollAuditLog)
            .where(
                PayrollAuditLog.location_point_id == point.id,
                PayrollAuditLog.entity_type == 'payroll_settings',
            )
            .order_by(PayrollAuditLog.id.desc())
            .limit(200)
        )
    ).all()
    for audit_row in audit_rows:
        try:
            details = json.loads(audit_row.details_json or '{}')
        except (TypeError, ValueError, json.JSONDecodeError):
            details = {}
        for item in details.get('category_rates') or []:
            _append_category(item.get('category_id') or item.get('id'), item.get('category_name') or item.get('name'))

    # 1.2) Категории из локальной истории ревизий/снапшотов по этой точке.
    revision_queries = [
        select(SelectionTarget.category_id, SelectionTarget.category_name).where(SelectionTarget.location == normalized),
        select(SelectionTargetDay.category_id, SelectionTargetDay.category_name).where(SelectionTargetDay.location == normalized),
        select(CategoryAssignment.category_id, CategoryAssignment.category_name).where(CategoryAssignment.location == normalized),
        select(ReportTargetSnapshot.category_id, ReportTargetSnapshot.category_name)
        .join(Report, Report.id == ReportTargetSnapshot.report_id)
        .where(Report.location == normalized),
        select(CheckResult.category_id, CheckResult.category_name)
        .join(Report, Report.id == CheckResult.report_id)
        .where(Report.location == normalized),
    ]
    for query in revision_queries:
        for category_id, category_name in (await db.execute(query.distinct())).all():
            _append_category(category_id, category_name)

    # 1.3) Категории из снимков закрытых смен — это помогает, даже если по текущей дате продаж нет.
    snapshot_rows = (
        await db.execute(
            select(ShiftPayrollCategorySnapshot.category_id, ShiftPayrollCategorySnapshot.category_name)
            .join(ShiftPayrollSnapshot, ShiftPayrollCategorySnapshot.snapshot_id == ShiftPayrollSnapshot.id)
            .where(ShiftPayrollSnapshot.location_point_id == point.id)
            .order_by(ShiftPayrollCategorySnapshot.category_name.asc())
        )
    ).all()
    for category_id, category_name in snapshot_rows:
        _append_category(category_id, category_name)

    # 2) Категории из кеша метрик продаж — здесь бывают категории, которых уже нет в остатках.
    metric_rows = (
        await db.scalars(
            select(PayrollDailyMetricCache)
            .where(PayrollDailyMetricCache.location_point_id == point.id)
            .order_by(PayrollDailyMetricCache.metric_date.desc(), PayrollDailyMetricCache.id.desc())
            .limit(730)
        )
    ).all()
    for metric_row in metric_rows:
        try:
            parsed_categories = json.loads(metric_row.categories_json or '[]')
        except (TypeError, ValueError, json.JSONDecodeError):
            parsed_categories = []
        for item in parsed_categories or []:
            _append_category(item.get('category_id'), item.get('category_name'))

    # 3) Пытаемся подтянуть полный каталог из МоегоСклада.
    try:
        if _ms_client_enabled(token=_point_ms_token(point), location=normalized):
            token = _point_ms_token(point)
            folder_map = await ms_client._get_folder_map(token=token, location=normalized)
            # Берём верхнюю категорию для каждой папки, а не только "явные" корни.
            for folder_id, folder in folder_map.items():
                chain = ms_client._resolve_folder_chain(str(folder_id or '').strip(), folder_map)
                if not chain:
                    continue
                top_category = chain[0]
                _append_category(f"cat-{normalized.lower()}-{top_category['id']}", top_category['name'])

            inventory = await ms_client.get_inventory(normalized, **_point_ms_kwargs(point))
            for category in inventory.get('categories', []):
                _append_category(category.get('id'), category.get('name'))
        else:
            from app.logic import MOCK_INVENTORY
            inventory = MOCK_INVENTORY.get(normalized, {'categories': []})
            for category in inventory.get('categories', []):
                _append_category(category.get('id'), category.get('name'))
    except Exception:
        logger.exception('Не удалось полностью загрузить каталог категорий для точки %s. Используем сохранённые и кешированные категории.', normalized)

    if not categories:
        settings = await _ensure_default_payroll_settings(point, db)
        rates = await _get_settings_rates(settings.id, db)
        for item in rates.values():
            _append_category(item.get('category_id'), item.get('category_name'))

    categories.sort(key=lambda item: item['name'].lower())
    return {'location': normalized, 'categories': categories}


async def update_location_payroll_settings(payload: PayrollSettingsUpdateRequest, db: AsyncSession, current_user: User) -> dict[str, Any]:
    if current_user.role not in {'admin', 'superadmin'}:
        raise HTTPException(status_code=403, detail='Изменять настройки зарплаты может только управляющий.')
    await ensure_user_can_access_location(current_user, payload.location, db)
    point = await _get_location_point_by_name(payload.location, db)

    if payload.responsible_admin_user_id:
        admin_user = await db.get(User, payload.responsible_admin_user_id)
        if not admin_user or admin_user.role not in {'admin', 'superadmin'} or not admin_user.is_active:
            raise HTTPException(status_code=400, detail='Ответственный управляющий не найден.')
        has_access = await db.scalar(
            select(func.count())
            .select_from(AdminLocationAccess)
            .where(
                AdminLocationAccess.admin_user_id == admin_user.id,
                AdminLocationAccess.location_point_id == point.id,
            )
        )
        if (has_access or 0) <= 0:
            raise HTTPException(status_code=400, detail='У выбранного управляющего нет доступа к точке.')

    if current_user.role != 'superadmin' and payload.manager_salary_brackets:
        raise HTTPException(status_code=403, detail='Настраивать зарплату управляющего может только главный управляющий.')

    normalized_bonus_category_ids = _normalize_bonus_category_ids(payload.bonus_category_ids)

    existing_version = await db.scalar(
        select(PayrollSettingsVersion)
        .where(
            PayrollSettingsVersion.location_point_id == point.id,
            PayrollSettingsVersion.effective_from == payload.effective_from,
        )
        .order_by(PayrollSettingsVersion.id.desc())
        .limit(1)
    )

    settings_source = existing_version or await _ensure_default_payroll_settings(point, db)
    manager_salary_brackets = _normalize_manager_salary_brackets(payload.manager_salary_brackets) if current_user.role == 'superadmin' else _load_manager_salary_brackets(settings_source)

    if existing_version is None:
        version = PayrollSettingsVersion(
            location_point_id=point.id,
            effective_from=payload.effective_from,
            exit_amount=payload.exit_amount,
            bonus_threshold=payload.bonus_threshold,
            bonus_amount=payload.bonus_amount,
            other_rate_percent=payload.other_rate_percent,
            bonus_category_ids_json=json.dumps(normalized_bonus_category_ids, ensure_ascii=False),
            manager_salary_brackets_json=json.dumps(manager_salary_brackets, ensure_ascii=False),
            responsible_admin_user_id=payload.responsible_admin_user_id,
            created_by_user_id=current_user.id,
        )
        db.add(version)
        await db.flush()
        audit_action = 'create_version'
    else:
        version = existing_version
        version.exit_amount = payload.exit_amount
        version.bonus_threshold = payload.bonus_threshold
        version.bonus_amount = payload.bonus_amount
        version.other_rate_percent = payload.other_rate_percent
        version.bonus_category_ids_json = json.dumps(normalized_bonus_category_ids, ensure_ascii=False)
        if current_user.role == 'superadmin' or not getattr(version, 'manager_salary_brackets_json', None):
            version.manager_salary_brackets_json = json.dumps(manager_salary_brackets, ensure_ascii=False)
        version.responsible_admin_user_id = payload.responsible_admin_user_id
        version.created_by_user_id = current_user.id
        await db.execute(delete(PayrollCategoryRateVersion).where(PayrollCategoryRateVersion.settings_version_id == version.id))
        await db.flush()
        audit_action = 'update_version'

    seen_ids: set[str] = set()
    normalized_rates: list[dict[str, Any]] = []
    for raw in payload.category_rates:
        category_id = str(raw.get('category_id') or raw.get('id') or '').strip()
        category_name = str(raw.get('category_name') or raw.get('name') or '').strip()
        if not category_id or not category_name or category_id in seen_ids:
            continue
        seen_ids.add(category_id)
        try:
            rate_percent = float(raw.get('rate_percent') or 0)
        except (TypeError, ValueError):
            rate_percent = 0.0
        normalized_rate = max(rate_percent, 0.0)
        normalized_rates.append({'category_id': category_id, 'category_name': category_name, 'rate_percent': normalized_rate})
        # Сохраняем и нулевые проценты тоже: это позволяет не терять полный список категорий
        # после перезагрузки страницы и стабильно хранить флаги бонуса к выходу.
        db.add(PayrollCategoryRateVersion(
            settings_version_id=version.id,
            category_id=category_id,
            category_name=category_name,
            rate_percent=normalized_rate,
        ))

    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='payroll_settings',
        entity_id=str(version.id),
        action_type=audit_action,
        details={
            'effective_from': payload.effective_from.isoformat(),
            'exit_amount': payload.exit_amount,
            'bonus_threshold': payload.bonus_threshold,
            'bonus_amount': payload.bonus_amount,
            'other_rate_percent': payload.other_rate_percent,
            'bonus_category_ids': normalized_bonus_category_ids,
            'manager_salary_brackets': manager_salary_brackets,
            'responsible_admin_user_id': payload.responsible_admin_user_id,
            'category_rates': normalized_rates,
        },
    )
    await db.commit()

    today = get_moscow_today()
    rebuild_until = await db.scalar(
        select(func.max(WorkShift.shift_date))
        .where(
            WorkShift.location_point_id == point.id,
            WorkShift.is_deleted.is_(False),
            WorkShift.status == 'closed',
            WorkShift.shift_date >= payload.effective_from,
        )
    )
    rebuild_to = rebuild_until or today
    recalc_job_payload: dict[str, Any] | None = None
    if payload.effective_from <= rebuild_to:
        recalc_job = await _enqueue_payroll_recalc_job(point, version.id, payload.effective_from, rebuild_to, current_user.id)
        recalc_job_payload = _serialize_recalc_job(recalc_job, point)

    saved_settings = await db.get(PayrollSettingsVersion, version.id)
    saved_rates = await _get_settings_rates(version.id, db)
    setup = await get_location_payroll_setup(point.name, db, current_user, effective_from=payload.effective_from)
    setup['settings'] = _serialize_settings_version_payload(saved_settings or version, saved_rates)
    setup['recalc_job'] = recalc_job_payload
    if recalc_job_payload is None:
        setup['rebuild_closed_shifts'] = {
            'date_from': payload.effective_from.isoformat(),
            'date_to': rebuild_to.isoformat(),
            'location': point.name,
            'processed': 0,
            'updated': 0,
            'details': [],
        }
    return setup


async def _ensure_month_expense_entries(point: LocationPoint, month_start: date, db: AsyncSession) -> list[MonthlyExpenseEntry]:
    templates = (
        await db.scalars(
            select(ExpenseTemplate)
            .where(ExpenseTemplate.location_point_id == point.id, ExpenseTemplate.is_active.is_(True))
            .order_by(ExpenseTemplate.name.asc())
        )
    ).all()
    template_by_id = {template.id: template for template in templates}
    existing_entries = (
        await db.scalars(
            select(MonthlyExpenseEntry)
            .where(
                MonthlyExpenseEntry.location_point_id == point.id,
                MonthlyExpenseEntry.month_start == month_start,
            )
        )
    ).all()
    entry_by_template = {entry.template_id: entry for entry in existing_entries}
    changed = False
    for entry in existing_entries:
        template = template_by_id.get(entry.template_id) if entry.template_id is not None else None
        normalized_mode = _expense_entry_mode(entry)
        if entry.distribution_mode != normalized_mode:
            entry.distribution_mode = normalized_mode
            changed = True
        resolved_expense_date = _resolve_entry_expense_date(entry, template)
        if entry.expense_date != resolved_expense_date:
            entry.expense_date = resolved_expense_date
            changed = True
    for template in templates:
        if template.id in entry_by_template:
            continue
        amount = float(template.default_amount or 0)
        entry = MonthlyExpenseEntry(
            template_id=template.id,
            location_point_id=point.id,
            month_start=month_start,
            expense_date=_default_expense_date_for_month(month_start, day_of_month=template.day_of_month),
            distribution_mode=EXPENSE_MODE_SPREAD,
            amount=amount,
            is_paid=False,
            assigned_employee_user_id=None,
            apply_to_employee_salary=bool(template.assign_to_employee_by_default),
        )
        db.add(entry)
        changed = True
    if changed:
        await db.flush()
        existing_entries = (
            await db.scalars(
                select(MonthlyExpenseEntry)
                .where(
                    MonthlyExpenseEntry.location_point_id == point.id,
                    MonthlyExpenseEntry.month_start == month_start,
                )
                .order_by(MonthlyExpenseEntry.id.asc())
            )
        ).all()
    return existing_entries


async def list_expense_templates(location: str, db: AsyncSession, current_user: User) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    rows = (
        await db.scalars(
            select(ExpenseTemplate)
            .where(ExpenseTemplate.location_point_id == point.id)
            .order_by(ExpenseTemplate.name.asc())
        )
    ).all()
    return {
        'location': point.name,
        'templates': [
            {
                'id': row.id,
                'name': row.name,
                'amount_type': row.amount_type,
                'default_amount': round(float(row.default_amount or 0), 2) if row.default_amount is not None else None,
                'assign_to_employee_by_default': row.assign_to_employee_by_default,
                'is_active': row.is_active,
            }
            for row in rows
        ],
    }


async def create_expense_template(payload: ExpenseTemplateCreateRequest, db: AsyncSession, current_user: User) -> dict[str, Any]:
    if current_user.role not in {'admin', 'superadmin'}:
        raise HTTPException(status_code=403, detail='Создавать шаблоны расходов может только управляющий.')
    await ensure_user_can_access_location(current_user, payload.location, db)
    point = await _get_location_point_by_name(payload.location, db)
    normalized_name = payload.name.strip()
    if not normalized_name:
        raise HTTPException(status_code=400, detail='Введите название.')
    amount_type = payload.amount_type if payload.amount_type in {'static', 'dynamic'} else 'dynamic'
    template = ExpenseTemplate(
        location_point_id=point.id,
        name=normalized_name,
        amount_type=amount_type,
        default_amount=payload.default_amount,
        assign_to_employee_by_default=payload.assign_to_employee_by_default,
        created_by_user_id=current_user.id,
    )
    db.add(template)
    await db.flush()
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='expense_template',
        entity_id=str(template.id),
        action_type='create',
        details={
            'name': template.name,
            'amount_type': template.amount_type,
            'default_amount': template.default_amount,
            'assign_to_employee_by_default': template.assign_to_employee_by_default,
        },
    )
    await db.commit()
    return await list_expense_templates(point.name, db, current_user)


async def update_expense_template(template_id: int, payload: ExpenseTemplateUpdateRequest, db: AsyncSession, current_user: User) -> dict[str, Any]:
    template = await db.get(ExpenseTemplate, template_id)
    if not template:
        raise HTTPException(status_code=404, detail='Шаблон расхода не найден.')
    point = await db.get(LocationPoint, template.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка шаблона не найдена.')
    await ensure_user_can_access_location(current_user, point.name, db)

    before = {
        'name': template.name,
        'amount_type': template.amount_type,
        'default_amount': template.default_amount,
        'assign_to_employee_by_default': template.assign_to_employee_by_default,
        'is_active': template.is_active,
    }
    normalized_name = payload.name.strip()
    if not normalized_name:
        raise HTTPException(status_code=400, detail='Введите название.')
    template.name = normalized_name
    template.amount_type = payload.amount_type if payload.amount_type in {'static', 'dynamic'} else 'dynamic'
    template.default_amount = payload.default_amount
    template.assign_to_employee_by_default = payload.assign_to_employee_by_default
    template.is_active = payload.is_active
    template.updated_at = datetime.utcnow()
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='expense_template',
        entity_id=str(template.id),
        action_type='update',
        details={'before': before, 'after': {
            'name': template.name,
            'amount_type': template.amount_type,
            'default_amount': template.default_amount,
            'assign_to_employee_by_default': template.assign_to_employee_by_default,
            'is_active': template.is_active,
        }},
    )
    await db.commit()
    return await list_expense_templates(point.name, db, current_user)




def _expense_entry_display_name(entry: MonthlyExpenseEntry, template_map: dict[int, ExpenseTemplate]) -> str:
    if entry.custom_name:
        return entry.custom_name
    template = template_map.get(entry.template_id) if entry.template_id is not None else None
    return template.name if template else 'Расход'


def _serialize_monthly_expense_entry(
    entry: MonthlyExpenseEntry,
    *,
    template_map: dict[int, ExpenseTemplate],
    employee_names: dict[int, str],
) -> dict[str, Any]:
    template = template_map.get(entry.template_id) if entry.template_id is not None else None
    return {
        'id': entry.id,
        'template_id': entry.template_id,
        'template_name': template.name if template else None,
        'name': _expense_entry_display_name(entry, template_map),
        'is_manual': entry.template_id is None,
        'amount_type': template.amount_type if template else 'manual',
        'month_start': entry.month_start.isoformat(),
        'expense_date': _resolve_entry_expense_date(entry, template).isoformat(),
        'distribution_mode': _expense_entry_mode(entry),
        'amount': round(float(entry.amount or 0), 2),
        'is_paid': entry.is_paid,
        'assigned_employee_user_id': entry.assigned_employee_user_id,
        'assigned_employee_name': employee_names.get(entry.assigned_employee_user_id),
        'apply_to_employee_salary': entry.apply_to_employee_salary,
        'comment': entry.comment or '',
        'created_at': _datetime_to_str(entry.created_at),
        'updated_at': _datetime_to_str(entry.updated_at),
    }


async def create_manual_monthly_expense(payload: ManualMonthlyExpenseCreateRequest, db: AsyncSession, current_user: User) -> dict[str, Any]:
    if current_user.role not in {'admin', 'superadmin'}:
        raise HTTPException(status_code=403, detail='Создавать расходы может только управляющий.')
    await ensure_user_can_access_location(current_user, payload.location, db)
    point = await _get_location_point_by_name(payload.location, db)
    month_key = _month_start(payload.month_start)
    normalized_name = payload.name.strip()
    if not normalized_name:
        raise HTTPException(status_code=400, detail='Введите название.')
    if payload.assigned_employee_user_id:
        employee = await db.get(User, payload.assigned_employee_user_id)
        if not employee or employee.role != 'employee' or employee.location != point.name:
            raise HTTPException(status_code=400, detail='Нельзя привязать расход к этому сотруднику.')
    entry = MonthlyExpenseEntry(
        template_id=None,
        location_point_id=point.id,
        month_start=month_key,
        expense_date=_clip_expense_date_to_month(payload.expense_date or month_key, month_key),
        distribution_mode=_normalize_expense_distribution_mode(payload.distribution_mode, default=EXPENSE_MODE_SINGLE_DAY),
        custom_name=normalized_name,
        comment=(payload.comment or '').strip() or None,
        amount=payload.amount,
        is_paid=payload.is_paid,
        assigned_employee_user_id=payload.assigned_employee_user_id,
        apply_to_employee_salary=payload.apply_to_employee_salary,
        created_by_user_id=current_user.id,
        updated_by_user_id=current_user.id,
        updated_at=datetime.utcnow(),
    )
    db.add(entry)
    await db.flush()
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='monthly_expense',
        entity_id=str(entry.id),
        action_type='create_manual',
        details={
            'name': entry.custom_name,
            'month_start': month_key.isoformat(),
            'expense_date': entry.expense_date.isoformat() if entry.expense_date else None,
            'distribution_mode': entry.distribution_mode,
            'amount': entry.amount,
            'assigned_employee_user_id': entry.assigned_employee_user_id,
            'apply_to_employee_salary': entry.apply_to_employee_salary,
            'is_paid': entry.is_paid,
            'comment': entry.comment,
        },
    )
    await db.commit()
    return await list_monthly_expenses(point.name, month_key, db, current_user)



async def deactivate_expense_template(template_id: int, db: AsyncSession, current_user: User) -> dict[str, Any]:
    template = await db.get(ExpenseTemplate, template_id)
    if not template:
        raise HTTPException(status_code=404, detail='Шаблон расхода не найден.')
    point = await db.get(LocationPoint, template.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка шаблона не найдена.')
    await ensure_user_can_access_location(current_user, point.name, db)
    template.is_active = not bool(template.is_active)
    template.updated_at = datetime.utcnow()
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='expense_template',
        entity_id=str(template.id),
        action_type='activate' if template.is_active else 'deactivate',
        details={'name': template.name, 'is_active': template.is_active},
    )
    await db.commit()
    return await list_expense_templates(point.name, db, current_user)


async def delete_expense_template(template_id: int, db: AsyncSession, current_user: User) -> dict[str, Any]:
    template = await db.get(ExpenseTemplate, template_id)
    if not template:
        raise HTTPException(status_code=404, detail='Шаблон расхода не найден.')
    point = await db.get(LocationPoint, template.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка шаблона не найдена.')
    await ensure_user_can_access_location(current_user, point.name, db)

    linked_entries = (
        await db.scalars(
            select(MonthlyExpenseEntry)
            .where(MonthlyExpenseEntry.template_id == template.id)
            .order_by(MonthlyExpenseEntry.month_start.asc(), MonthlyExpenseEntry.id.asc())
        )
    ).all()
    detached_entries_count = 0
    now = datetime.utcnow()
    for entry in linked_entries:
        if not entry.custom_name:
            entry.custom_name = template.name
        entry.template_id = None
        entry.updated_by_user_id = current_user.id
        entry.updated_at = now
        detached_entries_count += 1

    await db.flush()
    template_name = template.name
    await db.delete(template)
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='expense_template',
        entity_id=str(template_id),
        action_type='delete',
        details={'name': template_name, 'detached_entries_count': detached_entries_count},
    )
    await db.commit()
    return await list_expense_templates(point.name, db, current_user)


async def delete_monthly_expense_entry(entry_id: int, db: AsyncSession, current_user: User) -> dict[str, Any]:
    entry = await db.get(MonthlyExpenseEntry, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail='Расход не найден.')
    point = await db.get(LocationPoint, entry.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка расхода не найдена.')
    await ensure_user_can_access_location(current_user, point.name, db)
    if entry.template_id is not None:
        raise HTTPException(status_code=400, detail='Удалять можно только свободные расходы без шаблона.')

    details = {
        'name': entry.custom_name or 'Свободный расход',
        'month_start': entry.month_start.isoformat(),
        'expense_date': entry.expense_date.isoformat() if entry.expense_date else None,
        'distribution_mode': entry.distribution_mode,
        'amount': float(entry.amount or 0),
        'assigned_employee_user_id': entry.assigned_employee_user_id,
        'comment': entry.comment,
    }
    await db.delete(entry)
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='monthly_expense',
        entity_id=str(entry_id),
        action_type='delete_manual',
        details=details,
    )
    await db.commit()
    return await list_monthly_expenses(point.name, entry.month_start, db, current_user)


async def get_monthly_expenses(location: str, month: date, db: AsyncSession, current_user: User) -> dict[str, Any]:
    return await list_monthly_expenses(location, month, db, current_user)


async def list_monthly_expenses(location: str, month: date, db: AsyncSession, current_user: User) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    month_key = _month_start(month)
    await _ensure_month_expense_entries(point, month_key, db)
    await db.commit()
    templates = {
        row.id: row
        for row in (
            await db.scalars(select(ExpenseTemplate).where(ExpenseTemplate.location_point_id == point.id))
        ).all()
    }
    employees = {user.id: user.full_name for user in await _list_location_employees(point, db)}
    entries = (
        await db.scalars(
            select(MonthlyExpenseEntry)
            .where(
                MonthlyExpenseEntry.location_point_id == point.id,
                MonthlyExpenseEntry.month_start == month_key,
            )
            .order_by(MonthlyExpenseEntry.updated_at.desc(), MonthlyExpenseEntry.id.desc())
        )
    ).all()
    return {
        'location': point.name,
        'month_start': month_key.isoformat(),
        'entries': [
            _serialize_monthly_expense_entry(entry, template_map=templates, employee_names=employees)
            for entry in entries
        ],
    }


async def update_monthly_expense_entry(entry_id: int, payload: MonthlyExpenseEntryUpdateRequest, db: AsyncSession, current_user: User) -> dict[str, Any]:
    entry = await db.get(MonthlyExpenseEntry, entry_id)
    if not entry:
        raise HTTPException(status_code=404, detail='Расход не найден.')
    point = await db.get(LocationPoint, entry.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка расхода не найдена.')
    await ensure_user_can_access_location(current_user, point.name, db)
    before = {
        'amount': entry.amount,
        'is_paid': entry.is_paid,
        'assigned_employee_user_id': entry.assigned_employee_user_id,
        'apply_to_employee_salary': entry.apply_to_employee_salary,
        'distribution_mode': entry.distribution_mode,
        'expense_date': entry.expense_date.isoformat() if entry.expense_date else None,
        'comment': entry.comment,
    }
    if payload.assigned_employee_user_id:
        employee = await db.get(User, payload.assigned_employee_user_id)
        if not employee or employee.role != 'employee' or employee.location != point.name:
            raise HTTPException(status_code=400, detail='Нельзя привязать расход к этому сотруднику.')
    entry.amount = payload.amount
    entry.is_paid = payload.is_paid
    entry.assigned_employee_user_id = payload.assigned_employee_user_id
    entry.apply_to_employee_salary = payload.apply_to_employee_salary
    entry.distribution_mode = _normalize_expense_distribution_mode(
        payload.distribution_mode,
        default=EXPENSE_MODE_SINGLE_DAY if entry.template_id is None else EXPENSE_MODE_SPREAD,
    )
    entry.expense_date = _clip_expense_date_to_month(
        payload.expense_date or _resolve_entry_expense_date(entry),
        entry.month_start,
    )
    entry.comment = (payload.comment or '').strip() or None
    entry.updated_by_user_id = current_user.id
    entry.updated_at = datetime.utcnow()
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='monthly_expense',
        entity_id=str(entry.id),
        action_type='update',
        details={'before': before, 'after': {
            'amount': entry.amount,
            'is_paid': entry.is_paid,
            'assigned_employee_user_id': entry.assigned_employee_user_id,
            'apply_to_employee_salary': entry.apply_to_employee_salary,
            'distribution_mode': entry.distribution_mode,
            'expense_date': entry.expense_date.isoformat() if entry.expense_date else None,
            'comment': entry.comment,
        }},
    )
    await db.commit()
    return await list_monthly_expenses(point.name, entry.month_start, db, current_user)


async def _build_top_category_lookup(point: LocationPoint) -> dict[str, dict[str, str]]:
    normalized = _normalize_location(point.name)
    cache_key = f"{point.id}"
    cached = _category_lookup_cache.get(cache_key)
    if cached and _cache_is_fresh(cached[0], SALES_METRICS_TTL_SECONDS):
        return dict(cached[1])

    lookup: dict[str, dict[str, str]] = {}
    try:
        if _ms_client_enabled(location=normalized):
            inventory = await ms_client.get_inventory(normalized, **_point_ms_kwargs(point))
        else:
            from app.logic import MOCK_INVENTORY
            inventory = MOCK_INVENTORY.get(normalized, {'categories': []})
    except Exception:
        logger.exception('Не удалось загрузить инвентарь для категоризации продаж точки %s. Используем fallback без категорий.', normalized)
        inventory = {'categories': []}

    for category in inventory.get('categories', []):
        category_name = category.get('name') or ''
        if category_name == DEFAULT_CATEGORY_NAME:
            continue
        for subcategory in category.get('subcategories', []):
            for item in subcategory.get('items', []):
                lookup[str(item['id'])] = {'category_id': str(category['id']), 'category_name': category_name}
    _category_lookup_cache[cache_key] = (asyncio.get_running_loop().time(), dict(lookup))
    return lookup


async def _fetch_document_rows(
    endpoint: str,
    date_from: date,
    date_to: date,
    point: LocationPoint,
    *,
    expand: str | None = None,
    include_positions: bool = False,
    positions_expand: str | None = None,
    positions_fields: str | None = None,
) -> list[dict[str, Any]]:
    if not _ms_client_enabled(token=_point_ms_token(point), location=point.name):
        return []
    return await ms_client.get_documents_by_period(
        endpoint,
        date_from,
        date_to,
        expand=expand,
        include_positions=include_positions,
        positions_expand=positions_expand,
        positions_fields=positions_fields,
        token=_point_ms_token(point),
        location=point.name,
    )


async def _fetch_retail_shift_rows(date_from: date, date_to: date, point: LocationPoint) -> list[dict[str, Any]]:
    return await _fetch_document_rows('retailshift', date_from, date_to, point, expand='store,retailStore')



def _extract_position_category_info(
    position: dict[str, Any],
    category_lookup: dict[str, dict[str, str]],
    category_ids_by_name: dict[str, str],
) -> dict[str, str] | None:
    assortment = position.get('assortment') or {}
    item_id = _extract_id(assortment)
    if item_id and item_id in category_lookup:
        return category_lookup[item_id]

    path_name = str(assortment.get('pathName') or assortment.get('path_name') or '').strip()
    if path_name:
        first_part = next((part.strip() for part in re.split(r'\s*/\s*', path_name) if part.strip()), '')
        if first_part and first_part in category_ids_by_name:
            return {
                'category_id': category_ids_by_name[first_part],
                'category_name': first_part,
            }

    folder = assortment.get('productFolder') or assortment.get('folder') or {}
    folder_name = str(folder.get('name') or '').strip()
    if folder_name and folder_name in category_ids_by_name:
        return {
            'category_id': category_ids_by_name[folder_name],
            'category_name': folder_name,
        }
    return None



def _iter_positions(doc: dict[str, Any]) -> list[dict[str, Any]]:
    positions = doc.get('positions')
    if isinstance(positions, dict):
        return positions.get('rows', []) or []
    if isinstance(positions, list):
        return positions
    return []


def _extract_document_day(doc: dict[str, Any]) -> date | None:
    for field in ('moment', 'openMoment', 'closeMoment', 'created', 'updated'):
        raw = str(doc.get(field) or '').strip()
        if len(raw) >= 10:
            try:
                return date.fromisoformat(raw[:10])
            except Exception:
                continue
    return None


def _extract_shift_sales_amount(doc: dict[str, Any]) -> float:
    proceeds_fields = ('proceedsCash', 'proceedsNoCash')
    if any(doc.get(field) is not None for field in proceeds_fields):
        amount = round(sum(_money(doc.get(field)) for field in proceeds_fields if doc.get(field) is not None), 2)
        if amount > 0:
            return amount

    for field in ('saleSum', 'retailSum', 'sum'):
        amount = _money(doc.get(field))
        if amount > 0:
            return amount

    payment_fields = (
        'cashSum',
        'noCashSum',
        'electronicSum',
        'qrSum',
        'prepaymentCashSum',
        'prepaymentNoCashSum',
        'prepaymentElectronicSum',
        'prepaymentQrSum',
        'receivedCash',
        'receivedNoCash',
    )
    amount = round(sum(_money(doc.get(field)) for field in payment_fields if doc.get(field) is not None), 2)
    return max(amount, 0.0)


def _extract_shift_return_amount(doc: dict[str, Any]) -> float:
    for field in ('returnSum', 'salesReturnSum'):
        amount = _money(doc.get(field))
        if amount > 0:
            return amount
    return 0.0


def _ensure_other_category_bucket(category_rows: dict[str, CategoryDocMetrics]) -> CategoryDocMetrics:
    return category_rows.setdefault(
        '__other__',
        CategoryDocMetrics(),
    )


def _iter_point_reference_candidates(doc: dict[str, Any]) -> list[dict[str, Any]]:
    collected: list[dict[str, Any]] = []
    seen: set[int] = set()

    def add(candidate: Any) -> None:
        if not isinstance(candidate, dict):
            return
        marker = id(candidate)
        if marker in seen:
            return
        seen.add(marker)
        collected.append(candidate)

    def add_store_refs(container: dict[str, Any] | None) -> None:
        if not isinstance(container, dict):
            return
        add(container.get('store'))
        add(container.get('retailStore'))
        add(container.get('retailstore'))

    add_store_refs(doc)

    retail_shift = doc.get('retailShift')
    add(retail_shift)
    add_store_refs(retail_shift if isinstance(retail_shift, dict) else None)

    demand = doc.get('demand')
    add(demand)
    if isinstance(demand, dict):
        add_store_refs(demand)
        demand_shift = demand.get('retailShift')
        add(demand_shift)
        add_store_refs(demand_shift if isinstance(demand_shift, dict) else None)

    return collected


def _doc_matches_point(doc: dict[str, Any], point: LocationPoint) -> bool:
    if doc.get('applicable') is False:
        return False

    point_ids = {str(point.ms_store_id or '').strip()} - {''}
    point_names = set()
    for raw in (point.name, point.ms_store_name):
        point_names |= _candidate_name_variants(raw)

    if not point_ids and not point_names:
        return True

    candidate_names: set[str] = set()
    candidate_ids: set[str] = set()
    for candidate in _iter_point_reference_candidates(doc):
        candidate_id = _extract_id(candidate)
        if candidate_id:
            candidate_ids.add(candidate_id)
        candidate_names |= _candidate_name_variants(candidate.get('name'))
        meta = candidate.get('meta') if isinstance(candidate.get('meta'), dict) else None
        if meta:
            candidate_names |= _candidate_name_variants(meta.get('name'))

    if point_ids & candidate_ids:
        return True

    if point_names and candidate_names:
        if point_names & candidate_names:
            return True
        for point_name in point_names:
            for candidate_name in candidate_names:
                if point_name in candidate_name or candidate_name in point_name:
                    return True

    return False


def _empty_day_metrics() -> dict[str, Any]:
    return {
        'categories': [],
        'gross_sales_amount': 0.0,
        'return_amount': 0.0,
        'net_sales_amount': 0.0,
        'cost_amount': 0.0,
        'gross_profit_amount': 0.0,
        'non_tobacco_net_sales_for_bonus': 0.0,
    }


def _extract_profit_report_amount(row: dict[str, Any], *fields: str) -> float:
    value = _first_money_value(row, tuple(fields))
    if value is None:
        return 0.0
    return round(value, 2)


def _extract_profit_report_category_info(
    row: dict[str, Any],
    category_lookup: dict[str, dict[str, str]],
    category_ids_by_name: dict[str, str],
) -> dict[str, str] | None:
    category = _extract_position_category_info(row, category_lookup, category_ids_by_name)
    if category is not None:
        return category

    for candidate_key in ('productFolder', 'folder', 'goodFolder', 'productfolder'):
        folder = row.get(candidate_key)
        if not isinstance(folder, dict):
            continue
        folder_name = str(folder.get('name') or '').strip()
        if folder_name and folder_name in category_ids_by_name:
            return {
                'category_id': category_ids_by_name[folder_name],
                'category_name': folder_name,
            }

    path_name = str(row.get('pathName') or row.get('path_name') or '').strip()
    if path_name:
        first_part = next((part.strip() for part in re.split(r'\s*/\s*', path_name) if part.strip()), '')
        if first_part and first_part in category_ids_by_name:
            return {
                'category_id': category_ids_by_name[first_part],
                'category_name': first_part,
            }
    return None


async def _load_profitability_metrics_by_day(point: LocationPoint, date_from: date, date_to: date) -> dict[date, ProfitReportDayMetrics]:
    if not _ms_client_enabled(token=_point_ms_token(point), location=point.name):
        return {}

    token = _point_ms_token(point)
    location = point.name
    store_id = _point_store_id(point)
    if not store_id:
        logger.warning('Для точки %s не задан ms_store_id, поэтому отчет прибыльности не может быть отфильтрован по точке. Используем legacy fallback.', point.name)
        return {}
    output: dict[date, ProfitReportDayMetrics] = {}

    for current_day in _daterange(date_from, date_to):
        try:
            rows = await ms_client.get_profit_report_by_product(
                current_day,
                current_day,
                token=token,
                location=location,
                store_id=store_id,
            )
        except Exception:
            logger.warning(
                'Не удалось получить отчет прибыльности МоегоСклада для точки %s за %s. Используем legacy fallback.',
                point.name,
                current_day.isoformat(),
                exc_info=True,
            )
            continue

        if not rows:
            continue

        metrics = ProfitReportDayMetrics(has_rows=True)
        for row in rows:
            metrics.sales_amount += _extract_profit_report_amount(row, 'sellSum')
            metrics.return_amount += _extract_profit_report_amount(row, 'returnSum')
            metrics.cost_amount += _extract_profit_report_amount(row, 'sellCostSum')
            metrics.cost_amount -= _extract_profit_report_amount(row, 'returnCostSum')
            metrics.gross_profit_amount += _extract_profit_report_amount(row, 'profit')

        metrics.sales_amount = round(metrics.sales_amount, 2)
        metrics.return_amount = round(metrics.return_amount, 2)
        metrics.cost_amount = round(metrics.cost_amount, 2)
        metrics.gross_profit_amount = round(metrics.gross_profit_amount, 2)
        output[current_day] = metrics

    return output


async def _load_point_sales_metrics_live(point: LocationPoint, date_from: date, date_to: date) -> dict[date, dict[str, Any]]:
    category_lookup = await _build_top_category_lookup(point)
    category_ids_by_name: dict[str, str] = {}
    for value in category_lookup.values():
        category_name = str(value.get('category_name') or '').strip()
        category_id = str(value.get('category_id') or '').strip()
        if category_name and category_id and category_name not in category_ids_by_name:
            category_ids_by_name[category_name] = category_id

    sales_rows, return_rows, shift_rows = await asyncio.gather(
        _fetch_document_rows(
            'retaildemand',
            date_from,
            date_to,
            point,
            expand='store,retailStore,retailShift,retailShift.store,retailShift.retailStore',
            include_positions=True,
            positions_expand='assortment,assortment.productFolder',
            positions_fields='stock',
        ),
        _fetch_document_rows(
            'retailsalesreturn',
            date_from,
            date_to,
            point,
            expand='store,retailStore,retailShift,retailShift.store,retailShift.retailStore,demand,demand.store,demand.retailStore,demand.retailShift,demand.retailShift.store,demand.retailShift.retailStore',
            include_positions=True,
            positions_expand='assortment,assortment.productFolder',
            positions_fields='stock',
        ),
        _fetch_retail_shift_rows(date_from, date_to, point),
    )
    sales_rows = [row for row in sales_rows if _doc_matches_point(row, point)]
    return_rows = [row for row in return_rows if _doc_matches_point(row, point)]
    shift_rows = [row for row in shift_rows if _doc_matches_point(row, point)]
    profit_report_by_day = await _load_profitability_metrics_by_day(point, date_from, date_to)

    by_day: dict[date, dict[str, CategoryDocMetrics]] = defaultdict(dict)
    totals_by_day: dict[date, dict[str, float]] = defaultdict(lambda: {
        'gross_sales': 0.0,
        'returns': 0.0,
        'sales_cost': 0.0,
        'return_cost': 0.0,
    })
    shift_sales_by_day: dict[date, float] = defaultdict(float)
    shift_returns_by_day: dict[date, float] = defaultdict(float)
    shift_cost_by_day: dict[date, float] = defaultdict(float)
    shift_profit_by_day: dict[date, float] = defaultdict(float)
    shift_cost_days: set[date] = set()
    shift_profit_days: set[date] = set()
    category_names_by_id: dict[str, str] = {}

    for doc, is_return in [(row, False) for row in sales_rows] + [(row, True) for row in return_rows]:
        doc_date = _extract_document_day(doc)
        if doc_date is None:
            continue
        for position in _iter_positions(doc):
            category = _extract_position_category_info(position, category_lookup, category_ids_by_name)
            if category:
                category_id = category['category_id']
                category_name = category['category_name']
            else:
                category_id = '__other__'
                category_name = DEFAULT_CATEGORY_NAME
            category_names_by_id[category_id] = category_name
            amount = _extract_position_amount(position)
            cost_amount = _extract_position_cost_amount(position) or 0.0
            bucket = by_day[doc_date].setdefault(category_id, CategoryDocMetrics())
            if is_return:
                bucket.returns += amount
                bucket.return_cost += cost_amount
                totals_by_day[doc_date]['returns'] += amount
                totals_by_day[doc_date]['return_cost'] += cost_amount
            else:
                bucket.sales += amount
                bucket.sales_cost += cost_amount
                totals_by_day[doc_date]['gross_sales'] += amount
                totals_by_day[doc_date]['sales_cost'] += cost_amount

    for shift_row in shift_rows:
        shift_day = _extract_document_day(shift_row)
        if shift_day is None:
            continue
        shift_sales_by_day[shift_day] += _extract_shift_sales_amount(shift_row)
        shift_returns_by_day[shift_day] += _extract_shift_return_amount(shift_row)

        shift_cost_amount = _extract_shift_cost_amount(shift_row)
        if shift_cost_amount is not None:
            shift_cost_by_day[shift_day] += shift_cost_amount
            shift_cost_days.add(shift_day)

        shift_profit_amount = _extract_shift_profit_amount(shift_row)
        if shift_profit_amount is not None:
            shift_profit_by_day[shift_day] += shift_profit_amount
            shift_profit_days.add(shift_day)

    output: dict[date, dict[str, Any]] = {}
    for day in _daterange(date_from, date_to):
        category_rows = by_day.get(day, {})
        doc_totals = totals_by_day.get(day, {'gross_sales': 0.0, 'returns': 0.0, 'sales_cost': 0.0, 'return_cost': 0.0})
        shift_sales_amount = round(float(shift_sales_by_day.get(day, 0.0) or 0.0), 2)
        shift_return_amount = round(float(shift_returns_by_day.get(day, 0.0) or 0.0), 2)
        gross_sales_amount = shift_sales_amount if shift_sales_amount > 0 else round(doc_totals['gross_sales'], 2)
        return_amount = shift_return_amount if shift_return_amount > 0 else round(doc_totals['returns'], 2)
        net_sales_amount = round(gross_sales_amount - return_amount, 2)
        doc_cost_amount = round(doc_totals['sales_cost'] - doc_totals['return_cost'], 2)
        legacy_cost_amount = round(float(shift_cost_by_day.get(day, 0.0) or 0.0), 2) if day in shift_cost_days else doc_cost_amount

        if day in shift_profit_days:
            legacy_gross_profit_amount = round(float(shift_profit_by_day.get(day, 0.0) or 0.0), 2)
            if day not in shift_cost_days:
                legacy_cost_amount = round(net_sales_amount - legacy_gross_profit_amount, 2)
        else:
            legacy_gross_profit_amount = round(net_sales_amount - legacy_cost_amount, 2)

        report_metrics = profit_report_by_day.get(day)
        if report_metrics and report_metrics.has_rows:
            cost_amount = round(report_metrics.cost_amount, 2)
            gross_profit_amount = round(report_metrics.gross_profit_amount, 2)
        else:
            cost_amount = legacy_cost_amount
            gross_profit_amount = legacy_gross_profit_amount

        categories = []
        non_tobacco_net = 0.0
        category_cost_sum = 0.0
        category_sales_sum = 0.0
        category_return_sum = 0.0
        for category_id, metrics in category_rows.items():
            category_name = category_names_by_id.get(category_id, DEFAULT_CATEGORY_NAME)
            net_sales = _sanitize_uncategorized_net(category_name, round(metrics.sales - metrics.returns, 2))
            category_cost = round(metrics.sales_cost - metrics.return_cost, 2)
            category_cost_sum += category_cost
            category_sales_sum += round(metrics.sales, 2)
            category_return_sum += round(metrics.returns, 2)
            if not _is_tobacco_category(category_name):
                category_bonus_base = _resolve_calculation_sales_base(metrics.sales, net_sales)
                non_tobacco_net += max(category_bonus_base, 0.0) if category_name == DEFAULT_CATEGORY_NAME else category_bonus_base
            categories.append({
                'category_id': category_id,
                'category_name': category_name,
                'sales_amount': round(metrics.sales, 2),
                'return_amount': round(metrics.returns, 2),
                'net_sales_amount': net_sales,
                'cost_amount': category_cost,
            })

        sales_delta = round(gross_sales_amount - category_sales_sum, 2)
        return_delta = round(return_amount - category_return_sum, 2)
        if abs(sales_delta) > 0.009 or abs(return_delta) > 0.009:
            other_bucket = next((item for item in categories if item['category_id'] == '__other__'), None)
            if other_bucket is None:
                other_bucket = {
                    'category_id': '__other__',
                    'category_name': DEFAULT_CATEGORY_NAME,
                    'sales_amount': 0.0,
                    'return_amount': 0.0,
                    'net_sales_amount': 0.0,
                    'cost_amount': 0.0,
                }
                categories.append(other_bucket)
            other_bucket['sales_amount'] = round(float(other_bucket.get('sales_amount') or 0.0) + sales_delta, 2)
            other_bucket['return_amount'] = round(float(other_bucket.get('return_amount') or 0.0) + return_delta, 2)
            other_bucket['net_sales_amount'] = _sanitize_uncategorized_net(DEFAULT_CATEGORY_NAME, round(float(other_bucket.get('sales_amount') or 0.0) - float(other_bucket.get('return_amount') or 0.0), 2))
            if not _is_tobacco_category(DEFAULT_CATEGORY_NAME):
                non_tobacco_net += max(_resolve_calculation_sales_base(sales_delta, round(sales_delta - return_delta, 2)), 0.0)

        if not categories and (gross_sales_amount > 0 or return_amount > 0 or abs(cost_amount) > 0.009):
            synthetic_net = round(gross_sales_amount - return_amount, 2)
            if not _is_tobacco_category(DEFAULT_CATEGORY_NAME):
                non_tobacco_net += max(_resolve_calculation_sales_base(gross_sales_amount, synthetic_net), 0.0)
            categories.append({
                'category_id': '__other__',
                'category_name': DEFAULT_CATEGORY_NAME,
                'sales_amount': gross_sales_amount,
                'return_amount': return_amount,
                'net_sales_amount': synthetic_net,
                'cost_amount': cost_amount,
            })
        elif categories:
            cost_delta = round(cost_amount - category_cost_sum, 2)
            if abs(cost_delta) > 0.009:
                other_bucket = next((item for item in categories if item['category_id'] == '__other__'), None)
                if other_bucket is None:
                    other_bucket = {
                        'category_id': '__other__',
                        'category_name': DEFAULT_CATEGORY_NAME,
                        'sales_amount': 0.0,
                        'return_amount': 0.0,
                        'net_sales_amount': 0.0,
                        'cost_amount': 0.0,
                    }
                    categories.append(other_bucket)
                other_bucket['cost_amount'] = round(float(other_bucket.get('cost_amount') or 0.0) + cost_delta, 2)

        categories.sort(key=lambda item: item['category_name'].lower())
        output[day] = {
            'categories': categories,
            'gross_sales_amount': gross_sales_amount,
            'return_amount': return_amount,
            'net_sales_amount': net_sales_amount,
            'cost_amount': round(cost_amount, 2),
            'gross_profit_amount': round(gross_profit_amount, 2),
            'non_tobacco_net_sales_for_bonus': round(non_tobacco_net, 2),
        }
    return output


async def _get_cached_point_sales_metrics(point: LocationPoint, date_from: date, date_to: date, db: AsyncSession) -> dict[date, dict[str, Any]]:
    rows = (
        await db.scalars(
            select(PayrollDailyMetricCache)
            .where(
                PayrollDailyMetricCache.location_point_id == point.id,
                PayrollDailyMetricCache.metric_date >= date_from,
                PayrollDailyMetricCache.metric_date <= date_to,
            )
            .order_by(PayrollDailyMetricCache.metric_date.asc())
        )
    ).all()
    cached: dict[date, dict[str, Any]] = {}
    for row in rows:
        try:
            categories = json.loads(row.categories_json or '[]')
            if not isinstance(categories, list):
                categories = []
        except Exception:
            categories = []
        cached[row.metric_date] = {
            'categories': categories,
            'gross_sales_amount': round(float(row.gross_sales_amount or 0), 2),
            'return_amount': round(float(row.return_amount or 0), 2),
            'net_sales_amount': round(float(row.net_sales_amount or 0), 2),
            'cost_amount': round(float(row.cost_amount or 0), 2),
            'gross_profit_amount': round(float(row.gross_profit_amount or 0), 2),
            'non_tobacco_net_sales_for_bonus': round(float(row.non_tobacco_net_sales_for_bonus or 0), 2),
            '_refreshed_at': row.updated_at,
        }
    return cached


async def _store_point_sales_metrics_cache(point: LocationPoint, metrics_by_day: dict[date, dict[str, Any]], db: AsyncSession) -> None:
    if not metrics_by_day:
        return
    metric_days = sorted(metrics_by_day)
    existing_rows = (
        await db.scalars(
            select(PayrollDailyMetricCache)
            .where(
                PayrollDailyMetricCache.location_point_id == point.id,
                PayrollDailyMetricCache.metric_date >= metric_days[0],
                PayrollDailyMetricCache.metric_date <= metric_days[-1],
            )
        )
    ).all()
    existing_by_day = {row.metric_date: row for row in existing_rows}
    timestamp = datetime.utcnow()
    for day, metrics in metrics_by_day.items():
        row = existing_by_day.get(day)
        categories_payload = _mark_category_sales_allocation_version(metrics.get('categories') or [])
        payload = {
            'gross_sales_amount': round(float(metrics.get('gross_sales_amount') or 0), 2),
            'return_amount': round(float(metrics.get('return_amount') or 0), 2),
            'net_sales_amount': round(float(metrics.get('net_sales_amount') or 0), 2),
            'cost_amount': round(float(metrics.get('cost_amount') or 0), 2),
            'gross_profit_amount': round(float(metrics.get('gross_profit_amount') or 0), 2),
            'non_tobacco_net_sales_for_bonus': round(float(metrics.get('non_tobacco_net_sales_for_bonus') or 0), 2),
            'categories_json': json.dumps(categories_payload, ensure_ascii=False),
            'source_refreshed_at': timestamp,
            'updated_at': timestamp,
        }
        if row is None:
            row = PayrollDailyMetricCache(
                location_point_id=point.id,
                metric_date=day,
                created_at=timestamp,
                **payload,
            )
            db.add(row)
        else:
            for key, value in payload.items():
                setattr(row, key, value)
    await db.flush()


def _is_cached_day_fresh(day: date, metrics: dict[str, Any] | None) -> bool:
    if metrics is None:
        return False

    categories = metrics.get('categories') or []
    if _has_legacy_uncategorized_adjustment(categories):
        return False

    visible_rows = [row for row in categories if str(row.get('category_name') or '').strip() != DEFAULT_CATEGORY_NAME]
    has_visible_sales = any(
        abs(float(row.get('sales_amount') or 0.0)) > 0.009
        or abs(float(row.get('net_sales_amount') or 0.0)) > 0.009
        for row in visible_rows
    )
    if has_visible_sales:
        total_cost_amount = round(float(metrics.get('cost_amount') or 0.0), 2)
        visible_cost_sum = round(sum(float(row.get('cost_amount') or 0.0) for row in visible_rows), 2)
        if abs(visible_cost_sum) <= 0.009:
            if abs(total_cost_amount) <= 0.009:
                return False
            other_cost_sum = round(sum(float(row.get('cost_amount') or 0.0) for row in categories if str(row.get('category_name') or '').strip() == DEFAULT_CATEGORY_NAME), 2)
            if abs(other_cost_sum - total_cost_amount) <= 0.01:
                return False

    today = get_payroll_operational_today()
    if day < today:
        return True
    if day > today:
        return False
    refreshed_at = metrics.get('_refreshed_at')
    if not isinstance(refreshed_at, datetime):
        return False
    age_seconds = (datetime.utcnow() - refreshed_at).total_seconds()
    return age_seconds <= PAYROLL_DAILY_CACHE_RECENT_TTL_SECONDS


async def _load_point_sales_metrics(
    point: LocationPoint,
    date_from: date,
    date_to: date,
    db: AsyncSession | None = None,
    *,
    force_refresh: bool = False,
) -> dict[date, dict[str, Any]]:
    cache_key = (point.id, date_from.isoformat(), date_to.isoformat())
    today = get_payroll_operational_today()
    includes_today = date_from <= today <= date_to
    if not force_refresh:
        cached = _sales_metrics_cache.get(cache_key)
        if cached and _cache_is_fresh(cached[0], SALES_METRICS_TTL_SECONDS):
            return cached[1]

    try:
        result: dict[date, dict[str, Any]] = {}
        missing_days: list[date] = _daterange(date_from, date_to)

        if db is not None and not force_refresh:
            cached_days = await _get_cached_point_sales_metrics(point, date_from, date_to, db)
            missing_days = []
            for day in _daterange(date_from, date_to):
                metrics = cached_days.get(day)
                if _is_cached_day_fresh(day, metrics):
                    result[day] = {key: value for key, value in metrics.items() if not key.startswith('_')}
                else:
                    missing_days.append(day)

        if missing_days:
            live_ranges: list[tuple[date, date]] = []
            range_start = missing_days[0]
            previous_day = missing_days[0]
            for day in missing_days[1:]:
                if day == previous_day + timedelta(days=1):
                    previous_day = day
                    continue
                live_ranges.append((range_start, previous_day))
                range_start = previous_day = day
            live_ranges.append((range_start, previous_day))

            refreshed_by_day: dict[date, dict[str, Any]] = {}
            for range_from, range_to in live_ranges:
                refreshed_by_day.update(await _load_point_sales_metrics_live(point, range_from, range_to))
            for day in missing_days:
                refreshed_by_day.setdefault(day, _empty_day_metrics())
                result[day] = refreshed_by_day[day]
            if db is not None:
                await _store_point_sales_metrics_cache(point, refreshed_by_day, db)
                await db.commit()

        ordered = {day: result.get(day, _empty_day_metrics()) for day in _daterange(date_from, date_to)}
        _sales_metrics_cache[cache_key] = (asyncio.get_running_loop().time(), ordered)
        return ordered
    except Exception:
        logger.exception(
            'Не удалось загрузить продажи/выручку МоегоСклада для точки %s за период %s..%s. Возвращаем пустые метрики.',
            point.name,
            date_from.isoformat(),
            date_to.isoformat(),
        )
        return {day: _empty_day_metrics() for day in _daterange(date_from, date_to)}



async def _get_active_shift_count(point: LocationPoint, shift_date: date, db: AsyncSession) -> int:
    count = await db.scalar(
        select(func.count())
        .select_from(WorkShift)
        .where(
            WorkShift.location_point_id == point.id,
            WorkShift.shift_date == shift_date,
            WorkShift.is_deleted.is_(False),
        )
    )
    return max(int(count or 0), 1)


async def auto_close_open_shifts_in_period(
    date_from: date,
    date_to: date,
    db: AsyncSession,
    location: str | None = None,
) -> dict[str, Any]:
    query = (
        select(WorkShift)
        .where(
            WorkShift.is_deleted.is_(False),
            WorkShift.status != 'closed',
            WorkShift.shift_date >= date_from,
            WorkShift.shift_date <= date_to,
        )
        .order_by(WorkShift.shift_date.asc(), WorkShift.id.asc())
    )
    normalized_location: str | None = None
    if location:
        normalized_location = _normalize_location(location)
        point = await db.scalar(select(LocationPoint).where(LocationPoint.name == normalized_location).limit(1))
        if point is None:
            return {
                'date_from': date_from.isoformat(),
                'date_to': date_to.isoformat(),
                'location': normalized_location,
                'matched': 0,
                'closed': 0,
                'details': [],
                'warning': 'location_not_found',
            }
        query = query.where(WorkShift.location_point_id == point.id)

    shifts = (await db.scalars(query)).all()
    details: list[dict[str, Any]] = []
    closed_count = 0
    for shift in shifts:
        result = await close_shift(shift.id, db, actor_user=None, auto=True)
        serialized = result.get('shift') or {}
        details.append({
            'shift_id': shift.id,
            'shift_date': shift.shift_date.isoformat(),
            'location_point_id': shift.location_point_id,
            'employee_user_id': shift.employee_user_id,
            'status': serialized.get('status') or shift.status,
            'message': result.get('message'),
        })
        closed_count += 1

    return {
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'location': normalized_location,
        'matched': len(shifts),
        'closed': closed_count,
        'details': details,
    }


async def _ensure_shift_snapshots_for_point(point: LocationPoint, db: AsyncSession) -> None:
    today = get_payroll_operational_today()
    due_shifts = (
        await db.scalars(
            select(WorkShift)
            .where(
                WorkShift.location_point_id == point.id,
                WorkShift.is_deleted.is_(False),
                WorkShift.status != 'closed',
                WorkShift.shift_date < today,
            )
            .order_by(WorkShift.shift_date.asc(), WorkShift.id.asc())
        )
    ).all()
    for shift in due_shifts:
        await close_shift(shift.id, db, actor_user=None, auto=True)


async def _build_computed_shift(shift: WorkShift, db: AsyncSession, *, force_refresh_metrics: bool = False) -> ShiftComputedPayroll:
    point = await db.get(LocationPoint, shift.location_point_id)
    employee = await db.get(User, shift.employee_user_id)
    if not point or not employee:
        raise HTTPException(status_code=404, detail='Не удалось собрать данные смены.')

    snapshot = await db.scalar(select(ShiftPayrollSnapshot).where(ShiftPayrollSnapshot.shift_id == shift.id).limit(1))
    if snapshot:
        category_rows = (
            await db.scalars(
                select(ShiftPayrollCategorySnapshot)
                .where(ShiftPayrollCategorySnapshot.snapshot_id == snapshot.id)
                .order_by(ShiftPayrollCategorySnapshot.category_name.asc())
            )
        ).all()
        if _has_legacy_uncategorized_adjustment(category_rows):
            logger.info(
                'Снимок смены %s за %s содержит legacy-коррекцию в «Без категории», пересчитываем зарплату по актуальным дневным метрикам.',
                shift.id,
                shift.shift_date.isoformat(),
            )
        else:
            settings = await db.get(PayrollSettingsVersion, snapshot.settings_version_id) if snapshot.settings_version_id else await _get_settings_for_date(point, shift.shift_date, db)
            bonus_category_ids = _load_bonus_category_ids(settings)
            rate_map = await _get_settings_rates(settings.id, db)
            rate_name_map = _build_category_rate_name_map(rate_map)
            snapshot_categories: list[dict[str, Any]] = []
            for row in category_rows:
                category_name = str(row.category_name or '').strip() or DEFAULT_CATEGORY_NAME
                category_id = str(row.category_id or '__other__')
                net_sales_amount = _sanitize_uncategorized_net(category_name, round(float(row.net_sales_amount or 0), 2))
                rate_info = _get_rate_info_for_category(category_id, category_name, rate_map, rate_name_map)
                rate_percent, _used_other_rate = _resolve_category_rate_percent(rate_info, settings.other_rate_percent)
                is_uncategorized = _is_uncategorized_category(category_name, category_id)
                is_return_adjustment = _is_uncategorized_return_adjustment_values(row.sales_amount, row.return_amount, net_sales_amount)
                effective_rate_percent = rate_percent if not is_uncategorized else 0.0
                calculation_base_amount = _resolve_calculation_sales_base(row.sales_amount, net_sales_amount)
                earning_amount = _calculate_category_earning_amount(calculation_base_amount, effective_rate_percent)
                snapshot_categories.append(_sanitize_uncategorized_row({
                    'category_id': category_id,
                    'category_name': category_name,
                    'rate_percent': round(effective_rate_percent, 2),
                    'sales_amount': round(float(row.sales_amount or 0), 2),
                    'return_amount': round(float(row.return_amount or 0), 2),
                    'net_sales_amount': net_sales_amount,
                    'cost_amount': round(float(row.cost_amount or 0), 2),
                    'earning_amount': earning_amount,
                    'is_other_category': is_uncategorized,
                }))
            snapshot_categories = await _backfill_category_costs_from_day_metrics(
                point,
                shift.shift_date,
                snapshot_categories,
                round(float(snapshot.cost_amount or 0), 2),
                float(snapshot.share_ratio or 0),
                db,
            )
            snapshot_category_earnings_total = round(sum(float(item.get('earning_amount') or 0) for item in snapshot_categories), 2)
            snapshot_bonus_base = round(max(float(snapshot.non_tobacco_net_sales_for_bonus or 0), 0.0), 2)
            if snapshot_categories:
                if bonus_category_ids:
                    snapshot_bonus_base = round(sum(
                        _resolve_category_calculation_base(item)
                        for item in snapshot_categories
                        if item.get('category_id') in bonus_category_ids
                    ), 2)
                else:
                    snapshot_bonus_base = round(sum(
                        max(_resolve_category_calculation_base(item), 0.0) if str(item.get('category_name') or '').strip() == DEFAULT_CATEGORY_NAME
                        else _resolve_category_calculation_base(item)
                        for item in snapshot_categories
                        if not _is_tobacco_category(item.get('category_name'))
                    ), 2)
            snapshot_bonus = round(float(snapshot.bonus_amount or 0), 2)
            snapshot_exit = round(float(snapshot.exit_amount or 0), 2)
            snapshot_cost_amount = round(float(snapshot.cost_amount or 0), 2)
            patched_snapshot_cost_amount = round(sum(float(item.get('cost_amount') or 0) for item in snapshot_categories), 2)
            if abs(snapshot_cost_amount) <= 0.009 and abs(patched_snapshot_cost_amount) > 0.009:
                snapshot_cost_amount = patched_snapshot_cost_amount
            snapshot_gross_profit_amount = round(float(snapshot.gross_profit_amount or 0), 2)
            expected_snapshot_gross_profit_amount = _calculate_gross_profit_from_revenue(snapshot.gross_sales_amount, snapshot_cost_amount)
            if abs(snapshot_gross_profit_amount) <= 0.009 and abs(expected_snapshot_gross_profit_amount) > 0.009:
                snapshot_gross_profit_amount = expected_snapshot_gross_profit_amount
            snapshot_gross_salary = round(snapshot_exit + snapshot_bonus + snapshot_category_earnings_total, 2)
            return ShiftComputedPayroll(
                shift=shift,
                location_point=point,
                employee=employee,
                settings=settings,
                split_count=snapshot.split_count,
                share_ratio=float(snapshot.share_ratio or 0),
                categories=snapshot_categories,
                exit_amount=snapshot_exit,
                bonus_threshold=round(float(snapshot.bonus_threshold or 0), 2),
                bonus_amount=snapshot_bonus,
                bonus_base_sales_amount=snapshot_bonus_base,
                bonus_category_ids=bonus_category_ids,
                other_rate_percent=round(float(snapshot.other_rate_percent or 0), 2),
                gross_sales_amount=round(float(snapshot.gross_sales_amount or 0), 2),
                return_amount=round(float(snapshot.return_amount or 0), 2),
                net_sales_amount=round(float(snapshot.net_sales_amount or 0), 2),
                cost_amount=snapshot_cost_amount,
                gross_profit_amount=snapshot_gross_profit_amount,
                non_tobacco_net_sales_for_bonus=round(float(snapshot.non_tobacco_net_sales_for_bonus or 0), 2),
                category_earnings_total=snapshot_category_earnings_total,
                gross_salary_amount=snapshot_gross_salary,
                snapshot_id=snapshot.id,
                is_closed=True,
                closed_at=_datetime_to_str(snapshot.closed_at),
                is_auto_closed=bool(snapshot.is_auto_closed),
            )

    settings = await _get_settings_for_date(point, shift.shift_date, db)
    bonus_category_ids = _load_bonus_category_ids(settings)
    rate_map = await _get_settings_rates(settings.id, db)
    rate_name_map = _build_category_rate_name_map(rate_map)
    day_metrics_by_date = await _load_point_sales_metrics(point, shift.shift_date, shift.shift_date, db, force_refresh=force_refresh_metrics)
    day_metrics = day_metrics_by_date.get(shift.shift_date, {
        'categories': [],
        'gross_sales_amount': 0.0,
        'return_amount': 0.0,
        'net_sales_amount': 0.0,
        'cost_amount': 0.0,
        'gross_profit_amount': 0.0,
        'non_tobacco_net_sales_for_bonus': 0.0,
    })
    split_count = await _get_active_shift_count(point, shift.shift_date, db)
    share_ratio = round(1.0 / split_count, 4)
    categories: list[dict[str, Any]] = []
    category_earnings_total = 0.0
    bonus_base_sales_amount = 0.0

    for row in day_metrics['categories']:
        sales_amount = round(float(row['sales_amount']) * share_ratio, 2)
        return_amount = round(float(row['return_amount']) * share_ratio, 2)
        net_sales_amount = _sanitize_uncategorized_net(row.get('category_name'), round(float(row['net_sales_amount']) * share_ratio, 2))
        category_id = row['category_id']
        rate_info = _get_rate_info_for_category(category_id, row.get('category_name'), rate_map, rate_name_map)
        rate_percent, _used_other_rate = _resolve_category_rate_percent(rate_info, settings.other_rate_percent)
        cost_amount = round(float(row.get('cost_amount') or 0) * share_ratio, 2)
        is_uncategorized = _is_uncategorized_category(row.get('category_name'), category_id)
        is_return_adjustment = _is_uncategorized_return_adjustment_values(sales_amount, return_amount, net_sales_amount)
        effective_rate_percent = rate_percent if not is_uncategorized else 0.0
        calculation_base_amount = _resolve_calculation_sales_base(sales_amount, net_sales_amount)
        earning_amount = _calculate_category_earning_amount(calculation_base_amount, effective_rate_percent)
        category_earnings_total += earning_amount
        categories.append(_sanitize_uncategorized_row({
            'category_id': category_id,
            'category_name': row['category_name'],
            'rate_percent': round(effective_rate_percent, 2),
            'sales_amount': sales_amount,
            'return_amount': return_amount,
            'net_sales_amount': net_sales_amount,
            'cost_amount': cost_amount,
            'earning_amount': earning_amount,
            'is_other_category': is_uncategorized,
        }))
        if bonus_category_ids:
            if category_id in bonus_category_ids:
                bonus_base_sales_amount += calculation_base_amount

    if bonus_category_ids:
        bonus_base_sales_amount = round(max(bonus_base_sales_amount, 0.0), 2)
    else:
        bonus_base_sales_amount = round(sum(
            max(_resolve_category_calculation_base(category), 0.0) if str(category.get('category_name') or '').strip() == DEFAULT_CATEGORY_NAME
            else _resolve_category_calculation_base(category)
            for category in categories
            if not _is_tobacco_category(category.get('category_name'))
        ), 2)

    non_tobacco_net = round(sum(
        max(_resolve_category_calculation_base(category), 0.0) if str(category.get('category_name') or '').strip() == DEFAULT_CATEGORY_NAME
        else _resolve_category_calculation_base(category)
        for category in categories
        if not _is_tobacco_category(category.get('category_name'))
    ), 2)
    gross_sales_amount = round(float(day_metrics['gross_sales_amount']) * share_ratio, 2)
    return_amount = round(float(day_metrics['return_amount']) * share_ratio, 2)
    net_sales_amount = round(float(day_metrics['net_sales_amount']) * share_ratio, 2)
    cost_amount = round(float(day_metrics['cost_amount']) * share_ratio, 2)
    gross_profit_amount = _calculate_gross_profit_from_revenue(gross_sales_amount, cost_amount)
    bonus = float(settings.bonus_amount or 0) if bonus_base_sales_amount >= float(settings.bonus_threshold or 0) else 0.0
    gross_salary_amount = round(float(settings.exit_amount or 0) + bonus + category_earnings_total, 2)

    return ShiftComputedPayroll(
        shift=shift,
        location_point=point,
        employee=employee,
        settings=settings,
        split_count=split_count,
        share_ratio=share_ratio,
        categories=categories,
        exit_amount=round(float(settings.exit_amount or 0), 2),
        bonus_threshold=round(float(settings.bonus_threshold or 0), 2),
        bonus_amount=round(bonus, 2),
        bonus_base_sales_amount=bonus_base_sales_amount,
        bonus_category_ids=bonus_category_ids,
        other_rate_percent=round(float(settings.other_rate_percent or 0), 2),
        gross_sales_amount=gross_sales_amount,
        return_amount=return_amount,
        net_sales_amount=net_sales_amount,
        cost_amount=cost_amount,
        gross_profit_amount=gross_profit_amount,
        non_tobacco_net_sales_for_bonus=non_tobacco_net,
        category_earnings_total=round(category_earnings_total, 2),
        gross_salary_amount=gross_salary_amount,
        is_closed=shift.status == 'closed',
        closed_at=_datetime_to_str(shift.closed_at),
        is_auto_closed=bool(snapshot.is_auto_closed) if snapshot else False,
    )


async def close_shift(shift_id: int, db: AsyncSession, actor_user: User | None, auto: bool = False) -> dict[str, Any]:
    shift = await db.get(WorkShift, shift_id)
    if not shift or shift.is_deleted:
        raise HTTPException(status_code=404, detail='Смена не найдена.')
    point = await db.get(LocationPoint, shift.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка смены не найдена.')
    if actor_user is not None:
        if actor_user.role == 'employee' and actor_user.id != shift.employee_user_id:
            raise HTTPException(status_code=403, detail='Можно закрыть только свою смену.')
        if actor_user.role in {'admin', 'superadmin'}:
            await ensure_user_can_access_location(actor_user, point.name, db)
    if shift.status == 'closed':
        computed = await _build_computed_shift(shift, db)
        return {'success': True, 'message': 'Смена уже закрыта.', 'shift': _serialize_computed_shift(computed)}

    computed = await _build_computed_shift(shift, db)
    snapshot = ShiftPayrollSnapshot(
        shift_id=shift.id,
        location_point_id=point.id,
        employee_user_id=shift.employee_user_id,
        shift_date=shift.shift_date,
        settings_version_id=computed.settings.id,
        split_count=computed.split_count,
        share_ratio=computed.share_ratio,
        exit_amount=computed.exit_amount,
        bonus_threshold=computed.bonus_threshold,
        bonus_amount=computed.bonus_amount,
        other_rate_percent=computed.other_rate_percent,
        non_tobacco_net_sales_for_bonus=computed.bonus_base_sales_amount,
        gross_sales_amount=computed.gross_sales_amount,
        return_amount=computed.return_amount,
        net_sales_amount=computed.net_sales_amount,
        cost_amount=computed.cost_amount,
        gross_profit_amount=computed.gross_profit_amount,
        category_earnings_total=computed.category_earnings_total,
        employee_expense_amount=0.0,
        gross_salary_amount=computed.gross_salary_amount,
        net_salary_amount=computed.gross_salary_amount,
        is_auto_closed=auto,
        closed_at=datetime.utcnow(),
    )
    db.add(snapshot)
    await db.flush()
    for row in computed.categories:
        db.add(ShiftPayrollCategorySnapshot(
            snapshot_id=snapshot.id,
            category_id=row['category_id'],
            category_name=row['category_name'],
            rate_percent=row['rate_percent'],
            sales_amount=row['sales_amount'],
            return_amount=row['return_amount'],
            net_sales_amount=row['net_sales_amount'],
            cost_amount=row.get('cost_amount', 0.0),
            earning_amount=row['earning_amount'],
            is_other_category=row['is_other_category'],
        ))
    shift.status = 'closed'
    shift.closed_at = datetime.utcnow()
    shift.closed_by_user_id = None if auto else (actor_user.id if actor_user else None)
    shift.updated_at = datetime.utcnow()
    await _log_payroll_action(
        db,
        actor_user_id=None if auto else (actor_user.id if actor_user else None),
        location_point_id=point.id,
        entity_type='work_shift',
        entity_id=str(shift.id),
        action_type='auto_close' if auto else 'close',
        details={
            'shift_date': shift.shift_date.isoformat(),
            'employee_user_id': shift.employee_user_id,
            'gross_salary_amount': computed.gross_salary_amount,
            'split_count': computed.split_count,
        },
    )
    await db.commit()
    computed = await _build_computed_shift(shift, db)
    return {'success': True, 'message': 'Смена закрыта.' if not auto else 'Смена автоматически закрыта.', 'shift': _serialize_computed_shift(computed)}


async def upsert_work_shift(payload: WorkShiftUpsertRequest, db: AsyncSession, current_user: User) -> dict[str, Any]:
    if current_user.role not in {'admin', 'superadmin'}:
        raise HTTPException(status_code=403, detail='Назначать смены может только управляющий.')
    await ensure_user_can_access_location(current_user, payload.location, db)
    point = await _get_location_point_by_name(payload.location, db)
    employee = await db.get(User, payload.employee_user_id)
    if not employee or employee.role != 'employee' or employee.location != point.name:
        raise HTTPException(status_code=400, detail='Сотрудник не привязан к выбранной точке.')
    shift = await db.scalar(
        select(WorkShift)
        .where(
            WorkShift.location_point_id == point.id,
            WorkShift.shift_date == payload.shift_date,
            WorkShift.employee_user_id == employee.id,
        )
        .limit(1)
    )
    if shift:
        shift.is_deleted = False
        if shift.status == 'deleted':
            shift.status = 'planned'
        shift.deleted_at = None
        shift.updated_at = datetime.utcnow()
        action = 'restore'
    else:
        shift = WorkShift(
            location_point_id=point.id,
            shift_date=payload.shift_date,
            employee_user_id=employee.id,
            status='planned',
            created_by_user_id=current_user.id,
            updated_at=datetime.utcnow(),
        )
        db.add(shift)
        await db.flush()
        action = 'create'
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='work_shift',
        entity_id=str(shift.id),
        action_type=action,
        details={'shift_date': payload.shift_date.isoformat(), 'employee_user_id': employee.id},
    )
    await db.commit()
    if payload.shift_date < get_payroll_operational_today() and shift.status != 'closed':
        await close_shift(shift.id, db, actor_user=current_user, auto=True)
    return {'success': True, 'message': 'Смена сохранена.'}


async def delete_work_shift(shift_id: int, db: AsyncSession, current_user: User) -> dict[str, Any]:
    shift = await db.get(WorkShift, shift_id)
    if not shift:
        raise HTTPException(status_code=404, detail='Смена не найдена.')
    point = await db.get(LocationPoint, shift.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка смены не найдена.')
    await ensure_user_can_access_location(current_user, point.name, db)
    shift.is_deleted = True
    shift.deleted_at = datetime.utcnow()
    shift.updated_at = datetime.utcnow()
    if shift.status != 'closed':
        shift.status = 'deleted'
    await _log_payroll_action(
        db,
        actor_user_id=current_user.id,
        location_point_id=point.id,
        entity_type='work_shift',
        entity_id=str(shift.id),
        action_type='delete',
        details={'shift_date': shift.shift_date.isoformat(), 'employee_user_id': shift.employee_user_id, 'closed': shift.status == 'closed'},
    )
    await db.commit()
    return {'success': True, 'message': 'Смена скрыта из активного календаря. История по закрытым сменам сохранена.'}



def _serialize_computed_shift(computed: ShiftComputedPayroll) -> dict[str, Any]:
    serialized_categories: list[dict[str, Any]] = []
    for row in computed.categories:
        cloned = dict(row)
        cloned['category_name'] = _get_payroll_display_category_name(cloned)
        serialized_categories.append(cloned)
    return {
        'id': computed.shift.id,
        'shift_date': computed.shift.shift_date.isoformat(),
        'employee_user_id': computed.employee.id,
        'employee_name': computed.employee.full_name,
        'location': computed.location_point.name,
        'status': computed.shift.status,
        'split_count': computed.split_count,
        'share_ratio': computed.share_ratio,
        'is_closed': computed.is_closed,
        'is_auto_closed': computed.is_auto_closed,
        'closed_at': computed.closed_at,
        'gross_sales_amount': computed.gross_sales_amount,
        'return_amount': computed.return_amount,
        'net_sales_amount': computed.net_sales_amount,
        'cost_amount': computed.cost_amount,
        'gross_profit_amount': computed.gross_profit_amount,
        'non_tobacco_net_sales_for_bonus': computed.non_tobacco_net_sales_for_bonus,
        'bonus_base_sales_amount': computed.bonus_base_sales_amount,
        'bonus_category_ids': computed.bonus_category_ids,
        'exit_amount': computed.exit_amount,
        'bonus_threshold': computed.bonus_threshold,
        'bonus_amount': computed.bonus_amount,
        'category_earnings_total': computed.category_earnings_total,
        'gross_salary_amount': computed.gross_salary_amount,
        'categories': serialized_categories,
    }


async def _serialize_shift_lightweight(
    shift: WorkShift,
    db: AsyncSession,
    *,
    employee_name: str | None = None,
) -> dict[str, Any]:
    point = await db.get(LocationPoint, shift.location_point_id)
    if not point:
        raise HTTPException(status_code=404, detail='Точка смены не найдена.')
    if employee_name is None:
        employee = await db.get(User, shift.employee_user_id)
        employee_name = employee.full_name if employee else 'Сотрудник'

    snapshot = await db.scalar(
        select(ShiftPayrollSnapshot)
        .where(ShiftPayrollSnapshot.shift_id == shift.id)
        .limit(1)
    )
    is_closed = bool(snapshot) or shift.status == 'closed'
    if snapshot:
        category_rows = (
            await db.scalars(
                select(ShiftPayrollCategorySnapshot)
                .where(ShiftPayrollCategorySnapshot.snapshot_id == snapshot.id)
                .order_by(ShiftPayrollCategorySnapshot.category_name.asc())
            )
        ).all()
        if _has_legacy_uncategorized_adjustment(category_rows):
            computed = await _build_computed_shift(shift, db)
            return {
                'id': shift.id,
                'shift_date': shift.shift_date.isoformat(),
                'employee_user_id': shift.employee_user_id,
                'employee_name': employee_name,
                'location': point.name,
                'status': shift.status,
                'is_closed': computed.is_closed,
                'closed_at': computed.closed_at,
                'gross_sales_amount': computed.gross_sales_amount,
                'return_amount': computed.return_amount,
                'net_sales_amount': computed.net_sales_amount,
                'exit_amount': computed.exit_amount,
                'bonus_amount': computed.bonus_amount,
                'category_earnings_total': computed.category_earnings_total,
                'gross_salary_amount': computed.gross_salary_amount,
            }
        exit_amount = round(float(snapshot.exit_amount or 0), 2)
        bonus_amount = round(float(snapshot.bonus_amount or 0), 2)
        category_earnings_total = round(float(snapshot.category_earnings_total or 0), 2)
        gross_sales_amount = round(float(snapshot.gross_sales_amount or 0), 2)
        return_amount = round(float(snapshot.return_amount or 0), 2)
        net_sales_amount = round(float(snapshot.net_sales_amount or 0), 2)
        gross_salary_amount = round(float(snapshot.gross_salary_amount or 0), 2)
        closed_at = _datetime_to_str(snapshot.closed_at)
    else:
        settings = await _get_settings_for_date(point, shift.shift_date, db)
        exit_amount = round(float(settings.exit_amount or 0), 2)
        bonus_amount = 0.0
        category_earnings_total = 0.0
        gross_sales_amount = 0.0
        return_amount = 0.0
        net_sales_amount = 0.0
        gross_salary_amount = round(exit_amount + bonus_amount + category_earnings_total, 2)
        closed_at = _datetime_to_str(shift.closed_at)

    return {
        'id': shift.id,
        'shift_date': shift.shift_date.isoformat(),
        'employee_user_id': shift.employee_user_id,
        'employee_name': employee_name,
        'location': point.name,
        'status': shift.status,
        'is_closed': is_closed,
        'closed_at': closed_at,
        'gross_sales_amount': gross_sales_amount,
        'return_amount': return_amount,
        'net_sales_amount': net_sales_amount,
        'exit_amount': exit_amount,
        'bonus_amount': bonus_amount,
        'category_earnings_total': category_earnings_total,
        'gross_salary_amount': gross_salary_amount,
    }


async def list_work_shifts(location: str, date_from: date, date_to: date, db: AsyncSession, current_user: User, employee_user_id: int | None = None) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    query = select(WorkShift).where(
        WorkShift.location_point_id == point.id,
        WorkShift.shift_date >= date_from,
        WorkShift.shift_date <= date_to,
        WorkShift.is_deleted.is_(False),
    )
    if employee_user_id:
        query = query.where(WorkShift.employee_user_id == employee_user_id)
    rows = (await db.scalars(query.order_by(WorkShift.shift_date.asc(), WorkShift.id.asc()))).all()
    if current_user.role == 'employee':
        rows = [row for row in rows if row.employee_user_id == current_user.id]

    employee_ids = {row.employee_user_id for row in rows}
    employee_names: dict[int, str] = {}
    if employee_ids:
        employees = (
            await db.scalars(select(User).where(User.id.in_(employee_ids)))
        ).all()
        employee_names = {item.id: item.full_name for item in employees}

    by_day: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for shift in rows:
        serialized = await _serialize_shift_lightweight(
            shift,
            db,
            employee_name=employee_names.get(shift.employee_user_id),
        )
        by_day[shift.shift_date.isoformat()].append(serialized)
    return {
        'location': point.name,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'days': [{
            'date': day,
            'shifts': sorted(items, key=lambda item: item['employee_name'].lower()),
        } for day, items in sorted(by_day.items())],
    }


async def list_work_shift_day_summary(location: str, date_from: date, date_to: date, db: AsyncSession, current_user: User, employee_user_id: int | None = None) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    query = select(WorkShift).where(
        WorkShift.location_point_id == point.id,
        WorkShift.shift_date >= date_from,
        WorkShift.shift_date <= date_to,
        WorkShift.is_deleted.is_(False),
    )
    if employee_user_id:
        query = query.where(WorkShift.employee_user_id == employee_user_id)
    rows = (await db.scalars(query.order_by(WorkShift.shift_date.desc(), WorkShift.id.desc()))).all()
    if current_user.role == 'employee':
        rows = [row for row in rows if row.employee_user_id == current_user.id]

    employee_ids = {row.employee_user_id for row in rows}
    employee_names: dict[int, str] = {}
    if employee_ids:
        employees = (
            await db.scalars(select(User).where(User.id.in_(employee_ids)))
        ).all()
        employee_names = {item.id: item.full_name for item in employees}

    days: list[dict[str, Any]] = []
    for shift in rows:
        days.append(
            await _serialize_shift_lightweight(
                shift,
                db,
                employee_name=employee_names.get(shift.employee_user_id),
            )
        )
    return {
        'location': point.name,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'days': sorted(days, key=lambda item: (item['shift_date'], item['employee_name'].lower()), reverse=False),
    }


async def _months_between(date_from: date, date_to: date) -> list[date]:
    current = _month_start(date_from)
    result: list[date] = []
    while current <= date_to:
        result.append(current)
        current = date(current.year + (1 if current.month == 12 else 0), 1 if current.month == 12 else current.month + 1, 1)
    return result


async def _collect_employee_expenses(point: LocationPoint, employee_user_id: int, date_from: date, date_to: date, db: AsyncSession) -> float:
    months = await _months_between(date_from, date_to)
    templates = {
        row.id: row
        for row in (
            await db.scalars(select(ExpenseTemplate).where(ExpenseTemplate.location_point_id == point.id))
        ).all()
    }
    total = 0.0
    for month in months:
        entries = await _ensure_month_expense_entries(point, month, db)
        total += sum(
            _expense_entry_amount_for_period(
                entry,
                date_from=date_from,
                date_to=date_to,
                template=templates.get(entry.template_id) if entry.template_id is not None else None,
            )
            for entry in entries
            if entry.assigned_employee_user_id == employee_user_id and entry.apply_to_employee_salary
        )
    return round(total, 2)


async def _collect_period_company_expenses(point: LocationPoint, date_from: date, date_to: date, db: AsyncSession) -> float:
    months = await _months_between(date_from, date_to)
    templates = {
        row.id: row
        for row in (
            await db.scalars(select(ExpenseTemplate).where(ExpenseTemplate.location_point_id == point.id))
        ).all()
    }
    total = 0.0
    for month in months:
        entries = await _ensure_month_expense_entries(point, month, db)
        total += sum(
            _expense_entry_amount_for_period(
                entry,
                date_from=date_from,
                date_to=date_to,
                template=templates.get(entry.template_id) if entry.template_id is not None else None,
            )
            for entry in entries
        )
    return round(total, 2)


async def get_employee_payroll_summary(location: str, date_from: date, date_to: date, db: AsyncSession, current_user: User, employee_user_id: int | None = None) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    await _ensure_shift_snapshots_for_point(point, db)

    target_employee_id = employee_user_id if employee_user_id is not None else (current_user.id if current_user.role == 'employee' else None)
    if current_user.role == 'employee' and target_employee_id != current_user.id:
        raise HTTPException(status_code=403, detail='Сотрудник может смотреть только свою зарплату.')

    employee: User | None = None
    if target_employee_id is not None:
        employee = await db.get(User, target_employee_id)
        if not employee or employee.role != 'employee':
            raise HTTPException(status_code=404, detail='Сотрудник не найден.')

    today = get_payroll_operational_today()
    effective_date_to = min(date_to, today)
    query = select(WorkShift).where(
        WorkShift.location_point_id == point.id,
        WorkShift.shift_date >= date_from,
        WorkShift.shift_date <= effective_date_to,
        WorkShift.is_deleted.is_(False),
    )
    if target_employee_id is not None:
        query = query.where(WorkShift.employee_user_id == target_employee_id)
    shifts = (await db.scalars(query.order_by(WorkShift.shift_date.asc(), WorkShift.id.asc()))).all()
    if current_user.role == 'employee':
        shifts = [shift for shift in shifts if shift.employee_user_id == current_user.id]

    days: list[dict[str, Any]] = []
    totals = {
        'gross_sales_amount': 0.0,
        'return_amount': 0.0,
        'net_sales_amount': 0.0,
        'cost_amount': 0.0,
        'gross_profit_amount': 0.0,
        'exit_amount': 0.0,
        'bonus_amount': 0.0,
        'category_earnings_total': 0.0,
        'gross_salary_amount': 0.0,
    }
    category_totals: dict[str, dict[str, Any]] = {}
    employee_ids_in_result: set[int] = set()
    for shift in shifts:
        computed = await _build_computed_shift(shift, db)
        serialized = _serialize_computed_shift(computed)
        days.append(serialized)
        employee_ids_in_result.add(int(serialized['employee_user_id']))
        for key in totals:
            totals[key] += float(serialized.get(key) or 0)
        for category in serialized['categories']:
            bucket = category_totals.setdefault(category['category_id'], {
                'category_id': category['category_id'],
                'category_name': category['category_name'],
                'rate_percent': category['rate_percent'],
                'sales_amount': 0.0,
                'return_amount': 0.0,
                'net_sales_amount': 0.0,
                'cost_amount': 0.0,
                'earning_amount': 0.0,
            })
            bucket['sales_amount'] += float(category['sales_amount'] or 0)
            bucket['return_amount'] += float(category['return_amount'] or 0)
            bucket['net_sales_amount'] += float(category['net_sales_amount'] or 0)
            bucket['cost_amount'] += float(category.get('cost_amount') or 0)
            bucket['earning_amount'] += float(category['earning_amount'] or 0)

    realized_expense_date_to = effective_date_to if effective_date_to >= date_from else None
    if target_employee_id is not None:
        employee_expenses_total = (
            await _collect_employee_expenses(point, target_employee_id, date_from, realized_expense_date_to, db)
            if realized_expense_date_to is not None
            else 0.0
        )
        employee_label = employee.full_name if employee else 'Сотрудник'
        employee_count = 1 if employee else 0
        employee_user_id_value = employee.id if employee else None
    else:
        employee_expenses_total = 0.0
        if realized_expense_date_to is not None:
            for employee_id in employee_ids_in_result:
                employee_expenses_total += await _collect_employee_expenses(point, employee_id, date_from, realized_expense_date_to, db)
        employee_expenses_total = round(employee_expenses_total, 2)
        employee_label = 'Все сотрудники'
        employee_count = len(employee_ids_in_result)
        employee_user_id_value = None

    summary = {
        'location': point.name,
        'employee_user_id': employee_user_id_value,
        'employee_name': employee_label,
        'employee_scope': 'single' if target_employee_id is not None else 'all',
        'employee_count': employee_count,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'days': days,
        'categories': sorted(
            [
                {
                    **row,
                    'category_name': _get_payroll_display_category_name(row),
                    'sales_amount': round(float(row['sales_amount']), 2),
                    'return_amount': round(float(row['return_amount']), 2),
                    'net_sales_amount': round(float(row['net_sales_amount']), 2),
                    'cost_amount': round(float(row.get('cost_amount') or 0), 2),
                    'earning_amount': round(float(row['earning_amount']), 2),
                }
                for row in category_totals.values()
            ],
            key=lambda item: item['category_name'].lower(),
        ),
        'totals': {key: round(value, 2) for key, value in totals.items()},
        'employee_expenses_total': employee_expenses_total,
        'net_payout_amount': round(totals['gross_salary_amount'] - employee_expenses_total, 2),
    }
    return summary



async def _sum_shift_salaries(point: LocationPoint, date_from: date, date_to: date, db: AsyncSession) -> float:
    if date_to < date_from:
        return 0.0
    shifts = (
        await db.scalars(
            select(WorkShift)
            .where(
                WorkShift.location_point_id == point.id,
                WorkShift.shift_date >= date_from,
                WorkShift.shift_date <= date_to,
                WorkShift.is_deleted.is_(False),
            )
            .order_by(WorkShift.shift_date.asc(), WorkShift.id.asc())
        )
    ).all()
    total = 0.0
    for shift in shifts:
        computed = await _build_computed_shift(shift, db)
        total += computed.gross_salary_amount
    return round(total, 2)


async def _calculate_manager_salary_proration(
    point: LocationPoint,
    date_from: date,
    date_to: date,
    db: AsyncSession,
) -> tuple[float, float, list[dict[str, Any]]]:
    today = get_payroll_operational_today()
    details: list[dict[str, Any]] = []
    total_salary = 0.0
    representative_rate = 0.0

    for month_start in await _months_between(date_from, date_to):
        month_end = _month_end(month_start)
        realized_end = min(month_end, today)
        if realized_end < month_start:
            continue
        selected_from = max(date_from, month_start)
        selected_to = min(date_to, realized_end)
        if selected_to < selected_from:
            continue

        month_metrics = await _load_point_sales_metrics(point, month_start, realized_end, db)
        month_net_sales = round(sum(_resolve_calculation_sales_base(day.get('gross_sales_amount'), day.get('net_sales_amount')) for day in month_metrics.values()), 2)
        month_cost_amount = round(sum(float(day.get('cost_amount') or 0.0) for day in month_metrics.values()), 2)
        month_employee_salary_total = await _sum_shift_salaries(point, month_start, realized_end, db)
        month_expenses_total = await _collect_period_company_expenses(point, month_start, realized_end, db)
        month_operating_profit = round(month_net_sales - month_cost_amount - month_employee_salary_total - month_expenses_total, 2)

        month_settings = await _get_settings_for_date(point, selected_to, db)
        month_brackets = _load_manager_salary_brackets(month_settings)
        month_rate_percent = _manager_rate_for_profit(month_operating_profit, month_brackets)
        month_salary_amount = round(max(month_operating_profit, 0.0) * (month_rate_percent / 100.0), 2)

        realized_days_count = (realized_end - month_start).days + 1
        selected_days_count = (selected_to - selected_from).days + 1
        prorated_amount = round(month_salary_amount / realized_days_count * selected_days_count, 2) if realized_days_count > 0 else 0.0
        total_salary += prorated_amount
        representative_rate = max(representative_rate, month_rate_percent)
        details.append({
            'month_start': month_start.isoformat(),
            'month_end': month_end.isoformat(),
            'realized_end': realized_end.isoformat(),
            'selected_from': selected_from.isoformat(),
            'selected_to': selected_to.isoformat(),
            'realized_days_count': realized_days_count,
            'selected_days_count': selected_days_count,
            'operating_profit': month_operating_profit,
            'manager_rate_percent': month_rate_percent,
            'manager_salary_full_amount': month_salary_amount,
            'manager_salary_prorated_amount': prorated_amount,
        })

    return round(total_salary, 2), round(representative_rate, 2), details


async def get_manager_payroll_summary(location: str, date_from: date, date_to: date, db: AsyncSession, current_user: User) -> dict[str, Any]:
    await ensure_user_can_access_location(current_user, location, db)
    point = await _get_location_point_by_name(location, db)
    await _ensure_shift_snapshots_for_point(point, db)
    today = get_payroll_operational_today()
    effective_date_to = min(date_to, today)
    day_metrics = await _load_point_sales_metrics(point, date_from, effective_date_to, db) if effective_date_to >= date_from else {}
    gross_sales_amount = sum(day['gross_sales_amount'] for day in day_metrics.values())
    return_amount = sum(day['return_amount'] for day in day_metrics.values())
    net_sales_amount = sum(day['net_sales_amount'] for day in day_metrics.values())
    cost_amount = sum(day['cost_amount'] for day in day_metrics.values())

    employee_salary_total = await _sum_shift_salaries(point, date_from, effective_date_to, db) if effective_date_to >= date_from else 0.0

    settings_effective_date = effective_date_to if effective_date_to >= date_from else date_from
    current_settings = await _get_settings_for_date(point, settings_effective_date, db)
    manager_salary_brackets = _load_manager_salary_brackets(current_settings)
    settings_cache: dict[int, tuple[PayrollSettingsVersion, dict[str, dict[str, Any]], dict[str, dict[str, Any]]]] = {}
    category_totals: dict[str, dict[str, Any]] = {}
    for metric_day, metrics in day_metrics.items():
        day_settings = await _get_settings_for_date(point, metric_day, db)
        cache_item = settings_cache.get(day_settings.id)
        if cache_item is None:
            day_rate_map = await _get_settings_rates(day_settings.id, db)
            day_rate_name_map = _build_category_rate_name_map(day_rate_map)
            cache_item = (day_settings, day_rate_map, day_rate_name_map)
            settings_cache[day_settings.id] = cache_item
        _, day_rate_map, day_rate_name_map = cache_item
        for row in metrics.get('categories') or []:
            category_id = str(row.get('category_id') or '__other__')
            category_name = str(row.get('category_name') or DEFAULT_CATEGORY_NAME)
            rate_info = _get_rate_info_for_category(category_id, category_name, day_rate_map, day_rate_name_map)
            rate_percent, _used_other_rate = _resolve_category_rate_percent(rate_info, day_settings.other_rate_percent)
            net_amount = _sanitize_uncategorized_net(category_name, float(row.get('net_sales_amount') or 0))
            is_uncategorized = _is_uncategorized_category(category_name, category_id)
            is_return_adjustment = _is_uncategorized_return_adjustment_values(row.get('sales_amount'), row.get('return_amount'), net_amount)
            effective_rate_percent = rate_percent if not is_uncategorized else 0.0
            calculation_base_amount = _resolve_calculation_sales_base(row.get('sales_amount'), net_amount)
            earning_amount = _calculate_category_earning_amount(calculation_base_amount, effective_rate_percent)
            bucket = category_totals.setdefault(category_id, {
                'category_id': category_id,
                'category_name': category_name,
                'rate_percent': effective_rate_percent,
                'sales_amount': 0.0,
                'return_amount': 0.0,
                'net_sales_amount': 0.0,
                'cost_amount': 0.0,
                'earning_amount': 0.0,
                'is_other_category': is_uncategorized,
            })
            bucket['sales_amount'] += float(row.get('sales_amount') or 0)
            bucket['return_amount'] += float(row.get('return_amount') or 0)
            bucket['net_sales_amount'] += net_amount
            bucket['cost_amount'] += float(row.get('cost_amount') or 0)
            bucket['earning_amount'] += earning_amount
            # Для периода с несколькими версиями правил показываем последнюю ставку, действовавшую в этом периоде.
            bucket['rate_percent'] = effective_rate_percent
            bucket['is_other_category'] = is_uncategorized

    expenses_total = (
        await _collect_period_company_expenses(point, date_from, effective_date_to, db)
        if effective_date_to >= date_from
        else 0.0
    )
    operating_profit = round(gross_sales_amount - cost_amount - employee_salary_total - expenses_total, 2)
    manager_salary_amount, manager_rate_percent, manager_salary_proration = await _calculate_manager_salary_proration(point, date_from, date_to, db)
    net_profit_after_manager_salary = round(operating_profit - manager_salary_amount, 2)
    responsible_admin = await db.get(User, current_settings.responsible_admin_user_id) if current_settings.responsible_admin_user_id else None
    return {
        'location': point.name,
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'gross_sales_amount': round(gross_sales_amount, 2),
        'revenue_amount': round(gross_sales_amount, 2),
        'return_amount': round(return_amount, 2),
        'net_sales_amount': round(net_sales_amount, 2),
        'cost_amount': round(cost_amount, 2),
        'employee_salary_total': round(employee_salary_total, 2),
        'expenses_total': expenses_total,
        'operating_profit_before_manager_salary': operating_profit,
        'manager_rate_percent': manager_rate_percent,
        'manager_salary_amount': manager_salary_amount,
        'net_profit_after_manager_salary': net_profit_after_manager_salary,
        'profit_after_manager_salary': net_profit_after_manager_salary,
        'responsible_admin_user_id': responsible_admin.id if responsible_admin else None,
        'responsible_admin_name': responsible_admin.full_name if responsible_admin else None,
        'manager_salary_brackets': manager_salary_brackets,
        'manager_salary_proration': manager_salary_proration,
        'categories': sorted(
            [
                {
                    **row,
                    'category_name': _get_payroll_display_category_name(row),
                    'sales_amount': round(float(row['sales_amount']), 2),
                    'return_amount': round(float(row['return_amount']), 2),
                    'net_sales_amount': round(float(row['net_sales_amount']), 2),
                    'cost_amount': round(float(row.get('cost_amount') or 0), 2),
                    'earning_amount': round(float(row['earning_amount']), 2),
                }
                for row in category_totals.values()
            ],
            key=lambda item: item['category_name'].lower(),
        ),
    }


async def refresh_payroll_metrics_cache(date_from: date, date_to: date, db: AsyncSession, location: str | None = None, *, force_refresh: bool = False) -> dict[str, Any]:
    query = select(LocationPoint).order_by(LocationPoint.name.asc())
    if location:
        query = query.where(LocationPoint.name == _normalize_location(location))
    points = (await db.scalars(query)).all()

    results: list[dict[str, Any]] = []
    total_days = 0
    for point in points:
        token = _point_ms_token(point)
        if not _ms_client_enabled(token=token, location=point.name):
            results.append({
                'location': point.name,
                'skipped': True,
                'reason': 'moysklad_disabled',
            })
            continue
        metrics = await _load_point_sales_metrics(point, date_from, date_to, db, force_refresh=force_refresh)
        day_count = len(metrics)
        total_days += day_count
        gross_sales_amount = round(sum(float(day.get('gross_sales_amount') or 0) for day in metrics.values()), 2)
        return_amount = round(sum(float(day.get('return_amount') or 0) for day in metrics.values()), 2)
        cost_amount = round(sum(float(day.get('cost_amount') or 0) for day in metrics.values()), 2)
        results.append({
            'location': point.name,
            'days': day_count,
            'gross_sales_amount': gross_sales_amount,
            'return_amount': return_amount,
            'cost_amount': cost_amount,
            'refreshed': True,
        })

    return {
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'locations': results,
        'total_days': total_days,
    }


async def rebuild_closed_shift_snapshots(
    date_from: date,
    date_to: date,
    db: AsyncSession,
    location: str | None = None,
    *,
    force_refresh_metrics: bool = False,
    progress_callback: Any | None = None,
) -> dict[str, Any]:
    query = (
        select(WorkShift)
        .where(
            WorkShift.is_deleted.is_(False),
            WorkShift.status == 'closed',
            WorkShift.shift_date >= date_from,
            WorkShift.shift_date <= date_to,
        )
        .order_by(WorkShift.shift_date.asc(), WorkShift.id.asc())
    )
    if location:
        normalized = _normalize_location(location)
        point = await db.scalar(select(LocationPoint).where(LocationPoint.name == normalized).limit(1))
        if point is None:
            return {
                'date_from': date_from.isoformat(),
                'date_to': date_to.isoformat(),
                'location': normalized,
                'updated': 0,
                'processed': 0,
                'details': [],
            }
        query = query.where(WorkShift.location_point_id == point.id)

    shifts = (await db.scalars(query)).all()
    details: list[dict[str, Any]] = []
    updated = 0

    for shift in shifts:
        point = await db.get(LocationPoint, shift.location_point_id)
        snapshot = await db.scalar(select(ShiftPayrollSnapshot).where(ShiftPayrollSnapshot.shift_id == shift.id).limit(1))
        closed_at = snapshot.closed_at if snapshot else (shift.closed_at or datetime.utcnow())
        is_auto_closed = bool(snapshot.is_auto_closed) if snapshot else False

        if snapshot is not None:
            await db.execute(delete(ShiftPayrollCategorySnapshot).where(ShiftPayrollCategorySnapshot.snapshot_id == snapshot.id))
            await db.delete(snapshot)
            await db.flush()

        computed = await _build_computed_shift(shift, db, force_refresh_metrics=force_refresh_metrics)
        new_snapshot = ShiftPayrollSnapshot(
            shift_id=shift.id,
            location_point_id=shift.location_point_id,
            employee_user_id=shift.employee_user_id,
            shift_date=shift.shift_date,
            settings_version_id=computed.settings.id,
            split_count=computed.split_count,
            share_ratio=computed.share_ratio,
            exit_amount=computed.exit_amount,
            bonus_threshold=computed.bonus_threshold,
            bonus_amount=computed.bonus_amount,
            other_rate_percent=computed.other_rate_percent,
            non_tobacco_net_sales_for_bonus=computed.bonus_base_sales_amount,
            gross_sales_amount=computed.gross_sales_amount,
            return_amount=computed.return_amount,
            net_sales_amount=computed.net_sales_amount,
            cost_amount=computed.cost_amount,
            gross_profit_amount=computed.gross_profit_amount,
            category_earnings_total=computed.category_earnings_total,
            employee_expense_amount=0.0,
            gross_salary_amount=computed.gross_salary_amount,
            net_salary_amount=computed.gross_salary_amount,
            is_auto_closed=is_auto_closed,
            closed_at=closed_at,
        )
        db.add(new_snapshot)
        await db.flush()

        for row in computed.categories:
            db.add(ShiftPayrollCategorySnapshot(
                snapshot_id=new_snapshot.id,
                category_id=row['category_id'],
                category_name=row['category_name'],
                rate_percent=row['rate_percent'],
                sales_amount=row['sales_amount'],
                return_amount=row['return_amount'],
                net_sales_amount=row['net_sales_amount'],
                earning_amount=row['earning_amount'],
                is_other_category=row['is_other_category'],
            ))

        await db.commit()
        updated += 1
        details.append({
            'shift_id': shift.id,
            'shift_date': shift.shift_date.isoformat(),
            'location': point.name if point else None,
            'gross_sales_amount': computed.gross_sales_amount,
            'return_amount': computed.return_amount,
            'cost_amount': computed.cost_amount,
            'gross_profit_amount': computed.gross_profit_amount,
        })
        if progress_callback is not None:
            callback_result = progress_callback(updated, len(shifts))
            if asyncio.iscoroutine(callback_result):
                await callback_result

    await db.commit()
    return {
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'location': _normalize_location(location) if location else None,
        'processed': len(shifts),
        'updated': updated,
        'details': details,
    }


async def list_payroll_audit_logs(location: str | None, db: AsyncSession, current_user: User, limit: int = 200) -> dict[str, Any]:
    query = select(PayrollAuditLog).order_by(PayrollAuditLog.created_at.desc()).limit(max(1, min(limit, 500)))
    if location:
        await ensure_user_can_access_location(current_user, location, db)
        point = await _get_location_point_by_name(location, db)
        query = query.where(PayrollAuditLog.location_point_id == point.id)
    elif current_user.role == 'admin':
        accessible_locations = set(await get_user_accessible_locations(current_user, db))
        point_ids = (
            await db.scalars(select(LocationPoint.id).where(LocationPoint.name.in_(accessible_locations)))
        ).all()
        query = query.where(PayrollAuditLog.location_point_id.in_(list(point_ids)))
    rows = (await db.scalars(query)).all()
    users = {user.id: user.full_name for user in (await db.scalars(select(User))).all()}
    locations = {point.id: point.name for point in (await db.scalars(select(LocationPoint))).all()}
    return {
        'logs': [
            {
                'id': row.id,
                'created_at': _datetime_to_str(row.created_at),
                'actor_user_id': row.actor_user_id,
                'actor_name': users.get(row.actor_user_id, 'Система' if row.actor_user_id is None else 'Пользователь'),
                'location': locations.get(row.location_point_id),
                'entity_type': row.entity_type,
                'entity_id': row.entity_id,
                'action_type': row.action_type,
                'details': json.loads(row.details_json or '{}'),
            }
            for row in rows
        ]
    }


async def export_employee_payroll_xlsx(location: str, date_from: date, date_to: date, db: AsyncSession, current_user: User, employee_user_id: int | None = None) -> tuple[str, bytes]:
    summary = await get_employee_payroll_summary(location, date_from, date_to, db, current_user, employee_user_id=employee_user_id)
    workbook = Workbook()
    ws = workbook.active
    ws.title = 'Зарплата'
    ws.append(['Сотрудник', summary['employee_name']])
    ws.append(['Точка', summary['location']])
    ws.append(['Период', f"{summary['date_from']} — {summary['date_to']}"])
    ws.append([])
    ws.append(['Дата', 'Выручка', 'Возвраты', 'Выручка после возвратов', 'Выход', 'Бонус', 'Категории', 'Итого'])
    for cell in ws[5]:
        cell.font = Font(bold=True)
    for day in summary['days']:
        ws.append([
            day['shift_date'],
            day['gross_sales_amount'],
            day['return_amount'],
            day['net_sales_amount'],
            day['exit_amount'],
            day['bonus_amount'],
            day['category_earnings_total'],
            day['gross_salary_amount'],
        ])
    ws.append([])
    ws.append(['Категория', 'Процент', 'Продажи', 'Возвраты', 'Чистая сумма', 'Начислено'])
    for cell in ws[7 + len(summary['days'])]:
        cell.font = Font(bold=True)
    for category in summary['categories']:
        ws.append([
            category['category_name'],
            category['rate_percent'],
            category['sales_amount'],
            category['return_amount'],
            category['net_sales_amount'],
            category['earning_amount'],
        ])
    ws.append([])
    ws.append(['Общий итог', summary['totals']['gross_salary_amount']])
    ws.append(['Расходы на сотрудника', summary['employee_expenses_total']])
    ws.append(['К выплате', summary['net_payout_amount']])
    payload = BytesIO()
    workbook.save(payload)
    safe_location = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in summary['location'])
    safe_employee = ''.join(ch if ch.isalnum() or ch in ('-', '_') else '_' for ch in summary['employee_name'])
    return f'payroll_{safe_location}_{safe_employee}_{summary["date_from"]}_{summary["date_to"]}.xlsx', payload.getvalue()
